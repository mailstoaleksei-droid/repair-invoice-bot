"""Excel generation module.

Creates a new .xlsx file per batch with invoice data
matching the target column format.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

log = logging.getLogger(__name__)

_HEADERS = [
    "Year", "Month", "Week", "Date Invoice", "Truck",
    "Total Price", "Invoice", "Seller", "Buyer", "Kategorie",
]


def generate(records: list[dict], output_dir: Path) -> Path | None:
    """Generate Excel file from validated records.

    Returns path to generated file, or None if no records.
    """
    if not records:
        return None

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Rechnungen_{timestamp}.xlsx"
    filepath = output_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Header row
    header_font = Font(bold=True)
    thin_border = Border(bottom=Side(style="thin"))
    for col, header in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border

    # Data rows
    for i, rec in enumerate(records, 2):
        ws.cell(row=i, column=1, value=rec.get("invoice_year"))
        ws.cell(row=i, column=2, value=rec.get("invoice_month"))
        ws.cell(row=i, column=3, value=rec.get("invoice_week"))
        ws.cell(row=i, column=4, value=rec.get("invoice_date", ""))
        ws.cell(row=i, column=5, value=rec.get("truck", ""))

        price_cell = ws.cell(row=i, column=6, value=rec.get("total_price"))
        price_cell.number_format = '#,##0.00'

        ws.cell(row=i, column=7, value=rec.get("invoice_nr", ""))
        ws.cell(row=i, column=8, value=rec.get("seller", ""))
        ws.cell(row=i, column=9, value=rec.get("buyer", ""))
        ws.cell(row=i, column=10, value=rec.get("kategorie", ""))

    # Column widths
    widths = [6, 6, 6, 12, 12, 12, 16, 30, 22, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Auto-filter
    ws.auto_filter.ref = f"A1:J{len(records) + 1}"

    wb.save(filepath)
    log.info("Excel saved: %s (%d rows)", filepath.name, len(records))
    return filepath
