"""
Система автоматической обработки PDF счетов
Версия 2.1 - С исправленной классификацией поставщиков
"""

import os
import shutil
import pdfplumber
import openpyxl
import re
from datetime import datetime
import time

# ===== НАСТРОЙКИ ПУТЕЙ =====
PDF_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG"
EXCEL_FILE = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG\Repair_2025.xlsx"
PROCESSED_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG"
MANUAL_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG\manual"
LOG_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\PDF_Processor\logs"

# ===== НАСТРОЙКИ TELEGRAM =====
TELEGRAM_ENABLED = False  # Установите True после настройки telegram_bot.py

if TELEGRAM_ENABLED:
    try:
        from telegram_bot import send_notification
    except ImportError:
        TELEGRAM_ENABLED = False
        print("⚠ Telegram уведомления недоступны")


# ===== ФУНКЦИИ ИЗВЛЕЧЕНИЯ ДАННЫХ =====

def extract_vital_projekt(text):
    """Извлечь данные из PDF счёта Vital Projekt"""
    data = {}
    
    invoice_match = re.search(r'Rechnungs-Nr\.:\s*(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    date_match = re.search(r'Hamburg,\s*den\s*(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
    else:
        date_match2 = re.search(r'Leistungsdatum/-Ort:(\d{2})-(\d{2})-(\d{4})', text)
        if date_match2:
            date_str = f"{date_match2.group(1)}.{date_match2.group(2)}.{date_match2.group(3)}"
        else:
            return None
    
    data['date'] = date_str
    
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    truck_match = re.search(r'Kennzeichen/Fahrer:\s*([A-Z]{2}\s*[A-Z]{2}\s*\d+)', text)
    if not truck_match:
        return None
    
    truck_raw = truck_match.group(1)
    truck_clean = re.sub(r'\s+', '', truck_raw)
    if len(truck_clean) >= 4:
        data['truck'] = f"{truck_clean[:2]}-{truck_clean[2:]}"
    else:
        data['truck'] = truck_clean
    
    table_match = re.search(
        r'^\s*1\s+(\d+)\s+Stk\.\s+[^\s]+\s+(.+?)\s+\d+%\s+([\d,]+)\s*€',
        text,
        re.MULTILINE
    )
    
    if table_match:
        data['amount'] = int(table_match.group(1))
        description = table_match.group(2).strip()
        data['name'] = description[:50] if len(description) > 50 else description
        price_str = table_match.group(3).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['amount'] = 1
        data['name'] = "Шины и услуги"
        data['price'] = 0.0
    
    summe_match = re.search(r'Summe\s+([\d,.]+)\s*€', text)
    if summe_match:
        total_str = summe_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Vital Projekt Inh.Vitalij Barth'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_autocompass_internal(text):
    """Извлечь данные из внутреннего счёта Auto Compass"""
    data = {}
    
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    date_match = re.search(r'Datum\s+(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
    else:
        date_match2 = re.search(r'Leistungsdatum:\s*(\d{2}\.\d{2}\.\d{4})', text)
        if date_match2:
            date_str = date_match2.group(1)
        else:
            return None
    
    data['date'] = date_str
    
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    truck_match = re.search(r'Kennzeichen\s+([A-Z]{2}-[A-Z]{2}\s*\d+)', text)
    if not truck_match:
        truck_match = re.search(r'([A-Z]{2}-[A-Z]{2}\s+\d+)', text)
    
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
    else:
        data['truck'] = ''
    
    description_match = re.search(
        r'Bezeichnung\s+Menge\s+E-Preis\s+Gesamt\s+(.+?)(?:\n\s*[A-Z0-9-]+\s+|\n\s*$)',
        text,
        re.DOTALL
    )
    if description_match:
        desc_text = description_match.group(1).strip()
        desc_lines = [line.strip() for line in desc_text.split('\n') if line.strip()]
        description = desc_lines[0] if desc_lines else desc_text
        data['name'] = description[:50] if len(description) > 50 else description
    else:
        data['name'] = 'Ремонтные работы'
    
    table_lines = text.split('\n')
    found_table = False
    for i, line in enumerate(table_lines):
        if 'E-Preis' in line and 'Gesamt' in line:
            found_table = True
            continue
        
        if found_table and line.strip():
            if not any(char.isdigit() for char in line):
                continue
            
            parts = line.split()
            numbers = []
            for part in parts:
                if re.match(r'^\d+[,.]?\d*$', part.replace(',', '.')):
                    numbers.append(part.replace(',', '.'))
            
            if len(numbers) >= 3:
                try:
                    data['amount'] = float(numbers[-3])
                    data['price'] = float(numbers[-2])
                except:
                    pass
                break
    
    if 'amount' not in data:
        data['amount'] = 1
        data['price'] = 0.0
    
    data['seller'] = 'Auto Compass GmbH'
    
    buyer_match = re.search(r'Firma\s+(.+?)(?:\n|Randersweide)', text)
    if buyer_match:
        buyer = buyer_match.group(1).strip()
        data['buyer'] = buyer
    else:
        data['buyer'] = 'Unknown'
    
    is_internal = (data['seller'] == data['buyer'])
    
    if is_internal:
        gesamt_match = re.search(r'Gesamt\s+([\d,.]+)\s*€', text)
        if not gesamt_match:
            lines = text.split('\n')
            for line in reversed(lines):
                if '€' in line:
                    numbers = re.findall(r'([\d,.]+)\s*€', line)
                    if numbers:
                        gesamt_match = re.search(r'([\d,.]+)\s*€', line)
                        break
        
        if gesamt_match:
            total_str = gesamt_match.group(1).replace('.', '').replace(',', '.')
            try:
                data['total_price'] = float(total_str)
            except:
                return None
        else:
            return None
        
        data['interne_rechnung'] = 'Interne Rechnung'
    else:
        netto_pattern1 = re.search(
            r'Lohn\s+Material\s+Fremdleistung\s+Netto\s+[\d,.]+\s*€\s+[\d,.]+\s*€\s+[\d,.]+\s*€\s+([\d,.]+)\s*€',
            text
        )
        netto_pattern2 = re.search(
            r'([\d,.]+)\s*€\s+([\d,.]+)\s*€\s+([\d,.]+)\s*€\s+([\d,.]+)\s*€\s+Gesamt',
            text
        )
        netto_pattern3 = re.search(r'Netto\s+([\d,.]+)\s*€', text)
        
        if netto_pattern1:
            total_str = netto_pattern1.group(1).replace('.', '').replace(',', '.')
            data['total_price'] = float(total_str)
        elif netto_pattern2:
            total_str = netto_pattern2.group(4).replace('.', '').replace(',', '.')
            data['total_price'] = float(total_str)
        elif netto_pattern3:
            total_str = netto_pattern3.group(1).replace('.', '').replace(',', '.')
            data['total_price'] = float(total_str)
        else:
            return None
        
        data['interne_rechnung'] = ''
    
    return data


# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =====

def identify_supplier(text):
    """
    Определить поставщика по содержимому PDF
    ИСПРАВЛЕНО: Приоритет точных совпадений в начале документа
    """
    # Берём первые 1000 символов (шапка документа)
    text_start = text[:1000].upper()
    text_upper = text.upper()
    
    # ПРИОРИТЕТ 1: Точные совпадения в начале документа
    if 'VITAL PROJEKT' in text_start:
        return 'Vital Projekt'
    elif 'K&L KFZ MEISTERBETRIEB' in text_start or 'K&L-KFZ' in text_start:
        return 'K&L'
    elif 'AUTO COMPASS' in text_start and ('KOPIE' in text or 'RANDERSWEIDE' in text):
        return 'Auto Compass (Internal)'
    elif 'DEKRA AUTOMOBIL' in text_start:
        return 'DEKRA'
    elif 'SCANIA HAMBURG' in text_start or 'SCANIA VERTRIEB' in text_start:
        return 'Scania'
    elif 'SCHÜTT GMBH' in text_start or 'W. SCHÜTT' in text_start:
        return 'Schütt'
    elif 'VOLVO GROUP' in text_start:
        return 'Volvo'
    elif 'FERRONORDIC' in text_start:
        return 'Ferronordic'
    elif 'QUICK REIFEN' in text_start:
        return 'Quick Reifendicount'
    elif 'SOTECS' in text_start:
        return 'Sotecs'
    elif 'EXPRESS' in text_start:
        return 'Express'
    
    # ПРИОРИТЕТ 2: Поиск во всём документе (только если не найдено в начале)
    elif 'SCANIA' in text_upper:
        return 'Scania'
    elif 'DEKRA' in text_upper:
        return 'DEKRA'
    elif 'FERRONORDIC' in text_upper:
        return 'Ferronordic'
    
    # MAN - ТОЛЬКО если действительно от MAN (не запчасти для MAN)
    elif 'MAN TRUCK' in text_upper or 'MAN SE' in text_upper:
        return 'MAN'
    
    return 'Unknown'


def check_invoice_exists(invoice_number, excel_path):
    """Проверить, существует ли уже счёт с таким номером в Excel"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f'I{row}'].value
            if cell_value and str(cell_value) == str(invoice_number):
                wb.close()
                return True
        
        wb.close()
        return False
    except Exception as e:
        print(f"  ⚠ Ошибка при проверке дубликата: {e}")
        return False


def write_to_excel(data, excel_path):
    """Добавить строку в Excel с датой обработки"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        row = ws.max_row + 1
        
        # Текущая дата обработки
        processing_date = datetime.now().strftime('%d.%m.%Y')
        
        ws[f'A{row}'] = data.get('month', '')
        ws[f'B{row}'] = data.get('week', '')
        ws[f'C{row}'] = data.get('truck', '')
        ws[f'D{row}'] = data.get('date', '')
        ws[f'E{row}'] = data.get('name', '')
        ws[f'F{row}'] = data.get('amount', '')
        ws[f'G{row}'] = data.get('price', '')
        ws[f'H{row}'] = data.get('total_price', '')
        ws[f'I{row}'] = data.get('invoice', '')
        ws[f'J{row}'] = data.get('seller', '')
        ws[f'K{row}'] = data.get('buyer', '')
        ws[f'L{row}'] = data.get('interne_rechnung', '')
        ws[f'M{row}'] = processing_date
        
        wb.save(excel_path)
        wb.close()
        return row
    except Exception as e:
        print(f"  ⚠ Ошибка записи в Excel: {e}")
        return None


def rename_and_move_file(source_path, destination_folder, prefix, max_attempts=3):
    """Переименовать файл с префиксом и переместить"""
    filename = os.path.basename(source_path)
    new_filename = f"{prefix}{filename}"
    destination = os.path.join(destination_folder, new_filename)
    
    for attempt in range(max_attempts):
        try:
            if os.path.exists(destination):
                base, ext = os.path.splitext(destination)
                counter = 1
                while os.path.exists(f"{base}({counter}){ext}"):
                    counter += 1
                destination = f"{base}({counter}){ext}"
                new_filename = os.path.basename(destination)
            
            shutil.move(source_path, destination)
            return True, new_filename, None
            
        except PermissionError:
            if attempt < max_attempts - 1:
                time.sleep(1)
                continue
            else:
                return False, new_filename, "Файл используется другим процессом"
        except Exception as e:
            return False, new_filename, str(e)
    
    return False, new_filename, "Не удалось переместить"


def log_processing(message, log_file='processing.log'):
    """Записать в лог файл"""
    try:
        os.makedirs(LOG_FOLDER, exist_ok=True)
        log_path = os.path.join(LOG_FOLDER, log_file)
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(f"[{timestamp}] {message}\n")
    except Exception as e:
        print(f"  ⚠ Ошибка записи в лог: {e}")


# ===== ГЛАВНАЯ ФУНКЦИЯ ОБРАБОТКИ =====

def process_all_pdfs():
    """Обработать все PDF файлы"""
    
    print("="*80)
    print("АВТОМАТИЧЕСКАЯ ОБРАБОТКА PDF СЧЕТОВ")
    print("="*80)
    print(f"Версия: 2.1 (исправленная классификация)")
    print(f"Дата запуска: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    print("="*80)
    
    processed_count = 0
    duplicate_count = 0
    error_count = 0
    
    log_processing("="*80)
    log_processing("НАЧАЛО ОБРАБОТКИ (v2.1)")
    log_processing("="*80)
    
    os.makedirs(MANUAL_FOLDER, exist_ok=True)
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)
    
    try:
        pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    except Exception as e:
        print(f"❌ Ошибка чтения папки: {e}")
        return {'total': 0, 'processed': 0, 'duplicates': 0, 'manual': 0}
    
    total_files = len(pdf_files)
    
    print(f"\n📁 Исходная папка: {PDF_FOLDER}")
    print(f"📄 Найдено файлов: {total_files}")
    
    if total_files == 0:
        print("\n✓ Нет файлов для обработки")
        return {'total': 0, 'processed': 0, 'duplicates': 0, 'manual': 0}
    
    if TELEGRAM_ENABLED:
        send_notification(f"🔄 <b>Начата обработка PDF</b>\n\n📁 Найдено файлов: {total_files}")
    
    print("\n" + "="*80)
    print("ОБРАБОТКА ФАЙЛОВ")
    print("="*80)
    
    for index, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        
        print(f"\n[{index}/{total_files}] {pdf_file}")
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                full_text = ''
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + '\n'
            
            if not full_text.strip():
                print(f"  ⚠ PDF пустой")
                success, new_name, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                if success:
                    print(f"  → manual/{new_name}")
                    log_processing(f"MANUAL (пустой): {pdf_file} -> {new_name}")
                error_count += 1
                continue
            
            supplier = identify_supplier(full_text)
            print(f"  Поставщик: {supplier}")
            
            data = None
            
            if supplier == 'Vital Projekt':
                data = extract_vital_projekt(full_text)
            elif supplier == 'Auto Compass (Internal)':
                data = extract_autocompass_internal(full_text)
            
            if not data:
                print(f"  ❌ Не удалось извлечь данные")
                success, new_name, error_msg = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                
                if success:
                    print(f"  → manual/{new_name}")
                    log_processing(f"MANUAL: {pdf_file} -> {new_name}")
                else:
                    print(f"  ⚠ Ошибка перемещения: {error_msg}")
                
                error_count += 1
                continue
            
            if check_invoice_exists(data['invoice'], EXCEL_FILE):
                print(f"  ⚠ ДУБЛИКАТ! Invoice {data['invoice']}")
                success, new_name, _ = rename_and_move_file(pdf_path, PROCESSED_FOLDER, 'checked_')
                
                if success:
                    print(f"  → {new_name}")
                    log_processing(f"ДУБЛИКАТ: {pdf_file} -> {new_name}")
                
                duplicate_count += 1
                continue
            
            row_num = write_to_excel(data, EXCEL_FILE)
            
            if row_num:
                print(f"  ✓ Excel строка {row_num}")
                print(f"    Invoice: {data['invoice']} | Truck: {data['truck']} | Total: {data['total_price']:.2f} €")
                
                success, new_name, error_msg = rename_and_move_file(pdf_path, PROCESSED_FOLDER, 'checked_')
                
                if success:
                    print(f"  → {new_name}")
                    log_processing(f"УСПЕХ: {pdf_file} -> {new_name} -> Строка {row_num}")
                    processed_count += 1
                else:
                    print(f"  ⚠ Данные записаны, файл не перемещён: {error_msg}")
                    processed_count += 1
            else:
                error_count += 1
                
        except Exception as e:
            print(f"  ✗ Ошибка: {e}")
            log_processing(f"ОШИБКА: {pdf_file} - {e}")
            
            try:
                success, new_name, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                if success:
                    print(f"  → manual/{new_name}")
            except:
                pass
            
            error_count += 1
    
    try:
        remaining_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
        remaining_count = len(remaining_files)
    except:
        remaining_count = 0
    
    print("\n" + "="*80)
    print("РЕЗУЛЬТАТЫ ОБРАБОТКИ")
    print("="*80)
    print(f"📄 Всего файлов:              {total_files}")
    print(f"✅ Успешно обработано:        {processed_count}")
    print(f"🔄 Дубликатов (пропущено):    {duplicate_count}")
    print(f"❌ Требуют ручной обработки:  {error_count}")
    print(f"📁 Осталось в исходной папке: {remaining_count}")
    print("="*80)
    
    if remaining_count == 0:
        print("✓ Исходная папка полностью очищена!")
    else:
        print(f"⚠ В исходной папке остались {remaining_count} файлов")
    
    print(f"\n📊 Проверьте результаты:")
    print(f"  Excel: ...\\Repair_2025.xlsx")
    print(f"  Обработанные: ...\\RG 2025 Ersatyteile RepRG")
    print(f"  Ручная обработка: ...\\EingangsRG\\manual")
    
    log_processing("="*80)
    log_processing(f"ИТОГО: Всего={total_files}, Успешно={processed_count}, Дубликаты={duplicate_count}, Manual={error_count}, Осталось={remaining_count}")
    log_processing("="*80)
    
    if TELEGRAM_ENABLED:
        notification_text = (
            f"✅ <b>Обработка завершена!</b>\n\n"
            f"📊 <b>Результаты:</b>\n"
            f"📄 Всего: {total_files}\n"
            f"✅ Обработано: {processed_count}\n"
            f"🔄 Дубликатов: {duplicate_count}\n"
            f"❌ Manual: {error_count}\n\n"
        )
        
        if remaining_count == 0:
            notification_text += "✓ Папка очищена!"
        else:
            notification_text += f"⚠ Осталось: {remaining_count} файлов"
        
        send_notification(notification_text)
    
    return {
        'total': total_files,
        'processed': processed_count,
        'duplicates': duplicate_count,
        'manual': error_count,
        'remaining': remaining_count
    }


if __name__ == "__main__":
    try:
        results = process_all_pdfs()
        exit_code = 0 if results['remaining'] == 0 else 1
        exit(exit_code)
        
    except KeyboardInterrupt:
        print("\n\n⏹ Обработка прервана")
        log_processing("ПРЕРВАНО")
        exit(2)
    except Exception as e:
        print(f"\n\n❌ Критическая ошибка: {e}")
        log_processing(f"КРИТИЧЕСКАЯ ОШИБКА: {e}")
        exit(3)