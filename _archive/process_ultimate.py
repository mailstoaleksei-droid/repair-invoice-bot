"""
ULTIMATE PDF PROCESSOR v3.0
Система автоматической обработки PDF счетов
Цель: 92% автоматизации (278 из 302 файлов)

Возможности:
- 19 функций извлечения данных
- OCR для сканов
- Telegram уведомления
- Полное логирование
"""

import os
import shutil
import pdfplumber
import openpyxl
import re
from datetime import datetime
import time
from collections import defaultdict

# ===== НАСТРОЙКИ ПУТЕЙ =====
PDF_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG"
EXCEL_FILE = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG\Repair_2025.xlsx"
PROCESSED_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG"
MANUAL_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG\manual"
LOG_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\PDF_Processor\logs"

# ===== НАСТРОЙКИ =====
TELEGRAM_ENABLED = False
USE_OCR = True  # Использовать OCR для сканов

if TELEGRAM_ENABLED:
    try:
        from telegram_bot import send_notification
    except ImportError:
        TELEGRAM_ENABLED = False

# ===== OCR МОДУЛЬ =====
if USE_OCR:
    try:
        import pytesseract
        from pdf2image import convert_from_path
        from PIL import Image
        OCR_AVAILABLE = True
        # Путь к Tesseract (укажите свой путь после установки)
        pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    except ImportError:
        OCR_AVAILABLE = False
        print("⚠ OCR недоступен. Установите: pip install pytesseract pdf2image pillow")
else:
    OCR_AVAILABLE = False


# ===== ФУНКЦИИ ИЗВЛЕЧЕНИЯ ДАННЫХ =====

def extract_vital_projekt(text):
    """Vital Projekt - шины и услуги"""
    data = {}
    
    invoice_match = re.search(r'Rechnungs-Nr\.:\s*(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    date_match = re.search(r'Hamburg,\s*den\s*(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
    else:
        date_match2 = re.search(r'Leistungsdatum/-Ort:(\d{2})-(\d{2})-(\d{4})', text)
        if date_match2:
            date_str = f"{date_match2.group(1)}.{date_match2.group(2)}.{date_match2.group(3)}"
        else:
            return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    truck_match = re.search(r'Kennzeichen/Fahrer:\s*([A-Z]{2}\s*[A-Z]{2}\s*\d+)', text)
    if not truck_match:
        return None
    
    truck_raw = truck_match.group(1)
    truck_clean = re.sub(r'\s+', '', truck_raw)
    if len(truck_clean) >= 4:
        data['truck'] = f"{truck_clean[:2]}-{truck_clean[2:]}"
    else:
        data['truck'] = truck_clean
    
    table_match = re.search(
        r'^\s*1\s+(\d+)\s+Stk\.\s+[^\s]+\s+(.+?)\s+\d+%\s+([\d,]+)\s*€',
        text, re.MULTILINE
    )
    
    if table_match:
        data['amount'] = int(table_match.group(1))
        description = table_match.group(2).strip()
        data['name'] = description[:50] if len(description) > 50 else description
        price_str = table_match.group(3).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['amount'] = 1
        data['name'] = "Шины и услуги"
        data['price'] = 0.0
    
    summe_match = re.search(r'Summe\s+([\d,.]+)\s*€', text)
    if summe_match:
        total_str = summe_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Vital Projekt Inh.Vitalij Barth'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_autocompass_internal(text):
    """Auto Compass Internal - внутренние счета"""
    data = {}
    
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        # Альтернативный поиск для Scania Internal
        invoice_match = re.search(r'Interne Rechnung\s+(\d+)', text)
        if not invoice_match:
            return None
    data['invoice'] = invoice_match.group(1)
    
    date_match = re.search(r'Datum\s+(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
    else:
        date_match2 = re.search(r'Leistungsdatum:\s*(\d{2}\.\d{2}\.\d{4})', text)
        if date_match2:
            date_str = date_match2.group(1)
        else:
            return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    truck_match = re.search(r'Kennzeichen\s+([A-Z]{2}-[A-Z]{2}\s*\d+)', text)
    if not truck_match:
        truck_match = re.search(r'([A-Z]{2}-[A-Z]{2}\s+\d+)', text)
    
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
    else:
        data['truck'] = ''
    
    description_match = re.search(
        r'Bezeichnung\s+Menge\s+E-Preis\s+Gesamt\s+(.+?)(?:\n\s*[A-Z0-9-]+\s+|\n\s*$)',
        text, re.DOTALL
    )
    if description_match:
        desc_text = description_match.group(1).strip()
        desc_lines = [line.strip() for line in desc_text.split('\n') if line.strip()]
        description = desc_lines[0] if desc_lines else desc_text
        data['name'] = description[:50] if len(description) > 50 else description
    else:
        data['name'] = 'Ремонтные работы'
    
    table_lines = text.split('\n')
    found_table = False
    for line in table_lines:
        if 'E-Preis' in line and 'Gesamt' in line:
            found_table = True
            continue
        
        if found_table and line.strip():
            if not any(char.isdigit() for char in line):
                continue
            
            parts = line.split()
            numbers = []
            for part in parts:
                if re.match(r'^\d+[,.]?\d*$', part.replace(',', '.')):
                    numbers.append(part.replace(',', '.'))
            
            if len(numbers) >= 3:
                try:
                    data['amount'] = float(numbers[-3])
                    data['price'] = float(numbers[-2])
                except:
                    pass
                break
    
    if 'amount' not in data:
        data['amount'] = 1
        data['price'] = 0.0
    
    data['seller'] = 'Auto Compass GmbH'
    
    buyer_match = re.search(r'Firma\s+(.+?)(?:\n|Randersweide)', text)
    if buyer_match:
        buyer = buyer_match.group(1).strip()
        data['buyer'] = buyer
    else:
        data['buyer'] = 'Unknown'
    
    is_internal = (data['seller'] == data['buyer'])
    
    if is_internal:
        gesamt_match = re.search(r'Gesamt\s+([\d,.]+)\s*€', text)
        if not gesamt_match:
            lines = text.split('\n')
            for line in reversed(lines):
                if '€' in line:
                    numbers = re.findall(r'([\d,.]+)\s*€', line)
                    if numbers:
                        gesamt_match = re.search(r'([\d,.]+)\s*€', line)
                        break
        
        if gesamt_match:
            total_str = gesamt_match.group(1).replace('.', '').replace(',', '.')
            try:
                data['total_price'] = float(total_str)
            except:
                return None
        else:
            return None
        
        data['interne_rechnung'] = 'Interne Rechnung'
    else:
        netto_pattern1 = re.search(
            r'Lohn\s+Material\s+Fremdleistung\s+Netto\s+[\d,.]+\s*€\s+[\d,.]+\s*€\s+[\d,.]+\s*€\s+([\d,.]+)\s*€',
            text
        )
        netto_pattern2 = re.search(
            r'([\d,.]+)\s*€\s+([\d,.]+)\s*€\s+([\d,.]+)\s*€\s+([\d,.]+)\s*€\s+Gesamt',
            text
        )
        netto_pattern3 = re.search(r'Netto\s+([\d,.]+)\s*€', text)
        
        if netto_pattern1:
            total_str = netto_pattern1.group(1).replace('.', '').replace(',', '.')
            data['total_price'] = float(total_str)
        elif netto_pattern2:
            total_str = netto_pattern2.group(4).replace('.', '').replace(',', '.')
            data['total_price'] = float(total_str)
        elif netto_pattern3:
            total_str = netto_pattern3.group(1).replace('.', '').replace(',', '.')
            data['total_price'] = float(total_str)
        else:
            return None
        
        data['interne_rechnung'] = ''
    
    return data


def extract_scania_external(text):
    """Scania External - внешние счета от Scania (не внутренние AC)"""
    data = {}
    
    # Проверка на внешний счет Scania
    if '#splminfo' not in text.lower():
        return None
    
    # Номер счёта - ищем SCHWL
    invoice_match = re.search(r'SCHWL(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = f"SCHWL{invoice_match.group(1)}"
    
    # Дата - различные форматы
    date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{2,4})', text)
    if date_match:
        day = date_match.group(1)
        month = date_match.group(2)
        year = date_match.group(3)
        if len(year) == 2:
            year = f"20{year}"
        date_str = f"{day}.{month}.{year}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина
    truck_match = re.search(r'GR-OO\s*(\d+)', text)
    if truck_match:
        data['truck'] = f"GR-OO{truck_match.group(1)}"
    else:
        truck_match2 = re.search(r'ATML\.KENNZ:\s*([A-Z]{2}-[A-Z]{2}\s*\d+)', text)
        if truck_match2:
            data['truck'] = re.sub(r'\s+', '', truck_match2.group(1))
        else:
            data['truck'] = ''
    
    # Описание
    lines = text.split('\n')
    for i, line in enumerate(lines):
        if 'HU durchführen' in line or 'durchgeführt' in line:
            data['name'] = line.strip()[:50]
            break
    
    if 'name' not in data:
        data['name'] = 'Ремонтные работы Scania'
    
    # Количество и цена
    data['amount'] = 1
    data['price'] = 0.0
    
    # Сумма - ищем различные паттерны
    total_match = re.search(r'(\d+,\d+)\s*€', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Scania Deutschland GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_ferronordic(text):
    """Ferronordic - сервис Volvo"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnung:\s*RE(\d+)-(\d+)', text)
    if invoice_match:
        data['invoice'] = f"RE{invoice_match.group(1)}-{invoice_match.group(2)}"
    else:
        return None
    
    # Дата
    date_match = re.search(r'Vom:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина
    truck_match = re.search(r'Kennzeichen:\s*DE\s*FN\s*(\d+)', text)
    if truck_match:
        data['truck'] = f"DE-FN{truck_match.group(1)}"
    else:
        truck_match2 = re.search(r'DE\s*FN\s*(\d+)', text)
        if truck_match2:
            data['truck'] = f"DE-FN{truck_match2.group(1)}"
        else:
            data['truck'] = ''
    
    # Описание
    desc_match = re.search(r'1\s+Monatstarif[^\n]+', text)
    if desc_match:
        data['name'] = desc_match.group(0)[:50]
    else:
        data['name'] = 'Wartung/Service Volvo'
    
    # Количество и цена
    amount_match = re.search(r'1\s+Monatstarif[^\d]+(\d+,\d+)\s*€', text)
    if amount_match:
        data['amount'] = 1
        price_str = amount_match.group(1).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['amount'] = 1
        data['price'] = 0.0
    
    # Общая сумма (Endsumme с НДС)
    total_match = re.search(r'Endsumme\s+([\d,.]+)\s*€', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Ferronordic Deutschland GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_hns(text):
    """HNS Nutzfahrzeuge Service"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnung\s*:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина
    truck_match = re.search(r'Fahrzeug:\s*GR-OO\s*(\d+)', text)
    if truck_match:
        data['truck'] = f"GR-OO{truck_match.group(1)}"
    else:
        truck_match2 = re.search(r'GR-OO\s*(\d+)', text)
        if truck_match2:
            data['truck'] = f"GR-OO{truck_match2.group(1)}"
        else:
            data['truck'] = ''
    
    # Описание - берём первую работу
    desc_match = re.search(r'HU\s+Begleitung|HU\s+Prüfung', text)
    if desc_match:
        data['name'] = desc_match.group(0)
    else:
        data['name'] = 'Werkstattleistungen'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма (Endbetrag)
    total_match = re.search(r'Endbetrag\s*:\s*([\d,.]+)\s*€', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'HNS Nutzfahrzeuge Service GmbH'
    data['buyer'] = 'Groo GmbH' if 'Groo GmbH' in text else 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_tip(text):
    """TIP Trailer Services - аренда прицепов"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnungsnr\.\s*U71/(\d+)', text)
    if invoice_match:
        data['invoice'] = f"U71/{invoice_match.group(1)}"
    else:
        return None
    
    # Дата
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})/(\d{2})/(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина (прицеп)
    truck_match = re.search(r'Kennzeichen:\s*([A-Z0-9]+)', text)
    if truck_match:
        data['truck'] = truck_match.group(1)
    else:
        # Альтернативный поиск
        truck_match2 = re.search(r'Flotten-Nr\.\s*(\d+)', text)
        if truck_match2:
            data['truck'] = truck_match2.group(1)
        else:
            data['truck'] = ''
    
    # Описание
    data['name'] = 'Аренда прицепа'
    
    data['amount'] = 1
    
    # Цена за месяц
    price_match = re.search(r'(\d+,\d+)\s+EUR', text)
    if price_match:
        price_str = price_match.group(1).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['price'] = 0.0
    
    # Общая сумма
    total_match = re.search(r'Gesamtbetrag\s+EUR\s+([\d,.]+)', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'TIP Trailer Services Germany GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_euromaster(text):
    """Euromaster - шины и обслуживание"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'Datum\s*:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина
    truck_match = re.search(r'KFZ-KENNZEICHEN:\s*([A-Z]{2}-[A-Z0-9]+)', text)
    if truck_match:
        data['truck'] = truck_match.group(1)
    else:
        data['truck'] = ''
    
    # Описание - первая позиция
    desc_match = re.search(r'EUROMASTER\s+\d+\s+[^\n]+', text)
    if desc_match:
        data['name'] = desc_match.group(0)[:50]
    else:
        data['name'] = 'Шины и обслуживание'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма (Nettowert или Bruttowert)
    total_match = re.search(r'Nettowert\s+Bruttowert\s+([\d,.]+)\s+([\d,.]+)', text)
    if total_match:
        netto_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(netto_str)
    else:
        # Альтернативный поиск
        total_match2 = re.search(r'([\d,.]+)\s+EUR\s+zum', text)
        if total_match2:
            total_str = total_match2.group(1).replace('.', '').replace(',', '.')
            data['total_price'] = float(total_str)
        else:
            return None
    
    data['seller'] = 'Euromaster GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data

# ===== ПРОДОЛЖЕНИЕ ФУНКЦИЙ ИЗВЛЕЧЕНИЯ =====

def extract_man(text):
    """MAN Truck & Bus Deutschland"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnungsnummer:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина - не всегда есть
    data['truck'] = ''
    
    # Описание
    desc_match = re.search(r'Job\s+\d+:\s+([^\n]+)', text)
    if desc_match:
        data['name'] = desc_match.group(1)[:50]
    else:
        data['name'] = 'Ремонт MAN'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма - различные паттерны
    total_match = re.search(r'NETTO\s+([\d,.]+)\s+EUR', text, re.IGNORECASE)
    if not total_match:
        total_match = re.search(r'Nettobetrag:\s+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'MAN Truck & Bus Deutschland GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_dekra(text):
    """DEKRA Automobil GmbH"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnung\s+Nr\.\s*:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина
    truck_match = re.search(r'Kennzeichen[^\n]*:\s*([A-Z]{2}\s*[A-Z0-9]+\s*\d+)', text)
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
        # Форматировать как GR-OO1514
        if len(data['truck']) >= 4:
            data['truck'] = f"{data['truck'][:2]}-{data['truck'][2:]}"
    else:
        data['truck'] = ''
    
    # Описание
    data['name'] = 'Hauptuntersuchung'
    
    data['amount'] = 1
    
    # Цена
    price_match = re.search(r'Nettopreis\s+[\d,]+\s+[\d,]+\s+([\d,]+)', text)
    if price_match:
        price_str = price_match.group(1).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['price'] = 0.0
    
    # Общая сумма
    total_match = re.search(r'Nettobetrag\s+([\d,.]+)', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'DEKRA Automobil GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_schutt(text):
    """W. Schütt GmbH"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnung\s*:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина
    truck_match = re.search(r'Kennzeichen\s*:\s*([A-Z]{2}-[A-Z]{2}\s*\d+)', text)
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
    else:
        data['truck'] = ''
    
    # Описание
    data['name'] = 'Ремонтные работы'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма
    total_match = re.search(r'Netto\s+([\d,.]+)\s*€', text)
    if not total_match:
        total_match = re.search(r'Gesamt\s+([\d,.]+)\s*€', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'W. Schütt GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_volvo(text):
    """Volvo Group Trucks Service"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnungs-Nr\.\.\s*:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'Auftr\.-Datum\.\.\s*:\s*(\d{2})\.(\d{2})\.(\d{2})', text)
    if date_match:
        year = f"20{date_match.group(3)}"
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{year}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина
    truck_match = re.search(r'Kennzeichen\.\.\s*:\s*([A-Z]{2}-[A-Z0-9]+)', text)
    if truck_match:
        data['truck'] = truck_match.group(1)
    else:
        data['truck'] = ''
    
    # Описание
    desc_match = re.search(r'PFX\s+Ersatzteilnummer\s+[^\n]+', text)
    if desc_match:
        data['name'] = 'Запчасти Volvo'
    else:
        data['name'] = 'Ремонт Volvo'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма
    total_match = re.search(r'Gesamt\s+EUR\s+([\d,.]+)', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Volvo Group Trucks Service Nord GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_sotecs(text):
    """Sotecs GmbH"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnungsnummer:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    data['truck'] = ''
    
    # Описание
    desc_match = re.search(r'Einbau von\s+([^\n]+)', text)
    if desc_match:
        data['name'] = desc_match.group(1)[:50]
    else:
        data['name'] = 'Установка оборудования'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма
    total_match = re.search(r'Total\s+\(netto\)\s+([\d,.]+)\s+EUR', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Sotecs GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_express(text):
    """Express Service"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'Datum\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина
    truck_match = re.search(r'NGZ\s+(\d+)', text)
    if truck_match:
        data['truck'] = f"NGZ{truck_match.group(1)}"
    else:
        data['truck'] = ''
    
    # Описание
    data['name'] = 'Wartung und Filter'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма
    total_match = re.search(r'Gesamtbetrag\s+([\d,.]+)\s*€', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Express Service GmbH'
    data['buyer'] = 'UAB Groo Transport'
    data['interne_rechnung'] = ''
    
    return data


def extract_kl(text):
    """K&L Kfz Meisterbetrieb"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnungsnummer:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    data['truck'] = ''
    
    # Описание - из таблицы
    desc_match = re.search(r'MEHRFACHKUPPLUNG|GETRIEBE', text)
    if desc_match:
        data['name'] = desc_match.group(0)
    else:
        data['name'] = 'Запчасти и ремонт'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма
    total_match = re.search(r'Gesamtsumme netto\s+([\d,.]+)\s*EUR', text)
    if not total_match:
        total_match = re.search(r'Total\s+\(netto\)\s+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'K&L Kfz Meisterbetrieb GmbH'
    data['buyer'] = 'GROO GmbH'
    data['interne_rechnung'] = ''
    
    return data


def extract_quick(text):
    """Quick Reifendicount"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnung\s+Nr\.\s*:\s*([A-Z0-9]+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{2})', text)
    if date_match:
        year = f"20{date_match.group(3)}"
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{year}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина
    truck_match = re.search(r'GR-(\d+)', text)
    if truck_match:
        data['truck'] = f"GR-{truck_match.group(1)}"
    else:
        data['truck'] = ''
    
    # Описание
    data['name'] = 'Шины'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма
    total_match = re.search(r'Gesamtbetrag\s+([\d,.]+)\s+EUR', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Quick Reifendicount'
    data['buyer'] = 'Andreas Groo Handel & Transporte'
    data['interne_rechnung'] = ''
    
    return data


def extract_tankpool24(text):
    """Tankpool24 - топливные карты"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'(\d{7})', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата - обычно в начале
    date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    data['truck'] = ''
    data['name'] = 'Топливо'
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма
    total_match = re.search(r'([\d,.]+)\s+EUR', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Tankpool24 International GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data


# ===== КЛАССИФИКАТОР =====

def identify_supplier(text):
    """
    Улучшенный классификатор поставщиков
    Возвращает название поставщика
    """
    text_start = text[:1500].upper()
    text_upper = text.upper()
    
    # ПРИОРИТЕТ 1: Точные совпадения в начале документа
    
    # Vital Projekt
    if 'VITAL PROJEKT' in text_start:
        return 'Vital Projekt'
    
    # K&L
    elif 'K&L KFZ MEISTERBETRIEB' in text_start or 'K&L-KFZ' in text_start:
        return 'K&L'
    
    # Auto Compass Internal
    elif 'AUTO COMPASS' in text_start and ('KOPIE' in text or 'RANDERSWEIDE' in text):
        return 'Auto Compass (Internal)'
    
    # Scania External (внешние счета)
    elif '#SPLMINFO' in text_start and 'SCANIA' in text_start:
        return 'Scania External'
    
    # Ferronordic
    elif 'FERRONORDIC' in text_start:
        return 'Ferronordic'
    
    # HNS
    elif 'HNS NUTZFAHRZEUGE' in text_start or 'HNS SERVICE' in text_start:
        return 'HNS'
    
    # DEKRA
    elif 'DEKRA AUTOMOBIL' in text_start:
        return 'DEKRA'
    
    # Schütt
    elif 'SCHÜTT GMBH' in text_start or 'W. SCHÜTT' in text_start:
        return 'Schütt'
    
    # Volvo
    elif 'VOLVO GROUP TRUCKS' in text_start:
        return 'Volvo'
    
    # TIP Trailer
    elif 'TIP TRAILER SERVICES' in text_start or 'TRAILER SERVICES GERMANY' in text_start:
        return 'TIP'
    
    # Euromaster
    elif 'EUROMASTER GMBH' in text_start:
        return 'Euromaster'
    
    # Quick
    elif 'QUICK REIFEN' in text_start or 'REIFENDISCOUNT' in text_start:
        return 'Quick'
    
    # Sotecs
    elif 'SOTECS GMBH' in text_start:
        return 'Sotecs'
    
    # Express
    elif 'EXPRESS' in text_start and 'UAB GROO' in text:
        return 'Express'
    
    # Tankpool24
    elif 'TANKPOOL' in text_start:
        return 'Tankpool24'
    
    # ПРИОРИТЕТ 2: Поиск во всём документе
    
    # MAN - только если действительно от MAN
    elif 'MAN TRUCK & BUS DEUTSCHLAND' in text_upper:
        return 'MAN'
    
    return 'Unknown'


# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =====

def check_invoice_exists(invoice_number, excel_path):
    """Проверить дубликат счёта"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f'I{row}'].value
            if cell_value and str(cell_value) == str(invoice_number):
                wb.close()
                return True
        
        wb.close()
        return False
    except:
        return False


def write_to_excel(data, excel_path):
    """Записать в Excel с датой обработки"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        row = ws.max_row + 1
        processing_date = datetime.now().strftime('%d.%m.%Y')
        
        ws[f'A{row}'] = data.get('month', '')
        ws[f'B{row}'] = data.get('week', '')
        ws[f'C{row}'] = data.get('truck', '')
        ws[f'D{row}'] = data.get('date', '')
        ws[f'E{row}'] = data.get('name', '')
        ws[f'F{row}'] = data.get('amount', '')
        ws[f'G{row}'] = data.get('price', '')
        ws[f'H{row}'] = data.get('total_price', '')
        ws[f'I{row}'] = data.get('invoice', '')
        ws[f'J{row}'] = data.get('seller', '')
        ws[f'K{row}'] = data.get('buyer', '')
        ws[f'L{row}'] = data.get('interne_rechnung', '')
        ws[f'M{row}'] = processing_date
        
        wb.save(excel_path)
        wb.close()
        return row
    except Exception as e:
        print(f"  ⚠ Ошибка записи: {e}")
        return None


def rename_and_move_file(source_path, destination_folder, prefix, max_attempts=3):
    """Переименовать с префиксом и переместить"""
    filename = os.path.basename(source_path)
    new_filename = f"{prefix}{filename}"
    destination = os.path.join(destination_folder, new_filename)
    
    for attempt in range(max_attempts):
        try:
            if os.path.exists(destination):
                base, ext = os.path.splitext(destination)
                counter = 1
                while os.path.exists(f"{base}({counter}){ext}"):
                    counter += 1
                destination = f"{base}({counter}){ext}"
                new_filename = os.path.basename(destination)
            
            shutil.move(source_path, destination)
            return True, new_filename, None
            
        except PermissionError:
            if attempt < max_attempts - 1:
                time.sleep(1)
            else:
                return False, new_filename, "Файл занят"
        except Exception as e:
            return False, new_filename, str(e)
    
    return False, new_filename, "Не удалось переместить"


def log_processing(message, log_file='processing.log'):
    """Логирование"""
    try:
        os.makedirs(LOG_FOLDER, exist_ok=True)
        log_path = os.path.join(LOG_FOLDER, log_file)
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(f"[{timestamp}] {message}\n")
    except:
        pass


def extract_text_with_ocr(pdf_path):
    """Извлечь текст из PDF с помощью OCR"""
    if not OCR_AVAILABLE:
        return None
    
    try:
        # Конвертировать PDF в изображения
        images = convert_from_path(pdf_path, dpi=300)
        
        full_text = ''
        for image in images:
            # OCR для каждой страницы
            text = pytesseract.image_to_string(image, lang='deu+eng')
            full_text += text + '\n'
        
        return full_text
    except Exception as e:
        print(f"  ⚠ Ошибка OCR: {e}")
        return None
    
# ===== РОУТЕР ОБРАБОТКИ =====

def extract_data_by_supplier(text, supplier):
    """
    Роутер: вызвать правильную функцию извлечения
    """
    extractors = {
        'Vital Projekt': extract_vital_projekt,
        'Auto Compass (Internal)': extract_autocompass_internal,
        'Scania External': extract_scania_external,
        'Ferronordic': extract_ferronordic,
        'HNS': extract_hns,
        'TIP': extract_tip,
        'Euromaster': extract_euromaster,
        'MAN': extract_man,
        'DEKRA': extract_dekra,
        'Schütt': extract_schutt,
        'Volvo': extract_volvo,
        'Sotecs': extract_sotecs,
        'Express': extract_express,
        'K&L': extract_kl,
        'Quick': extract_quick,
        'Tankpool24': extract_tankpool24,
    }
    
    extractor = extractors.get(supplier)
    if extractor:
        return extractor(text)
    
    return None


# ===== ГЛАВНАЯ ФУНКЦИЯ ОБРАБОТКИ =====

def process_all_pdfs():
    """
    ГЛАВНАЯ ФУНКЦИЯ
    Обработать все PDF в папке с максимальной автоматизацией
    """
    
    print("="*80)
    print("🚀 ULTIMATE PDF PROCESSOR v3.0")
    print("="*80)
    print(f"Цель: Максимальная автоматизация (92%+)")
    print(f"Дата запуска: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    print("="*80)
    
    # Счётчики
    processed_count = 0
    duplicate_count = 0
    error_count = 0
    ocr_count = 0
    
    # Статистика по поставщикам
    supplier_stats = defaultdict(int)
    
    log_processing("="*80)
    log_processing("НАЧАЛО ОБРАБОТКИ - ULTIMATE v3.0")
    log_processing("="*80)
    
    # Создать папки
    os.makedirs(MANUAL_FOLDER, exist_ok=True)
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)
    
    # Получить список PDF
    try:
        pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    except Exception as e:
        print(f"❌ Ошибка чтения папки: {e}")
        return {'total': 0, 'processed': 0, 'duplicates': 0, 'manual': 0}
    
    total_files = len(pdf_files)
    
    print(f"\n📁 Папка: {PDF_FOLDER}")
    print(f"📄 Найдено файлов: {total_files}")
    
    if total_files == 0:
        print("\n✓ Нет файлов для обработки")
        return {'total': 0, 'processed': 0, 'duplicates': 0, 'manual': 0}
    
    # Telegram уведомление
    if TELEGRAM_ENABLED:
        send_notification(
            f"🚀 <b>ULTIMATE Processor запущен</b>\n\n"
            f"📁 Файлов: {total_files}\n"
            f"🎯 Цель: 92% автоматизации"
        )
    
    print("\n" + "="*80)
    print("ОБРАБОТКА ФАЙЛОВ")
    print("="*80)
    
    # Обработка каждого файла
    for index, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        
        print(f"\n[{index}/{total_files}] {pdf_file}")
        
        try:
            # ===== ШАГ 1: Извлечь текст =====
            full_text = ''
            used_ocr = False
            
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + '\n'
            
            # Если текста мало - попробовать OCR
            if len(full_text.strip()) < 100 and OCR_AVAILABLE:
                print(f"  📷 Попытка OCR (мало текста)...")
                ocr_text = extract_text_with_ocr(pdf_path)
                if ocr_text and len(ocr_text.strip()) > 100:
                    full_text = ocr_text
                    used_ocr = True
                    ocr_count += 1
                    print(f"  ✓ OCR успешно")
            
            if not full_text.strip():
                print(f"  ❌ Пустой PDF (даже после OCR)")
                success, new_name, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                if success:
                    print(f"  → manual/{new_name}")
                    log_processing(f"MANUAL (пустой): {pdf_file}")
                error_count += 1
                continue
            
            # ===== ШАГ 2: Классифицировать =====
            supplier = identify_supplier(full_text)
            print(f"  Поставщик: {supplier}")
            supplier_stats[supplier] += 1
            
            # ===== ШАГ 3: Извлечь данные =====
            data = extract_data_by_supplier(full_text, supplier)
            
            if not data:
                print(f"  ❌ Не удалось извлечь данные")
                success, new_name, error_msg = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'manual_'
                )
                
                if success:
                    print(f"  → manual/{new_name}")
                    log_processing(f"MANUAL ({supplier}): {pdf_file} -> {new_name}")
                else:
                    print(f"  ⚠ Ошибка перемещения: {error_msg}")
                
                error_count += 1
                continue
            
            # ===== ШАГ 4: Проверить дубликат =====
            if check_invoice_exists(data['invoice'], EXCEL_FILE):
                print(f"  ⚠ ДУБЛИКАТ! Invoice {data['invoice']}")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'checked_'
                )
                
                if success:
                    print(f"  → {new_name}")
                    log_processing(f"ДУБЛИКАТ ({supplier}): {pdf_file} -> {new_name}")
                
                duplicate_count += 1
                continue
            
            # ===== ШАГ 5: Записать в Excel =====
            row_num = write_to_excel(data, EXCEL_FILE)
            
            if row_num:
                print(f"  ✓ Excel строка {row_num}")
                print(f"    Invoice: {data['invoice']} | Truck: {data['truck']} | Total: {data['total_price']:.2f} €")
                
                if used_ocr:
                    print(f"    🔍 Использован OCR")
                
                # ===== ШАГ 6: Переместить =====
                success, new_name, error_msg = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'checked_'
                )
                
                if success:
                    print(f"  → {new_name}")
                    log_processing(f"УСПЕХ ({supplier}): {pdf_file} -> {new_name} -> Строка {row_num}")
                    processed_count += 1
                else:
                    print(f"  ⚠ Записано, но не перемещён: {error_msg}")
                    log_processing(f"ЧАСТИЧНО ({supplier}): {pdf_file} -> Строка {row_num}")
                    processed_count += 1
            else:
                error_count += 1
                
        except Exception as e:
            print(f"  ✗ Критическая ошибка: {e}")
            log_processing(f"ОШИБКА: {pdf_file} - {e}")
            
            try:
                success, new_name, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                if success:
                    print(f"  → manual/{new_name}")
            except:
                pass
            
            error_count += 1
    
    # ===== ПРОВЕРКА ОСТАТКОВ =====
    try:
        remaining_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
        remaining_count = len(remaining_files)
    except:
        remaining_count = 0
    
    # ===== ИТОГИ =====
    print("\n" + "="*80)
    print("📊 РЕЗУЛЬТАТЫ ОБРАБОТКИ")
    print("="*80)
    print(f"📄 Всего файлов:              {total_files}")
    print(f"✅ Успешно обработано:        {processed_count}")
    print(f"🔄 Дубликатов (пропущено):    {duplicate_count}")
    print(f"❌ Требуют ручной обработки:  {error_count}")
    print(f"🔍 Обработано через OCR:      {ocr_count}")
    print(f"📁 Осталось в папке:          {remaining_count}")
    
    # Процент автоматизации
    automation_percent = (processed_count / total_files * 100) if total_files > 0 else 0
    print(f"\n🎯 АВТОМАТИЗАЦИЯ: {automation_percent:.1f}%")
    
    print("\n" + "="*80)
    print("📈 СТАТИСТИКА ПО ПОСТАВЩИКАМ")
    print("="*80)
    
    for supplier in sorted(supplier_stats.keys()):
        count = supplier_stats[supplier]
        percent = (count / total_files * 100) if total_files > 0 else 0
        print(f"{supplier:30} : {count:3} файлов ({percent:5.1f}%)")
    
    print("\n" + "="*80)
    
    if remaining_count == 0:
        print("✅ Исходная папка ПОЛНОСТЬЮ очищена!")
    else:
        print(f"⚠ В исходной папке остались {remaining_count} файлов")
        print("  (возможно, файлы заняты другими программами)")
    
    print("\n📂 Результаты:")
    print(f"  ✅ Обработанные: {PROCESSED_FOLDER}")
    print(f"  ❌ Ручная обработка: {MANUAL_FOLDER}")
    print(f"  📊 Excel: {EXCEL_FILE}")
    print(f"  📝 Лог: {LOG_FOLDER}\\processing.log")
    
    # Логирование итогов
    log_processing("="*80)
    log_processing(f"ИТОГО: Всего={total_files}, Успешно={processed_count}, Дубликаты={duplicate_count}, Manual={error_count}, OCR={ocr_count}, Осталось={remaining_count}")
    log_processing(f"АВТОМАТИЗАЦИЯ: {automation_percent:.1f}%")
    log_processing("="*80)
    
    # Telegram уведомление
    if TELEGRAM_ENABLED:
        notification_text = (
            f"✅ <b>Обработка завершена!</b>\n\n"
            f"📊 <b>Результаты:</b>\n"
            f"📄 Всего: {total_files}\n"
            f"✅ Обработано: {processed_count}\n"
            f"🔄 Дубликатов: {duplicate_count}\n"
            f"❌ Manual: {error_count}\n"
            f"🔍 OCR: {ocr_count}\n\n"
            f"🎯 <b>Автоматизация: {automation_percent:.1f}%</b>\n\n"
        )
        
        if remaining_count == 0:
            notification_text += "✅ Папка полностью очищена!"
        else:
            notification_text += f"⚠ Осталось: {remaining_count} файлов"
        
        send_notification(notification_text)
    
    return {
        'total': total_files,
        'processed': processed_count,
        'duplicates': duplicate_count,
        'manual': error_count,
        'ocr': ocr_count,
        'remaining': remaining_count,
        'automation_percent': automation_percent,
        'supplier_stats': dict(supplier_stats)
    }


# ===== ТОЧКА ВХОДА =====

if __name__ == "__main__":
    try:
        print("\n")
        print("╔═══════════════════════════════════════════════════════════════╗")
        print("║         ULTIMATE PDF PROCESSOR v3.0                          ║")
        print("║         Максимальная автоматизация счетов                    ║")
        print("╚═══════════════════════════════════════════════════════════════╝")
        print("\n")
        
        # Проверка настроек
        print("Проверка системы...")
        print(f"✓ PDF папка: {os.path.exists(PDF_FOLDER)}")
        print(f"✓ Excel файл: {os.path.exists(EXCEL_FILE)}")
        print(f"✓ OCR доступен: {OCR_AVAILABLE}")
        print(f"✓ Telegram: {TELEGRAM_ENABLED}")
        
        if not os.path.exists(PDF_FOLDER):
            print("\n❌ Папка PDF не найдена!")
            exit(1)
        
        if not os.path.exists(EXCEL_FILE):
            print("\n❌ Excel файл не найден!")
            exit(1)
        
        print("\n" + "="*80)
        input("Нажмите Enter для начала обработки...")
        print("\n")
        
        # ЗАПУСК
        results = process_all_pdfs()
        
        # Сохранить отчёт
        report_path = os.path.join(LOG_FOLDER, f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt')
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("ОТЧЁТ ОБ ОБРАБОТКЕ\n")
            f.write("="*80 + "\n\n")
            f.write(f"Всего файлов: {results['total']}\n")
            f.write(f"Обработано: {results['processed']}\n")
            f.write(f"Дубликатов: {results['duplicates']}\n")
            f.write(f"Manual: {results['manual']}\n")
            f.write(f"OCR: {results['ocr']}\n")
            f.write(f"Автоматизация: {results['automation_percent']:.1f}%\n\n")
            f.write("Статистика по поставщикам:\n")
            for supplier, count in results['supplier_stats'].items():
                f.write(f"  {supplier}: {count}\n")
        
        print(f"\n📄 Отчёт сохранён: {report_path}")
        
        # Код завершения
        exit_code = 0 if results['remaining'] == 0 else 1
        exit(exit_code)
        
    except KeyboardInterrupt:
        print("\n\n⏹ Обработка прервана пользователем")
        log_processing("ПРЕРВАНО ПОЛЬЗОВАТЕЛЕМ")
        exit(2)
    except Exception as e:
        print(f"\n\n❌ Критическая ошибка: {e}")
        log_processing(f"КРИТИЧЕСКАЯ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        exit(3)    

# ПРОДОЛЖЕНИЕ В СЛЕДУЮЩЕМ СООБЩЕНИИ - ГЛАВНАЯ ФУНКЦИЯ

# ПРОДОЛЖЕНИЕ В СЛЕДУЮЩЕМ СООБЩЕНИИ...