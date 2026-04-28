"""
СИСТЕМА АВТОМАТИЧЕСКОЙ ОБРАБОТКИ PDF СЧЕТОВ
Версия 2.0 - Улучшенная версия с исправлениями
- Правильное определение номеров машин из имен файлов  
- Улучшенная Telegram интеграция
- Расширенная обработка поставщиков
"""

import os
import shutil
import pdfplumber
import openpyxl
import re
from datetime import datetime
import time
from collections import defaultdict

# ===== НАСТРОЙКИ ПУТЕЙ =====
PDF_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG"
EXCEL_FILE = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG\Repair_2025.xlsx"
PROCESSED_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG"
MANUAL_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG\manual"
LOG_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\PDF_Processor\logs"

# ===== НАСТРОЙКИ TELEGRAM =====
TELEGRAM_ENABLED = True  # Включить после настройки telegram_bot.py

if TELEGRAM_ENABLED:
    try:
        from telegram_bot import send_notification
    except ImportError:
        TELEGRAM_ENABLED = False
        print("⚠ Telegram уведомления недоступны. Проверьте telegram_bot.py")


# ===== ФУНКЦИЯ ОПРЕДЕЛЕНИЯ НОМЕРА МАШИНЫ ИЗ ИМЕНИ ФАЙЛА =====
def extract_truck_from_filename(filename):
    """Извлечь номер машины из имени файла"""
    
    # Паттерны для разных форматов файлов
    patterns = [
        # checked_15 - K&L 109455.pdf -> GR-OO15
        r'checked_(\d+)',
        # checked_186 - AC 300203.pdf -> GR-OO186
        r'checked_(\d+)',
        # checked_502 - MAN Aachen 551725327.pdf -> GR-OO502
        r'checked_(\d+)',
        # checked_771 - TIP U71_90919908.pdf -> GR-OO771
        r'checked_(\d+)',
        # checked_1511 - Vital Projekt 2500992.pdf -> GR-OO1511
        r'checked_(\d+)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            number = match.group(1)
            # Форматируем как GR-OOXX или GR-OOXXX
            if len(number) <= 2:
                return f"GR-OO{number.zfill(2)}"
            else:
                return f"GR-OO{number}"
    
    # Дополнительные паттерны для прямых указаний
    direct_patterns = [
        (r'GR[- ]?OO(\d+)', 'GR-OO'),
        (r'GR[- ]?(\d+)', 'GR-'),
        (r'WJQY4010', 'WJQY4010'),
        (r'OHAMX771', 'OHAMX771'),
        (r'HH[- ]?AG(\d+)', 'HH-AG'),
        (r'DE[- ]?FN(\d+)', 'DE-FN'),
    ]
    
    for pattern, prefix in direct_patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            if prefix in ['WJQY4010', 'OHAMX771']:
                return prefix
            number = match.group(1) if match.groups() else ''
            return f"{prefix}{number.zfill(3)}"
    
    return None


# ===== ФУНКЦИИ ИЗВЛЕЧЕНИЯ ДАННЫХ ПО ПОСТАВЩИКАМ =====

def extract_vital_projekt(text, filename=''):
    """Vital Projekt Inh.Vitalij Barth - шины и услуги"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Rechnungs-Nr\.:\s*(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Дата
    date_match = re.search(r'Hamburg,\s*den\s*(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
    else:
        date_match2 = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
        if date_match2:
            date_str = date_match2.group(1)
        else:
            return None
    
    # Преобразуем дату
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['date'] = date_obj.strftime('%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Сумма
    total_match = re.search(r'Gesamtbetrag\s+EUR\s+([\d,\.]+)', text)
    if not total_match:
        total_match = re.search(r'Summe\s+(?:brutto|netto)?\s*:?\s*([\d,\.]+)', text)
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    # Детали счета
    data['name'] = 'Reifenservice/Montage'
    data['seller'] = 'Vital Projekt Inh.Vitalij Barth'
    data['buyer'] = 'Auto Compass GmbH'
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data

def extract_autocompass(text, filename=''):
    """Auto Compass GmbH - внутренние документы"""
    data = {}
    
    # Ищем номер счета
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Nr[.\s]+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Дата
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            data['date'] = date_str
            data['month'] = 0
            data['week'] = 0
    else:
        data['date'] = ''
        data['month'] = 0
        data['week'] = 0
    
    # Сумма
    total_match = re.search(r'(?:Gesamt|Total|Summe)[:\s]+([\d,\.]+)', text)
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'Service/Reparatur'
    data['seller'] = 'Auto Compass GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['price'] = data['amount']
    data['internal'] = 'Internal'
    
    return data

def extract_groo_gmbh(text, filename=''):
    """Groo GmbH - услуги и ремонт"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Invoice\s+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Дата
    date_match = re.search(r'Rechnungsdatum[:\s]+(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            return None
    else:
        return None
    
    # Сумма
    total_match = re.search(r'Gesamtbetrag[:\s]+([\d,\.]+)', text)
    if not total_match:
        total_match = re.search(r'Total[:\s]+([\d,\.]+)', text)
    
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'Service/Wartung'
    data['seller'] = 'Groo GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data

def extract_man_truck(text, filename=''):
    """MAN Truck & Bus Deutschland GmbH"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Auftragsnummer[:\s]+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Дата
    date_match = re.search(r'Datum[:\s]+(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            return None
    else:
        return None
    
    # Сумма
    total_match = re.search(r'Endbetrag[:\s]+([\d,\.]+)', text)
    if not total_match:
        total_match = re.search(r'Gesamt[:\s]+([\d,\.]+)', text)
    
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'MAN Service/Ersatzteile'
    data['seller'] = 'MAN Truck & Bus Deutschland GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data

def extract_euromaster(text, filename=''):
    """Euromaster GmbH"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Nr[.\s]+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Дата
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            return None
    else:
        return None
    
    # Сумма
    total_match = re.search(r'Gesamt[:\s]+([\d,\.]+)', text)
    if not total_match:
        total_match = re.search(r'Total[:\s]+([\d,\.]+)', text)
    
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'Reifenservice'
    data['seller'] = 'Euromaster GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data

def extract_dekra(text, filename=''):
    """DEKRA Automobil GmbH"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Auftrag[:\s]+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Дата
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            return None
    else:
        return None
    
    # Сумма
    total_match = re.search(r'Summe[:\s]+([\d,\.]+)', text)
    if not total_match:
        total_match = re.search(r'Gesamt[:\s]+([\d,\.]+)', text)
    
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'Hauptuntersuchung'
    data['seller'] = 'DEKRA Automobil GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data

def extract_tip_trailer(text, filename=''):
    """TIP Trailer Services Germany GmbH"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Invoice[:\s]+(\w+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Rechnung[:\s]+(\w+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла - особая обработка для TIP
    truck = extract_truck_from_filename(filename)
    if not truck and 'U71' in text:
        truck = 'OHAMX771'
    data['truck'] = truck or ''
    
    # Дата
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            return None
    else:
        return None
    
    # Сумма
    total_match = re.search(r'Total[:\s]+([\d,\.]+)', text)
    if not total_match:
        total_match = re.search(r'Amount[:\s]+([\d,\.]+)', text)
    
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'Аренда прицепа'
    data['seller'] = 'TIP Trailer Services Germany GmbH'
    data['buyer'] = 'Groo GmbH'  # Обратите внимание - покупатель Groo
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data

def extract_tankkpool24(text, filename=''):
    """Tankpool24 International GmbH"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Rechnung[:\s]+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Дата
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            return None
    else:
        return None
    
    # Сумма
    total_match = re.search(r'Summe[:\s]+([\d,\.]+)', text)
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'Топливо'
    data['seller'] = 'Tankpool24 International GmbH'
    data['buyer'] = 'Groo GmbH'
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data

def extract_volvo_group(text, filename=''):
    """Volvo Group Trucks Service Nord GmbH"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Rechnung[:\s]+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Дата
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            return None
    else:
        return None
    
    # Сумма
    total_match = re.search(r'Gesamt[:\s]+([\d,\.]+)', text)
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'Volvo Service'
    data['seller'] = 'Volvo Group Trucks Service Nord GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data

def extract_kl_meisterbetrieb(text, filename=''):
    """K&L Kfz Meisterbetrieb GmbH"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Rechnung[:\s]+(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'RE(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Дата
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            return None
    else:
        return None
    
    # Сумма
    total_match = re.search(r'Gesamt[:\s]+([\d,\.]+)', text)
    if not total_match:
        total_match = re.search(r'Brutto[:\s]+([\d,\.]+)', text)
    
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'KFZ Service'
    data['seller'] = 'K&L Kfz Meisterbetrieb GmbH'
    data['buyer'] = 'Groo GmbH'
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data

def extract_quick_reifendiscount(text, filename=''):
    """Quick Reifendiscount GmbH"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Rechnung[:\s]+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Дата
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            return None
    else:
        return None
    
    # Сумма
    total_match = re.search(r'Gesamt[:\s]+([\d,\.]+)', text)
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'Шины'
    data['seller'] = 'Quick Reifendiscount GmbH'
    data['buyer'] = 'Groo GmbH'
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data

def extract_andreas_groo_ht(text, filename=''):
    """Andreas Groo Handel & Transporte"""
    data = {}
    
    # Номер счета
    invoice_match = re.search(r'Rechnung[:\s]+(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # Машина из имени файла
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Дата
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            data['date'] = date_obj.strftime('%d.%m.%Y')
            data['month'] = date_obj.month
            data['week'] = date_obj.isocalendar()[1]
        except:
            return None
    else:
        return None
    
    # Сумма
    total_match = re.search(r'Summe[:\s]+([\d,\.]+)', text)
    if total_match:
        amount_str = total_match.group(1).replace(',', '.')
        data['amount'] = float(amount_str)
    else:
        data['amount'] = 0
    
    data['name'] = 'Транспортные услуги'
    data['seller'] = 'Andreas Groo Handel & Transporte'
    data['buyer'] = 'Auto Compass GmbH'
    data['price'] = data['amount']
    data['internal'] = ''
    
    return data


# ===== ОСНОВНЫЕ ФУНКЦИИ =====

def identify_supplier(text):
    """Определить поставщика по тексту"""
    
    suppliers = {
        'Vital Projekt': ['vital projekt', 'vitalij barth'],
        'Auto Compass': ['auto compass gmbh', 'autocompass'],
        'Groo GmbH': ['groo gmbh'],
        'MAN Truck': ['man truck & bus', 'man truck and bus'],
        'Euromaster': ['euromaster'],
        'DEKRA': ['dekra automobil'],
        'TIP Trailer': ['tip trailer services'],
        'Tankpool24': ['tankpool24'],
        'Volvo Group': ['volvo group trucks'],
        'K&L Meisterbetrieb': ['k&l kfz meisterbetrieb', 'k & l'],
        'Quick Reifendiscount': ['quick reifendiscount'],
        'Andreas Groo': ['andreas groo handel'],
        'Ferronordic': ['ferronordic rental'],
    }
    
    text_lower = text.lower()
    
    for supplier, keywords in suppliers.items():
        for keyword in keywords:
            if keyword in text_lower:
                return supplier
    
    return 'Unknown'

def extract_data_by_supplier(text, supplier, filename=''):
    """Извлечь данные в зависимости от поставщика"""
    
    extractors = {
        'Vital Projekt': extract_vital_projekt,
        'Auto Compass': extract_autocompass,
        'Groo GmbH': extract_groo_gmbh,
        'MAN Truck': extract_man_truck,
        'Euromaster': extract_euromaster,
        'DEKRA': extract_dekra,
        'TIP Trailer': extract_tip_trailer,
        'Tankpool24': extract_tankkpool24,
        'Volvo Group': extract_volvo_group,
        'K&L Meisterbetrieb': extract_kl_meisterbetrieb,
        'Quick Reifendiscount': extract_quick_reifendiscount,
        'Andreas Groo': extract_andreas_groo_ht,
    }
    
    extractor = extractors.get(supplier)
    if extractor:
        return extractor(text, filename)
    return None

def check_invoice_exists(invoice_number, excel_file):
    """Проверить существование счета в Excel"""
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[7] and str(row[7]) == str(invoice_number):
                wb.close()
                return True
        
        wb.close()
        return False
    except Exception as e:
        print(f"  ⚠ Ошибка проверки дубликатов: {e}")
        return False

def add_to_excel(data, excel_file):
    """Добавить данные в Excel"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Найти последнюю заполненную строку
        last_row = ws.max_row
        while last_row > 1 and ws.cell(row=last_row, column=1).value is None:
            last_row -= 1
        
        new_row = last_row + 1
        
        # Записать данные
        ws.cell(row=new_row, column=1, value=data.get('month', ''))
        ws.cell(row=new_row, column=2, value=data.get('week', ''))
        ws.cell(row=new_row, column=3, value=data.get('truck', ''))
        ws.cell(row=new_row, column=4, value=data.get('date', ''))
        ws.cell(row=new_row, column=5, value=data.get('name', ''))
        ws.cell(row=new_row, column=6, value=data.get('amount', 0))
        ws.cell(row=new_row, column=7, value=data.get('price', 0))
        ws.cell(row=new_row, column=8, value=data.get('invoice', ''))
        ws.cell(row=new_row, column=9, value=data.get('seller', ''))
        ws.cell(row=new_row, column=10, value=data.get('buyer', ''))
        ws.cell(row=new_row, column=11, value=data.get('internal', ''))
        
        wb.save(excel_file)
        wb.close()
        return True
        
    except Exception as e:
        print(f"  ⚠ Ошибка записи в Excel: {e}")
        return False

def rename_and_move_file(pdf_path, destination_folder, prefix='checked_'):
    """Переименовать и переместить файл"""
    try:
        filename = os.path.basename(pdf_path)
        
        # Добавить префикс если его нет
        if not filename.startswith(prefix):
            new_filename = prefix + filename
        else:
            new_filename = filename
        
        new_path = os.path.join(destination_folder, new_filename)
        
        # Если файл существует - добавить временную метку
        if os.path.exists(new_path):
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            name_part, ext = os.path.splitext(new_filename)
            new_filename = f"{name_part}_{timestamp}{ext}"
            new_path = os.path.join(destination_folder, new_filename)
        
        shutil.move(pdf_path, new_path)
        return True, new_filename, None
        
    except Exception as e:
        return False, '', str(e)

def log_processing(message):
    """Записать в лог"""
    try:
        os.makedirs(LOG_FOLDER, exist_ok=True)
        log_file = os.path.join(LOG_FOLDER, f"processing_{datetime.now().strftime('%Y%m%d')}.log")
        
        with open(log_file, 'a', encoding='utf-8') as f:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            f.write(f"{timestamp} | {message}\n")
    except:
        pass

def process_all_pdfs():
    """Основная функция обработки всех PDF"""
    
    print("\n" + "="*80)
    print("ОБРАБОТКА ФАЙЛОВ - Версия 2.0")
    print("="*80)
    
    # Счётчики
    processed_count = 0
    duplicate_count = 0
    error_count = 0
    
    # Статистика по поставщикам
    supplier_stats = defaultdict(int)
    
    log_processing("="*80)
    log_processing("НАЧАЛО ОБРАБОТКИ - Версия 2.0")
    log_processing("="*80)
    
    # Создать папки
    os.makedirs(MANUAL_FOLDER, exist_ok=True)
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)
    
    # Получить список PDF
    try:
        pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    except Exception as e:
        print(f"❌ Ошибка чтения папки: {e}")
        return {'total': 0, 'processed': 0, 'duplicates': 0, 'manual': 0}
    
    total_files = len(pdf_files)
    
    print(f"\n📁 Папка: {PDF_FOLDER}")
    print(f"📄 Найдено файлов: {total_files}")
    
    if total_files == 0:
        print("\n✓ Нет файлов для обработки")
        if TELEGRAM_ENABLED:
            send_notification("✅ Нет новых счетов для обработки")
        return {'total': 0, 'processed': 0, 'duplicates': 0, 'manual': 0}
    
    # Telegram уведомление о начале
    if TELEGRAM_ENABLED:
        notification_text = f"🚀 <b>Запуск обработки PDF v2.0</b>\n\n"
        notification_text += f"📁 Файлов к обработке: {total_files}\n"
        notification_text += f"⏰ Время: {datetime.now().strftime('%H:%M')}"
        send_notification(notification_text)
    
    print("\n" + "="*80)
    print("ОБРАБОТКА ФАЙЛОВ")
    print("="*80)
    
    # Обработка каждого файла
    for index, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        
        print(f"\n[{index}/{total_files}] {pdf_file}")
        
        try:
            # ===== ШАГ 1: Извлечь текст =====
            full_text = ''
            
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + '\n'
            
            if not full_text.strip():
                print(f"  ⚠ PDF пустой")
                success, new_name, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                if success:
                    print(f"  → manual/{new_name}")
                    log_processing(f"MANUAL (пустой): {pdf_file} -> {new_name}")
                error_count += 1
                continue
            
            # ===== ШАГ 2: Классифицировать =====
            supplier = identify_supplier(full_text)
            print(f"  Поставщик: {supplier}")
            supplier_stats[supplier] += 1
            
            # ===== ШАГ 3: Извлечь данные =====
            data = extract_data_by_supplier(full_text, supplier, pdf_file)
            
            # Проверка извлечения номера машины
            if data and not data.get('truck'):
                print(f"  ⚠ Не определен номер машины")
                # Попытка дополнительного извлечения
                data['truck'] = extract_truck_from_filename(pdf_file) or ''
            
            if not data:
                print(f"  ❌ Не удалось извлечь данные")
                success, new_name, error_msg = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'manual_'
                )
                
                if success:
                    print(f"  → manual/{new_name}")
                    log_processing(f"MANUAL ({supplier}): {pdf_file} -> {new_name}")
                else:
                    print(f"  ⚠ Ошибка перемещения: {error_msg}")
                
                error_count += 1
                continue
            
            # Отладочный вывод
            if data.get('truck'):
                print(f"  Машина: {data['truck']}")
            
            # ===== ШАГ 4: Проверить дубликат =====
            if check_invoice_exists(data['invoice'], EXCEL_FILE):
                print(f"  ⚠ ДУБЛИКАТ! Счет {data['invoice']} уже существует")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'duplicate_'
                )
                if success:
                    print(f"  → processed/{new_name}")
                    log_processing(f"DUPLICATE: {pdf_file} -> {new_name}")
                duplicate_count += 1
                continue
            
            # ===== ШАГ 5: Добавить в Excel =====
            if add_to_excel(data, EXCEL_FILE):
                print(f"  ✓ Добавлен в Excel: счет {data['invoice']}")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'checked_'
                )
                if success:
                    print(f"  → processed/{new_name}")
                    log_processing(f"SUCCESS: {pdf_file} -> {new_name}")
                processed_count += 1
            else:
                print(f"  ❌ Ошибка добавления в Excel")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'error_'
                )
                if success:
                    print(f"  → manual/{new_name}")
                error_count += 1
                
        except Exception as e:
            print(f"  ❌ Ошибка обработки: {e}")
            log_processing(f"ERROR: {pdf_file} - {str(e)}")
            try:
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'error_'
                )
                if success:
                    print(f"  → manual/{new_name}")
            except:
                pass
            error_count += 1
    
    # ===== ИТОГИ =====
    print("\n" + "="*80)
    print("ИТОГИ ОБРАБОТКИ")
    print("="*80)
    print(f"Всего файлов: {total_files}")
    print(f"✓ Обработано успешно: {processed_count}")
    print(f"⚠ Дубликатов: {duplicate_count}")
    print(f"❌ Требуют ручной обработки: {error_count}")
    
    # Статистика по поставщикам
    print("\nСтатистика по поставщикам:")
    for supplier, count in sorted(supplier_stats.items(), key=lambda x: x[1], reverse=True):
        print(f"  {supplier}: {count}")
    
    # Проверка оставшихся файлов
    try:
        remaining_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
        remaining_count = len(remaining_files)
        if remaining_count > 0:
            print(f"\n⚠ Остались необработанные файлы: {remaining_count}")
    except:
        remaining_count = 0
    
    log_processing("="*80)
    log_processing(f"ЗАВЕРШЕНО: Обработано={processed_count}, Дубликаты={duplicate_count}, Ручные={error_count}")
    log_processing("="*80)
    
    # Telegram уведомление об окончании
    if TELEGRAM_ENABLED:
        notification_text = f"✅ <b>Обработка завершена!</b>\n\n"
        notification_text += f"📊 <b>Результаты:</b>\n"
        notification_text += f"✓ Обработано: {processed_count}\n"
        notification_text += f"⚠ Дубликатов: {duplicate_count}\n"
        notification_text += f"❌ Ручная обработка: {error_count}\n"
        notification_text += f"📈 Всего: {total_files}\n\n"
        
        if supplier_stats:
            notification_text += f"<b>Топ поставщиков:</b>\n"
            top_suppliers = sorted(supplier_stats.items(), key=lambda x: x[1], reverse=True)[:5]
            for supplier, count in top_suppliers:
                notification_text += f"• {supplier}: {count}\n"
        
        if remaining_count > 0:
            notification_text += f"\n⚠ Осталось: {remaining_count} файлов"
        
        send_notification(notification_text)
    
    return {
        'total': total_files,
        'processed': processed_count,
        'duplicates': duplicate_count,
        'manual': error_count,
        'remaining': remaining_count,
        'supplier_stats': dict(supplier_stats)
    }


# ===== ТОЧКА ВХОДА =====

if __name__ == "__main__":
    try:
        print("\n")
        print("╔══════════════════════════════════════════════════════════════╗")
        print("║         PDF PROCESSOR v2.0 - УЛУЧШЕННАЯ ВЕРСИЯ              ║")
        print("║         Автоматическая обработка счетов                     ║")
        print("╚══════════════════════════════════════════════════════════════╝")
        print("\n")
        
        # Проверка настроек
        print("Проверка системы...")
        print(f"✓ PDF папка: {os.path.exists(PDF_FOLDER)}")
        print(f"✓ Excel файл: {os.path.exists(EXCEL_FILE)}")
        print(f"✓ Telegram: {TELEGRAM_ENABLED}")
        
        if not os.path.exists(PDF_FOLDER):
            print("\n❌ Папка PDF не найдена!")
            exit(1)
        
        if not os.path.exists(EXCEL_FILE):
            print("\n❌ Excel файл не найден!")
            exit(1)
        
        # Запуск обработки
        print("\nЗапуск обработки...")
        results = process_all_pdfs()
        
        print("\n✅ Программа завершена успешно!")
        
        # Вывод финальной статистики
        if results['total'] > 0:
            success_rate = (results['processed'] / results['total']) * 100
            print(f"\n📊 Процент автоматизации: {success_rate:.1f}%")
        
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        log_processing(f"CRITICAL ERROR: {str(e)}")
        
        if TELEGRAM_ENABLED:
            send_notification(f"❌ <b>Критическая ошибка!</b>\n\n{str(e)}")
        
        exit(1)