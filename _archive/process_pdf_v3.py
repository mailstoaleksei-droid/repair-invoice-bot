"""
PDF PROCESSOR v3.0 - С ДЕТАЛЬНЫМИ УВЕДОМЛЕНИЯМИ
Автоматическая обработка PDF счетов с подробными Telegram уведомлениями
"""

import os
import shutil
import pdfplumber
import openpyxl
import re
from datetime import datetime
import time
from collections import defaultdict

# ===== НАСТРОЙКИ ПУТЕЙ =====
PDF_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG"
EXCEL_FILE = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG\Repair_2025.xlsx"
PROCESSED_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG"
MANUAL_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG\manual"
LOG_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\PDF_Processor\logs"

# ===== НАСТРОЙКИ TELEGRAM =====
TELEGRAM_ENABLED = True
DETAILED_NOTIFICATIONS = True  # Включить детальные уведомления

# Импорт Telegram уведомлений
if TELEGRAM_ENABLED:
    try:
        import telebot
        from telegram_notifications import TelegramNotifier
        
        TELEGRAM_BOT_TOKEN = "8127115250:AAHmDuiiRuPSpE6oSwHzmUpSl2-DzVSr3Io"
        TELEGRAM_CHAT_ID = "745125435"
        
        bot = telebot.TeleBot(TELEGRAM_BOT_TOKEN)
        notifier = TelegramNotifier(bot, TELEGRAM_CHAT_ID, throttle_interval=2.0)
        
    except ImportError as e:
        TELEGRAM_ENABLED = False
        print(f"⚠ Telegram уведомления недоступны: {e}")


# ===== ФУНКЦИИ ИЗ process_pdf_v2.py (скопированы без изменений) =====

def extract_truck_from_filename(filename):
    """Извлечь номер машины из имени файла"""
    
    # Паттерн 1: checked_XXX (основной формат)
    patterns = [
        r'checked_(\d+)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            number = match.group(1)
            if len(number) <= 2:
                return f"GR-OO{number.zfill(2)}"
            else:
                return f"GR-OO{number}"
    
    # Паттерн 2: XXXX - Поставщик (новый формат)
    # Примеры: "1726 - AC Intern.pdf", "771 - TIP.pdf"
    simple_number_pattern = r'^(\d{2,4})\s*[-–]\s*'
    match = re.match(simple_number_pattern, filename)
    if match:
        number = match.group(1)
        if len(number) <= 2:
            return f"GR-OO{number.zfill(2)}"
        elif len(number) == 3:
            return f"GR-OO{number}"
        else:  # 4 digits
            return f"GR-OO{number}"
    
    # Паттерн 3: Прямые указания (GR-OO, HH-AG, etc.)
    direct_patterns = [
        (r'GR[- ]?OO(\d+)', 'GR-OO'),
        (r'GR[- ]?(\d+)', 'GR-'),
        (r'WJQY4010', 'WJQY4010'),
        (r'OHAMX771', 'OHAMX771'),
        (r'HH[- ]?AG(\d+)', 'HH-AG'),
        (r'DE[- ]?FN(\d+)', 'DE-FN'),
    ]
    
    for pattern, prefix in direct_patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            if prefix in ['WJQY4010', 'OHAMX771']:
                return prefix
            number = match.group(1) if match.groups() else ''
            return f"{prefix}{number.zfill(3)}"
    
    return None


def identify_supplier(text):
    """Определить поставщика по тексту PDF"""
    text_lower = text.lower()
    
    if 'vital projekt' in text_lower or 'vitalij barth' in text_lower:
        return 'Vital Projekt'
    elif 'man truck' in text_lower or 'man nutzfahrzeuge' in text_lower:
        return 'MAN'
    elif 'euromaster' in text_lower:
        return 'Euromaster'
    elif 'auto compass' in text_lower:
        return 'Auto Compass'
    elif 'groo gmbh' in text_lower:
        return 'Groo GmbH'
    elif 'k&l' in text_lower or 'kfz meisterbetrieb' in text_lower:
        return 'K&L'
    elif 'scania' in text_lower:
        return 'Scania'
    elif 'volvo group' in text_lower:
        return 'Volvo'
    elif 'ferronordic' in text_lower:
        return 'Ferronordic'
    elif 'tip trailer' in text_lower:
        return 'TIP Trailer'
    elif 'dekra' in text_lower:
        return 'DEKRA'
    else:
        return 'Unknown'


def extract_data_by_supplier(text, supplier, filename):
    """Извлечь данные в зависимости от поставщика"""
    # Базовая структура данных
    data = {
        'invoice': None,
        'date': None,
        'amount': 0,
        'truck': extract_truck_from_filename(filename),
        'seller': supplier,
        'buyer': 'Auto Compass GmbH',
        'name': '',
        'price': 0,
        'total_price': 0,
        'month': 0,
        'week': 0,
    }
    
    # Извлечение номера счета (универсальное)
    invoice_patterns = [
        r'Rechnung[s-]?[Nn]r\.?\s*:?\s*(\d+)',
        r'Invoice\s+(?:No\.?|Number)\s*:?\s*(\d+)',
        r'Rg\.?\s*-?\s*Nr\.?\s*:?\s*(\d+)',
        r'Rechnungs-Nr\.:\s*(\d+)',
    ]
    
    for pattern in invoice_patterns:
        match = re.search(pattern, text)
        if match:
            data['invoice'] = match.group(1)
            break
    
    if not data['invoice']:
        return None
    
    # Извлечение даты (универсальное)
    date_patterns = [
        r'(\d{2}\.\d{2}\.\d{4})',
        r'(\d{2})-(\d{2})-(\d{4})',
    ]
    
    for pattern in date_patterns:
        match = re.search(pattern, text)
        if match:
            if len(match.groups()) == 1:
                date_str = match.group(1)
            else:
                date_str = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
            
            try:
                date_obj = datetime.strptime(date_str.replace('-', '.'), '%d.%m.%Y')
                data['date'] = date_obj.strftime('%d.%m.%Y')
                data['month'] = date_obj.month
                data['week'] = date_obj.isocalendar()[1]
            except:
                pass
            break
    
    # Извлечение суммы (универсальное)
    amount_patterns = [
        r'Gesamtbetrag\s+(?:EUR)?\s*([\d,\.]+)',
        r'Summe\s+(?:brutto|netto)?\s*:?\s*([\d,\.]+)',
        r'Total\s*:?\s*([\d,\.]+)',
        r'Gesamt\s*:?\s*([\d,\.]+)',
    ]
    
    for pattern in amount_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            amount_str = match.group(1).replace(',', '.').replace(' ', '')
            try:
                data['amount'] = float(amount_str)
                data['total_price'] = data['amount']
                data['price'] = data['amount']
            except:
                pass
            break
    
    return data if data['invoice'] and data['date'] else None


def check_invoice_exists(invoice_number, excel_file):
    """Проверить существует ли счет в Excel"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[8] == invoice_number:  # Колонка I (Invoice)
                wb.close()
                return True
        
        wb.close()
        return False
    except Exception as e:
        print(f"Ошибка проверки дубликата: {e}")
        return False


def add_to_excel(data, excel_file):
    """Добавить данные в Excel"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Найти первую пустую строку
        row_num = 2
        while ws[f'A{row_num}'].value is not None:
            row_num += 1
        
        # Заполнить данные
        ws[f'A{row_num}'] = data.get('month', 0)
        ws[f'B{row_num}'] = data.get('week', 0)
        ws[f'C{row_num}'] = data.get('truck', '')
        ws[f'D{row_num}'] = data.get('date', '')
        ws[f'E{row_num}'] = data.get('name', '')
        ws[f'F{row_num}'] = data.get('amount', 1)
        ws[f'G{row_num}'] = data.get('price', 0)
        ws[f'H{row_num}'] = data.get('total_price', 0)
        ws[f'I{row_num}'] = data.get('invoice', '')
        ws[f'J{row_num}'] = data.get('seller', '')
        ws[f'K{row_num}'] = data.get('buyer', 'Auto Compass GmbH')
        
        wb.save(excel_file)
        wb.close()
        
        return row_num
        
    except Exception as e:
        print(f"Ошибка добавления в Excel: {e}")
        return None


def rename_and_move_file(source_path, destination_folder, prefix=''):
    """Переименовать и переместить файл"""
    try:
        filename = os.path.basename(source_path)
        
        # Убрать старые префиксы
        clean_filename = filename
        for old_prefix in ['checked_', 'manual_', 'error_', 'duplicate_']:
            if clean_filename.startswith(old_prefix):
                clean_filename = clean_filename[len(old_prefix):]
        
        # Добавить новый префикс
        new_filename = f"{prefix}{clean_filename}"
        destination_path = os.path.join(destination_folder, new_filename)
        
        # Если файл уже существует, добавить timestamp
        if os.path.exists(destination_path):
            name, ext = os.path.splitext(new_filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            new_filename = f"{name}_{timestamp}{ext}"
            destination_path = os.path.join(destination_folder, new_filename)
        
        shutil.move(source_path, destination_path)
        return True, new_filename, None
        
    except Exception as e:
        return False, None, str(e)


def log_processing(message):
    """Логирование в файл"""
    try:
        os.makedirs(LOG_FOLDER, exist_ok=True)
        log_file = os.path.join(LOG_FOLDER, f"processing_{datetime.now().strftime('%Y%m%d')}.log")
        
        with open(log_file, 'a', encoding='utf-8') as f:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            f.write(f"[{timestamp}] {message}\n")
    except Exception as e:
        print(f"Ошибка логирования: {e}")


# ===== ГЛАВНАЯ ФУНКЦИЯ С ДЕТАЛЬНЫМИ УВЕДОМЛЕНИЯМИ =====

def process_all_pdfs():
    """
    Обработать все PDF файлы с детальными Telegram уведомлениями
    """
    
    print("\n╔══════════════════════════════════════════════════════════════╗")
    print("║         PDF PROCESSOR v3.0 - С ДЕТАЛЬНЫМИ УВЕДОМЛЕНИЯМИ     ║")
    print("╚══════════════════════════════════════════════════════════════╝\n")
    
    # Создать папки если не существуют
    os.makedirs(MANUAL_FOLDER, exist_ok=True)
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)
    os.makedirs(LOG_FOLDER, exist_ok=True)
    
    # Получить список файлов
    try:
        pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    except Exception as e:
        print(f"❌ Ошибка доступа к папке: {e}")
        return None
    
    if not pdf_files:
        print("📭 Нет файлов для обработки")
        return {'total': 0, 'processed': 0, 'duplicates': 0, 'manual': 0}
    
    total_files = len(pdf_files)
    processed_count = 0
    duplicate_count = 0
    error_count = 0
    
    # Для сводки
    processed_files = []
    duplicate_files = []
    manual_files = []
    
    supplier_stats = defaultdict(int)
    
    log_processing("="*80)
    log_processing(f"НАЧАЛО ОБРАБОТКИ: {total_files} файлов")
    log_processing("="*80)
    
    print(f"\n📊 Найдено файлов: {total_files}")
    print("="*80)
    
    # Обработка каждого файла
    for index, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        
        print(f"\n[{index}/{total_files}] {pdf_file}")
        
        # Уведомление о начале обработки
        if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
            notifier.notify_processing_start(pdf_file)
        
        try:
            # ===== ШАГ 1: Извлечь текст =====
            full_text = ''
            
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + '\n'
            
            if not full_text.strip():
                print(f"  ⚠ PDF пустой")
                success, new_name, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                
                if success:
                    print(f"  → manual/{new_name}")
                    log_processing(f"MANUAL (пустой): {pdf_file} -> {new_name}")
                    manual_files.append(pdf_file)
                    
                    # Уведомление
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        notifier.notify_manual(pdf_file, "PDF пустой")
                
                error_count += 1
                continue
            
            # ===== ШАГ 2: Классифицировать =====
            supplier = identify_supplier(full_text)
            print(f"  Поставщик: {supplier}")
            supplier_stats[supplier] += 1
            
            # ===== ШАГ 3: Извлечь данные =====
            data = extract_data_by_supplier(full_text, supplier, pdf_file)
            
            if data and not data.get('truck'):
                print(f"  ⚠ Не определен номер машины")
                data['truck'] = extract_truck_from_filename(pdf_file) or ''
            
            if not data:
                print(f"  ❌ Не удалось извлечь данные")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'manual_'
                )
                
                if success:
                    print(f"  → manual/{new_name}")
                    log_processing(f"MANUAL ({supplier}): {pdf_file} -> {new_name}")
                    manual_files.append(pdf_file)
                    
                    # Уведомление
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        notifier.notify_manual(pdf_file, "Не удалось извлечь данные", supplier)
                
                error_count += 1
                continue
            
            if data.get('truck'):
                print(f"  Машина: {data['truck']}")
            
            # ===== ШАГ 4: Проверить дубликат =====
            if check_invoice_exists(data['invoice'], EXCEL_FILE):
                print(f"  ⚠ ДУБЛИКАТ! Счет {data['invoice']} уже существует")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'duplicate_'
                )
                
                if success:
                    print(f"  → processed/{new_name}")
                    log_processing(f"DUPLICATE: {pdf_file} -> {new_name}")
                    duplicate_files.append(pdf_file)
                    
                    # Уведомление
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        notifier.notify_duplicate(pdf_file, data['invoice'], data.get('date'))
                
                duplicate_count += 1
                continue
            
            # ===== ШАГ 5: Добавить в Excel =====
            excel_row = add_to_excel(data, EXCEL_FILE)
            
            if excel_row:
                print(f"  ✓ Добавлен в Excel: строка {excel_row}, счет {data['invoice']}")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'checked_'
                )
                
                if success:
                    print(f"  → processed/{new_name}")
                    log_processing(f"SUCCESS: {pdf_file} -> {new_name}, Excel row {excel_row}")
                    processed_files.append(pdf_file)
                    
                    # Уведомление об успехе
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        notifier.notify_success(data, pdf_file, excel_row)
                
                processed_count += 1
            else:
                print(f"  ❌ Ошибка добавления в Excel")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'error_'
                )
                
                if success:
                    print(f"  → manual/{new_name}")
                    manual_files.append(pdf_file)
                    
                    # Уведомление
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        notifier.notify_manual(pdf_file, "Ошибка добавления в Excel", supplier)
                
                error_count += 1
                
        except Exception as e:
            print(f"  ❌ Ошибка обработки: {e}")
            log_processing(f"ERROR: {pdf_file} - {str(e)}")
            
            try:
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'error_'
                )
                if success:
                    print(f"  → manual/{new_name}")
                    manual_files.append(pdf_file)
                    
                    # Уведомление об ошибке
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        notifier.notify_error(pdf_file, str(e))
            except:
                pass
            
            error_count += 1
    
    # ===== ИТОГИ =====
    print("\n" + "="*80)
    print("ИТОГИ ОБРАБОТКИ")
    print("="*80)
    print(f"Всего файлов: {total_files}")
    print(f"✓ Обработано успешно: {processed_count}")
    print(f"⚠ Дубликатов: {duplicate_count}")
    print(f"❌ Требуют ручной обработки: {error_count}")
    
    print("\nСтатистика по поставщикам:")
    for supplier, count in sorted(supplier_stats.items(), key=lambda x: x[1], reverse=True):
        print(f"  {supplier}: {count}")
    
    try:
        remaining_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
        remaining_count = len(remaining_files)
        if remaining_count > 0:
            print(f"\n⚠ Остались необработанные файлы: {remaining_count}")
    except:
        remaining_count = 0
    
    log_processing("="*80)
    log_processing(f"ЗАВЕРШЕНО: Обработано={processed_count}, Дубликаты={duplicate_count}, Ручные={error_count}")
    log_processing("="*80)
    
    # Финальная сводка в Telegram
    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
        summary = {
            'total': total_files,
            'processed': processed_count,
            'duplicates': duplicate_count,
            'manual': error_count,
            'processed_files': processed_files,
            'duplicate_files': duplicate_files,
            'manual_files': manual_files,
        }
        notifier.notify_summary(summary)
        
        # Вывести статистику уведомлений
        stats = notifier.get_stats()
        print(f"\n📱 Статистика уведомлений:")
        print(f"   Отправлено: {stats['sent']}")
        print(f"   Ошибок: {stats['failed']}")
    
    return {
        'total': total_files,
        'processed': processed_count,
        'duplicates': duplicate_count,
        'manual': error_count,
        'remaining': remaining_count,
        'supplier_stats': dict(supplier_stats)
    }


# ===== ТОЧКА ВХОДА =====

if __name__ == "__main__":
    try:
        print("\n")
        print("╔══════════════════════════════════════════════════════════════╗")
        print("║         PDF PROCESSOR v3.0 - УЛУЧШЕННАЯ ВЕРСИЯ              ║")
        print("║         С детальными Telegram уведомлениями                 ║")
        print("╚══════════════════════════════════════════════════════════════╝")
        print("\n")
        
        print("Проверка системы...")
        print(f"✓ PDF папка: {os.path.exists(PDF_FOLDER)}")
        print(f"✓ Excel файл: {os.path.exists(EXCEL_FILE)}")
        print(f"✓ Telegram: {TELEGRAM_ENABLED}")
        print(f"✓ Детальные уведомления: {DETAILED_NOTIFICATIONS}")
        
        if not os.path.exists(PDF_FOLDER):
            print("\n❌ Папка PDF не найдена!")
            exit(1)
        
        if not os.path.exists(EXCEL_FILE):
            print("\n❌ Excel файл не найден!")
            exit(1)
        
        print("\nЗапуск обработки...")
        results = process_all_pdfs()
        
        print("\n✅ Программа завершена успешно!")
        
        if results and results['total'] > 0:
            success_rate = (results['processed'] / results['total']) * 100
            print(f"\n📊 Процент автоматизации: {success_rate:.1f}%")
        
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        log_processing(f"CRITICAL ERROR: {str(e)}")
        
        if TELEGRAM_ENABLED:
            try:
                notifier.notify_error("System", str(e))
            except:
                pass
        
        exit(1)
