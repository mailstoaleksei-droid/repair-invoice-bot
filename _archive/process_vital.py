import os
import shutil
import pdfplumber
import openpyxl
import re
from datetime import datetime
import time

# Пути
PDF_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG"
EXCEL_FILE = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG\Repair_2025.xlsx"
PROCESSED_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG"
LOG_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\PDF_Processor\logs"

def extract_vital_projekt(text):
    """Извлечь данные из PDF счёта Vital Projekt"""
    
    data = {}
    
    # 1. Номер счёта
    invoice_match = re.search(r'Rechnungs-Nr\.:\s*(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    # 2. Дата
    date_match = re.search(r'Hamburg,\s*den\s*(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
    else:
        date_match2 = re.search(r'Leistungsdatum/-Ort:(\d{2})-(\d{2})-(\d{4})', text)
        if date_match2:
            date_str = f"{date_match2.group(1)}.{date_match2.group(2)}.{date_match2.group(3)}"
        else:
            return None
    
    data['date'] = date_str
    
    # Вычислить Month и Week
    date_obj = datetime.strptime(date_str, '%d.%m.%Y')
    data['month'] = date_obj.month
    data['week'] = date_obj.isocalendar()[1]
    
    # 3. Номер машины
    truck_match = re.search(r'Kennzeichen/Fahrer:\s*([A-Z]{2}\s*[A-Z]{2}\s*\d+)', text)
    if not truck_match:
        return None
    
    truck_raw = truck_match.group(1)
    truck_clean = re.sub(r'\s+', '', truck_raw)
    if len(truck_clean) >= 4:
        data['truck'] = f"{truck_clean[:2]}-{truck_clean[2:]}"
    else:
        data['truck'] = truck_clean
    
    # 4. Извлечь данные из таблицы
    table_match = re.search(
        r'^\s*1\s+(\d+)\s+Stk\.\s+[^\s]+\s+(.+?)\s+\d+%\s+([\d,]+)\s*€\s+([\d,.]+)\s*€',
        text,
        re.MULTILINE
    )
    
    if table_match:
        data['amount'] = int(table_match.group(1))
        description = table_match.group(2).strip()
        data['name'] = description[:50] if len(description) > 50 else description
        price_str = table_match.group(3).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['amount'] = 1
        data['name'] = "Шины и услуги"
        data['price'] = 0.0
    
    # 5. Summe (нетто)
    summe_match = re.search(r'Summe\s+([\d,.]+)\s*€', text)
    if summe_match:
        total_str = summe_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    # 6. Продавец и покупатель
    data['seller'] = 'Vital Projekt Inh.Vitalij Barth'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data

def check_invoice_exists(invoice_number, excel_path):
    """Проверить, существует ли уже счёт с таким номером в Excel"""
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Колонка I содержит Invoice
    for row in range(2, ws.max_row + 1):  # Начинаем со 2-й строки (пропускаем заголовок)
        cell_value = ws[f'I{row}'].value
        if cell_value and str(cell_value) == str(invoice_number):
            wb.close()
            return True
    
    wb.close()
    return False

def write_to_excel(data, excel_path):
    """Добавить строку в Excel"""
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Найти первую пустую строку
    row = ws.max_row + 1
    
    # Заполнить данные
    ws[f'A{row}'] = data['month']
    ws[f'B{row}'] = data['week']
    ws[f'C{row}'] = data['truck']
    ws[f'D{row}'] = data['date']
    ws[f'E{row}'] = data['name']
    ws[f'F{row}'] = data['amount']
    ws[f'G{row}'] = data['price']
    ws[f'H{row}'] = data['total_price']
    ws[f'I{row}'] = data['invoice']
    ws[f'J{row}'] = data['seller']
    ws[f'K{row}'] = data['buyer']
    ws[f'L{row}'] = data['interne_rechnung']
    
    wb.save(excel_path)
    wb.close()
    return row

def move_file_safe(source, destination, max_attempts=3):
    """Безопасное перемещение файла с повторными попытками"""
    
    for attempt in range(max_attempts):
        try:
            # Если файл уже существует в папке назначения, добавить (1), (2) и т.д.
            if os.path.exists(destination):
                base, ext = os.path.splitext(destination)
                counter = 1
                while os.path.exists(f"{base}({counter}){ext}"):
                    counter += 1
                destination = f"{base}({counter}){ext}"
            
            shutil.move(source, destination)
            return True, None
            
        except PermissionError as e:
            if attempt < max_attempts - 1:
                time.sleep(1)  # Подождать 1 секунду перед повторной попыткой
                continue
            else:
                return False, f"Файл используется другим процессом"
        except Exception as e:
            return False, str(e)
    
    return False, "Не удалось переместить после нескольких попыток"

def log_processing(message, log_file='processing.log'):
    """Записать в лог файл"""
    log_path = os.path.join(LOG_FOLDER, log_file)
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    with open(log_path, 'a', encoding='utf-8') as f:
        f.write(f"[{timestamp}] {message}\n")

def process_vital_pdfs():
    """Обработать все PDF от Vital Projekt"""
    
    print("="*80)
    print("ОБРАБОТКА PDF ОТ VITAL PROJEKT")
    print("="*80)
    
    processed_count = 0
    duplicate_count = 0
    error_count = 0
    move_error_count = 0
    
    # Создать лог файл
    log_processing("="*80)
    log_processing("НАЧАЛО ОБРАБОТКИ VITAL PROJEKT")
    log_processing("="*80)
    
    # Получить список всех PDF
    pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    
    for pdf_file in pdf_files:
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        
        try:
            # Открыть PDF
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
                
                # Проверить, что это Vital Projekt
                if 'VITAL PROJEKT' not in text.upper():
                    continue
                
                print(f"\nОбработка: {pdf_file}")
                
                # Извлечь данные
                data = extract_vital_projekt(text)
                
                if not data:
                    print(f"  ✗ Не удалось извлечь данные")
                    log_processing(f"ОШИБКА: {pdf_file} - не удалось извлечь данные")
                    error_count += 1
                    continue
                
                # Проверить на дубликат
                if check_invoice_exists(data['invoice'], EXCEL_FILE):
                    print(f"  ⚠ Дубликат! Invoice {data['invoice']} уже существует в Excel")
                    log_processing(f"ДУБЛИКАТ: {pdf_file} - Invoice {data['invoice']} уже есть")
                    duplicate_count += 1
                    
                    # Всё равно перемещаем файл в обработанные
                    destination = os.path.join(PROCESSED_FOLDER, pdf_file)
                    success, error_msg = move_file_safe(pdf_path, destination)
                    
                    if success:
                        print(f"  ✓ Файл перемещён в обработанные (без записи в Excel)")
                    else:
                        print(f"  ⚠ Не удалось переместить файл: {error_msg}")
                        move_error_count += 1
                    
                    continue
                
                # Записать в Excel
                row_num = write_to_excel(data, EXCEL_FILE)
                print(f"  ✓ Данные записаны в строку {row_num}")
                print(f"    Invoice: {data['invoice']}, Truck: {data['truck']}, Total: {data['total_price']:.2f} €")
                
                # Переместить PDF в обработанные
                destination = os.path.join(PROCESSED_FOLDER, pdf_file)
                success, error_msg = move_file_safe(pdf_path, destination)
                
                if success:
                    print(f"  ✓ Файл перемещён в обработанные")
                    log_processing(f"УСПЕХ: {pdf_file} -> Строка {row_num} -> Invoice {data['invoice']}")
                    processed_count += 1
                else:
                    print(f"  ⚠ Данные записаны, но не удалось переместить файл: {error_msg}")
                    print(f"  ⚠ Переместите файл вручную: {pdf_file}")
                    log_processing(f"ЧАСТИЧНО: {pdf_file} -> Строка {row_num} -> Не перемещён: {error_msg}")
                    processed_count += 1
                    move_error_count += 1
                
        except Exception as e:
            print(f"  ✗ Ошибка: {e}")
            log_processing(f"ОШИБКА: {pdf_file} - {e}")
            error_count += 1
    
    # Итоги
    print("\n" + "="*80)
    print("ИТОГИ")
    print("="*80)
    print(f"Успешно обработано: {processed_count} файлов")
    print(f"Дубликатов (пропущено): {duplicate_count}")
    print(f"Ошибок извлечения данных: {error_count}")
    print(f"Ошибок перемещения файлов: {move_error_count}")
    
    log_processing("="*80)
    log_processing(f"ИТОГО: Успешно={processed_count}, Дубликаты={duplicate_count}, Ошибки={error_count}, Не перемещено={move_error_count}")
    log_processing("="*80)
    
    if processed_count > 0:
        print(f"\nПроверьте Excel файл:")
        print(f"  {EXCEL_FILE}")
        print(f"\nОбработанные PDF перемещены в:")
        print(f"  {PROCESSED_FOLDER}")
    
    if move_error_count > 0:
        print(f"\n⚠ ВНИМАНИЕ: {move_error_count} файлов не удалось переместить автоматически")
        print(f"   Закройте программы просмотра PDF и переместите вручную")

if __name__ == "__main__":
    process_vital_pdfs()