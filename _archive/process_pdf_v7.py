"""
PDF PROCESSOR v7.0 - ФИНАЛЬНАЯ ВЕРСИЯ
Unified Telegram + 16 поставщиков + правильная логика NETTO/BRUTTO
Исправлено: Auto Compass (Internal BRUTTO, External NETTO)
Добавлено: 13 новых функций извлечения
"""
import os
import shutil
import pdfplumber
import openpyxl
import re
from datetime import datetime
import time
from collections import defaultdict

# ===== ÐÐÐ¡Ð¢Ð ÐžÐ™ÐšÐ˜ ÐŸÐ£Ð¢Ð•Ð™ =====
PDF_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG"
EXCEL_FILE = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG\Repair_2025.xlsx"
PROCESSED_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG"
MANUAL_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG\manual"
LOG_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\PDF_Processor\logs"

# ===== ÐÐÐ¡Ð¢Ð ÐžÐ™ÐšÐ˜ TELEGRAM =====
TELEGRAM_ENABLED = True
DETAILED_NOTIFICATIONS = True  # Ð’ÐºÐ»ÑŽÑ‡Ð¸Ñ‚ÑŒ Ð´ÐµÑ‚Ð°Ð»ÑŒÐ½Ñ‹Ðµ ÑƒÐ²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ñ

# Ð˜Ð¼Ð¿Ð¾Ñ€Ñ‚ Telegram ÑƒÐ²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ð¹ (UNIFIED)
if TELEGRAM_ENABLED:
    try:
        from unified_telegram import create_client
        
        TELEGRAM_BOT_TOKEN = "8127115250:AAHmDuiiRuPSpE6oSwHzmUpSl2-DzVSr3Io"
        TELEGRAM_CHAT_ID = "745125435"
        
        telegram = create_client(TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID, throttle_interval=2.0)
        
    except ImportError as e:
        TELEGRAM_ENABLED = False
        print(f"âš  Telegram ÑƒÐ²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ñ Ð½ÐµÐ´Ð¾ÑÑ‚ÑƒÐ¿Ð½Ñ‹: {e}")


# ===== Ð¤Ð£ÐÐšÐ¦Ð˜Ð˜ Ð˜Ð—Ð’Ð›Ð•Ð§Ð•ÐÐ˜Ð¯ Ð˜Ð— Ð˜ÐœÐ•ÐÐ˜ Ð¤ÐÐ™Ð›Ð =====

def extract_truck_from_filename(filename):
    """Ð˜Ð·Ð²Ð»ÐµÑ‡ÑŒ Ð½Ð¾Ð¼ÐµÑ€ Ð¼Ð°ÑˆÐ¸Ð½Ñ‹ Ð¸Ð· Ð¸Ð¼ÐµÐ½Ð¸ Ñ„Ð°Ð¹Ð»Ð°"""
    
    # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 1: checked_XXX (Ð¾ÑÐ½Ð¾Ð²Ð½Ð¾Ð¹ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚)
    patterns = [
        r'checked_(\d+)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            number = match.group(1)
            if len(number) <= 2:
                return f"GR-OO{number.zfill(2)}"
            else:
                return f"GR-OO{number}"
    
    # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 2: XXXX - ÐŸÐ¾ÑÑ‚Ð°Ð²Ñ‰Ð¸Ðº (Ð½Ð¾Ð²Ñ‹Ð¹ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚)
    # ÐŸÑ€Ð¸Ð¼ÐµÑ€Ñ‹: "1726 - AC Intern.pdf", "771 - TIP.pdf"
    simple_number_pattern = r'^(\d{2,4})\s*[-â€“]\s*'
    match = re.match(simple_number_pattern, filename)
    if match:
        number = match.group(1)
        if len(number) <= 2:
            return f"GR-OO{number.zfill(2)}"
        elif len(number) == 3:
            return f"GR-OO{number}"
        else:  # 4 digits
            return f"GR-OO{number}"
    
    # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 3: ÐŸÑ€ÑÐ¼Ñ‹Ðµ ÑƒÐºÐ°Ð·Ð°Ð½Ð¸Ñ (GR-OO, HH-AG, etc.)
    direct_patterns = [
        (r'GR[- ]?OO(\d+)', 'GR-OO'),
        (r'GR[- ]?(\d+)', 'GR-'),
        (r'WJQY4010', 'WJQY4010'),
        (r'OHAMX771', 'OHAMX771'),
        (r'HH[- ]?AG(\d+)', 'HH-AG'),
        (r'DE[- ]?FN(\d+)', 'DE-FN'),
    ]
    
    for pattern, prefix in direct_patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            if prefix in ['WJQY4010', 'OHAMX771']:
                return prefix
            number = match.group(1) if match.groups() else ''
            return f"{prefix}{number.zfill(3)}"
    
    return None


# ===== Ð¤Ð£ÐÐšÐ¦Ð˜Ð˜ ÐšÐ›ÐÐ¡Ð¡Ð˜Ð¤Ð˜ÐšÐÐ¦Ð˜Ð˜ ÐŸÐžÐ¡Ð¢ÐÐ’Ð©Ð˜ÐšÐžÐ’ =====

def identify_supplier(text):
    """
    Ð£Ð»ÑƒÑ‡ÑˆÐµÐ½Ð½Ñ‹Ð¹ ÐºÐ»Ð°ÑÑÐ¸Ñ„Ð¸ÐºÐ°Ñ‚Ð¾Ñ€ Ð¿Ð¾ÑÑ‚Ð°Ð²Ñ‰Ð¸ÐºÐ¾Ð²
    Ð’Ð¾Ð·Ð²Ñ€Ð°Ñ‰Ð°ÐµÑ‚ Ð½Ð°Ð·Ð²Ð°Ð½Ð¸Ðµ Ð¿Ð¾ÑÑ‚Ð°Ð²Ñ‰Ð¸ÐºÐ°
    """
    text_start = text[:1500].upper()
    text_upper = text.upper()
    
    # ÐŸÐ Ð˜ÐžÐ Ð˜Ð¢Ð•Ð¢ 1: Ð¢Ð¾Ñ‡Ð½Ñ‹Ðµ ÑÐ¾Ð²Ð¿Ð°Ð´ÐµÐ½Ð¸Ñ Ð² Ð½Ð°Ñ‡Ð°Ð»Ðµ Ð´Ð¾ÐºÑƒÐ¼ÐµÐ½Ñ‚Ð°
    
    # Vital Projekt
    if 'VITAL PROJEKT' in text_start:
        return 'Vital Projekt'
    
    # K&L
    elif 'K&L KFZ MEISTERBETRIEB' in text_start or 'K&L-KFZ' in text_start:
        return 'K&L'
    
    # Auto Compass Internal
    elif 'AUTO COMPASS' in text_start and ('KOPIE' in text or 'RANDERSWEIDE' in text):
        return 'Auto Compass (Internal)'
    
    # Scania External (Ð²Ð½ÐµÑˆÐ½Ð¸Ðµ ÑÑ‡ÐµÑ‚Ð°)
    elif '#SPLMINFO' in text_start and 'SCANIA' in text_start:
        return 'Scania External'
    
    # Ferronordic
    elif 'FERRONORDIC' in text_start:
        return 'Ferronordic'
    
    # HNS
    elif 'HNS NUTZFAHRZEUGE' in text_start or 'HNS SERVICE' in text_start:
        return 'HNS'
    
    # DEKRA
    elif 'DEKRA AUTOMOBIL' in text_start:
        return 'DEKRA'
    
    # SchÃ¼tt
    elif 'SCHÃœTT GMBH' in text_start or 'W. SCHÃœTT' in text_start:
        return 'SchÃ¼tt'
    
    # Volvo
    elif 'VOLVO GROUP TRUCKS' in text_start:
        return 'Volvo'
    
    # TIP Trailer
    elif 'TIP TRAILER SERVICES' in text_start or 'TRAILER SERVICES GERMANY' in text_start:
        return 'TIP'
    
    # Euromaster
    elif 'EUROMASTER GMBH' in text_start:
        return 'Euromaster'
    
    # Quick
    elif 'QUICK REIFEN' in text_start or 'REIFENDISCOUNT' in text_start:
        return 'Quick'
    
    # Sotecs
    elif 'SOTECS GMBH' in text_start:
        return 'Sotecs'
    
    # Express
    elif 'EXPRESS' in text_start and 'UAB GROO' in text:
        return 'Express'
    
    # Tankpool24
    elif 'TANKPOOL' in text_start:
        return 'Tankpool24'
    
    # ÐŸÐ Ð˜ÐžÐ Ð˜Ð¢Ð•Ð¢ 2: ÐŸÐ¾Ð¸ÑÐº Ð²Ð¾ Ð²ÑÑ‘Ð¼ Ð´Ð¾ÐºÑƒÐ¼ÐµÐ½Ñ‚Ðµ
    
    # MAN - Ñ‚Ð¾Ð»ÑŒÐºÐ¾ ÐµÑÐ»Ð¸ Ð´ÐµÐ¹ÑÑ‚Ð²Ð¸Ñ‚ÐµÐ»ÑŒÐ½Ð¾ Ð¾Ñ‚ MAN
    elif 'MAN TRUCK & BUS DEUTSCHLAND' in text_upper:
        return 'MAN'
    
    return 'Unknown'


# ===== Ð’Ð¡ÐŸÐžÐœÐžÐ“ÐÐ¢Ð•Ð›Ð¬ÐÐ«Ð• Ð¤Ð£ÐÐšÐ¦Ð˜Ð˜ =====


def extract_dekra(text, filename):
    """DEKRA Automobil GmbH - Hauptuntersuchung"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnung\s+Nr\.\s*:?\s*(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Rechnung_Nr\.\s*:?\s*(\d+)', text)
    
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', text)
        if date_match:
            date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
        else:
            return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð°
    truck_match = re.search(r'Kennzeichen[^\n]*:\s*([A-Z]{2}\s*[A-Z0-9]+\s*\d+)', text)
    if not truck_match:
        truck_match = re.search(r'KM-Stand\s+([A-Z]{2}\s*[A-Z0-9]+\s*\d+)', text, re.IGNORECASE)
    
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
        # Ð¤Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€Ð¾Ð²Ð°Ñ‚ÑŒ ÐºÐ°Ðº GR-OO1514
        if len(data['truck']) >= 4:
            data['truck'] = f"{data['truck'][:2]}-{data['truck'][2:]}"
    else:
        # ÐŸÐ¾Ð¿Ñ‹Ñ‚ÐºÐ° Ð¸Ð· Ð¸Ð¼ÐµÐ½Ð¸ Ñ„Ð°Ð¹Ð»Ð°
        data['truck'] = extract_truck_from_filename(filename) or ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    data['name'] = 'Hauptuntersuchung'
    data['amount'] = 1
    
    # Ð¦ÐµÐ½Ð°
    price_match = re.search(r'Nettopreis\s+[\d,]+\s+[\d,]+\s+([\d,]+)', text)
    if price_match:
        price_str = price_match.group(1).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð° - NETTO (Ð±ÐµÐ· ÐÐ”Ð¡)
    total_match = re.search(r'Nettobetrag\s+([\d,.]+)', text)
    if not total_match:
        # ÐÐ»ÑŒÑ‚ÐµÑ€Ð½Ð°Ñ‚Ð¸Ð²Ð½Ñ‹Ð¹ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½
        total_match = re.search(r'Gesamt[^\d]+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'DEKRA Automobil GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    
    return data


def extract_autocompass(text, filename):
    """Auto Compass Internal - внутренние счета
    ВАЖНО: Для Internal использует BRUTTO (Gesamt), для External - NETTO
    """
    data = {}
    
    # Номер счёта - улучшенные паттерны
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Interne\s+Rechnung\s+(\d+)', text)
    if not invoice_match:
        # Альтернативный паттерн для строки "Interne Rechnung 700293 Datum"
        invoice_match = re.search(r'Rechnung\s+(\d+)\s+Datum', text)
    
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата - улучшенные паттерны
    date_match = re.search(r'Datum\s+(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        date_match = re.search(r'Leistungsdatum:\s*(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        # Новый паттерн: "700293 Datum 08.10.2025"
        date_match = re.search(r'\d+\s+Datum\s+(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    
    if date_match:
        date_str = date_match.group(1)
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина - улучшенные паттерны
    truck_match = re.search(r'Kennzeichen\s+([A-Z]{2}-[A-Z]{2}\s*\d+)', text)
    if not truck_match:
        # Новый паттерн: "Kennzeichen Fahrgestell-Nr. ... HH-AG 1926"
        truck_match = re.search(r'Kennzeichen[^\n]*\n[^\n]*\n([A-Z]{2}-[A-Z]{2,4}\s*\d+)', text)
    if not truck_match:
        truck_match = re.search(r'([A-Z]{2}-[A-Z]{2,4}\s+\d+)', text)
    
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
    else:
        data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Описание
    data['name'] = 'Ремонтные работы'
    data['amount'] = 1
    data['price'] = 0.0
    
    # Определяем seller и buyer
    data['seller'] = 'Auto Compass GmbH'
    
    # Пытаемся найти buyer (Firma)
    buyer_match = re.search(r'Firma\s+(.+?)(?:\n|Randersweide)', text)
    if buyer_match:
        buyer = buyer_match.group(1).strip()
        data['buyer'] = buyer
    else:
        data['buyer'] = 'Auto Compass GmbH'  # По умолчанию
    
    # 🔴 КРИТИЧЕСКАЯ ЛОГИКА: Internal (BRUTTO) vs External (NETTO)
    is_internal = (data['seller'] == data['buyer'])
    
    if is_internal:
        # ⚠️ INTERNAL счёт: используем BRUTTO (Gesamt)
        # Паттерн 1: "Gesamt\n337,50 € 112,35 € 0.00 € 0.00 € 452,10 €"
        gesamt_match = re.search(r'Gesamt\s+(?:[\d,]+\s*€\s+){3,}([\d,]+)\s*€', text)
        if not gesamt_match:
            # Паттерн 2: "Gesamt ... € XXX,XX €" (последняя сумма)
            gesamt_match = re.search(r'Gesamt[^\n]*?([\d,]+)\s*€(?:\s|$)', text)
        if not gesamt_match:
            # Паттерн 3: Простой "Gesamt XXX,XX €"
            gesamt_match = re.search(r'Gesamt\s+([\d,.]+)\s*€', text)
        
        if gesamt_match:
            total_str = gesamt_match.group(1).replace('.', '').replace(',', '.')
            try:
                data['total_price'] = float(total_str)
            except:
                return None
        else:
            return None
    else:
        # ✅ EXTERNAL счёт: используем NETTO
        # Паттерн 1: Таблица "Lohn Material Fremdleistung Netto"
        netto_pattern = re.search(
            r'Lohn\s+Material\s+Fremdleistung\s+Netto\s+[\d,.]+\s*€\s+[\d,.]+\s*€\s+[\d,.]+\s*€\s+([\d,.]+)\s*€',
            text
        )
        if not netto_pattern:
            # Паттерн 2: 4 числа перед "Gesamt"
            netto_pattern = re.search(
                r'[\d,.]+\s*€\s+[\d,.]+\s*€\s+[\d,.]+\s*€\s+([\d,.]+)\s*€\s+Gesamt',
                text
            )
        if not netto_pattern:
            # Паттерн 3: Простой "Netto XXX,XX €"
            netto_pattern = re.search(r'Netto\s+([\d,.]+)\s*€', text)
        
        if netto_pattern:
            total_str = netto_pattern.group(1).replace('.', '').replace(',', '.')
            try:
                data['total_price'] = float(total_str)
            except:
                return None
        else:
            return None
    
    return data


def extract_vital_projekt(text):
    """Vital Projekt - ÑˆÐ¸Ð½Ñ‹ Ð¸ ÑƒÑÐ»ÑƒÐ³Ð¸"""
    data = {}
    
    invoice_match = re.search(r'Rechnungs-Nr\.:\s*(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    date_match = re.search(r'Hamburg,\s*den\s*(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
    else:
        date_match2 = re.search(r'Leistungsdatum/-Ort:(\d{2})-(\d{2})-(\d{4})', text)
        if date_match2:
            date_str = f"{date_match2.group(1)}.{date_match2.group(2)}.{date_match2.group(3)}"
        else:
            return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    truck_match = re.search(r'Kennzeichen/Fahrer:\s*([A-Z]{2}\s*[A-Z]{2}\s*\d+)', text)
    if not truck_match:
        return None
    
    truck_raw = truck_match.group(1)
    truck_clean = re.sub(r'\s+', '', truck_raw)
    if len(truck_clean) >= 4:
        data['truck'] = f"{truck_clean[:2]}-{truck_clean[2:]}"
    else:
        data['truck'] = truck_clean
    
    table_match = re.search(
        r'^\s*1\s+(\d+)\s+Stk\.\s+[^\s]+\s+(.+?)\s+\d+%\s+([\d,]+)\s*â‚¬',
        text, re.MULTILINE
    )
    
    if table_match:
        data['amount'] = int(table_match.group(1))
        description = table_match.group(2).strip()
        data['name'] = description[:50] if len(description) > 50 else description
        price_str = table_match.group(3).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['amount'] = 1
        data['name'] = "Ð¨Ð¸Ð½Ñ‹ Ð¸ ÑƒÑÐ»ÑƒÐ³Ð¸"
        data['price'] = 0.0
    
    summe_match = re.search(r'Summe\s+([\d,.]+)\s*â‚¬', text)
    if summe_match:
        total_str = summe_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Vital Projekt Inh.Vitalij Barth'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_ferronordic(text):
    """Ferronordic - ÑÐµÑ€Ð²Ð¸Ñ Volvo"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnung:\s*RE(\d+)-(\d+)', text)
    if invoice_match:
        data['invoice'] = f"RE{invoice_match.group(1)}-{invoice_match.group(2)}"
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'Vom:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð°
    truck_match = re.search(r'Kennzeichen:\s*DE\s*FN\s*(\d+)', text)
    if truck_match:
        data['truck'] = f"DE-FN{truck_match.group(1)}"
    else:
        truck_match2 = re.search(r'DE\s*FN\s*(\d+)', text)
        if truck_match2:
            data['truck'] = f"DE-FN{truck_match2.group(1)}"
        else:
            data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    desc_match = re.search(r'1\s+Monatstarif[^\n]+', text)
    if desc_match:
        data['name'] = desc_match.group(0)[:50]
    else:
        data['name'] = 'Wartung/Service Volvo'
    
    # ÐšÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾ Ð¸ Ñ†ÐµÐ½Ð°
    amount_match = re.search(r'1\s+Monatstarif[^\d]+(\d+,\d+)\s*â‚¬', text)
    if amount_match:
        data['amount'] = 1
        price_str = amount_match.group(1).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['amount'] = 1
        data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð° (Endsumme Ñ ÐÐ”Ð¡)
    total_match = re.search(r'Endsumme\s+([\d,.]+)\s*â‚¬', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Ferronordic Deutschland GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_hns(text):
    """HNS Nutzfahrzeuge Service"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnung\s*:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð°
    truck_match = re.search(r'Fahrzeug:\s*GR-OO\s*(\d+)', text)
    if truck_match:
        data['truck'] = f"GR-OO{truck_match.group(1)}"
    else:
        truck_match2 = re.search(r'GR-OO\s*(\d+)', text)
        if truck_match2:
            data['truck'] = f"GR-OO{truck_match2.group(1)}"
        else:
            data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ - Ð±ÐµÑ€Ñ‘Ð¼ Ð¿ÐµÑ€Ð²ÑƒÑŽ Ñ€Ð°Ð±Ð¾Ñ‚Ñƒ
    desc_match = re.search(r'HU\s+Begleitung|HU\s+PrÃ¼fung', text)
    if desc_match:
        data['name'] = desc_match.group(0)
    else:
        data['name'] = 'Werkstattleistungen'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð° (Endbetrag)
    total_match = re.search(r'Endbetrag\s*:\s*([\d,.]+)\s*â‚¬', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'HNS Nutzfahrzeuge Service GmbH'
    data['buyer'] = 'Groo GmbH' if 'Groo GmbH' in text else 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_tip(text):
    """TIP Trailer Services - Ð°Ñ€ÐµÐ½Ð´Ð° Ð¿Ñ€Ð¸Ñ†ÐµÐ¿Ð¾Ð²"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnungsnr\.\s*U71/(\d+)', text)
    if invoice_match:
        data['invoice'] = f"U71/{invoice_match.group(1)}"
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})/(\d{2})/(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð° (Ð¿Ñ€Ð¸Ñ†ÐµÐ¿)
    truck_match = re.search(r'Kennzeichen:\s*([A-Z0-9]+)', text)
    if truck_match:
        data['truck'] = truck_match.group(1)
    else:
        # ÐÐ»ÑŒÑ‚ÐµÑ€Ð½Ð°Ñ‚Ð¸Ð²Ð½Ñ‹Ð¹ Ð¿Ð¾Ð¸ÑÐº
        truck_match2 = re.search(r'Flotten-Nr\.\s*(\d+)', text)
        if truck_match2:
            data['truck'] = truck_match2.group(1)
        else:
            data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    data['name'] = 'ÐÑ€ÐµÐ½Ð´Ð° Ð¿Ñ€Ð¸Ñ†ÐµÐ¿Ð°'
    
    data['amount'] = 1
    
    # Ð¦ÐµÐ½Ð° Ð·Ð° Ð¼ÐµÑÑÑ†
    price_match = re.search(r'(\d+,\d+)\s+EUR', text)
    if price_match:
        price_str = price_match.group(1).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°
    total_match = re.search(r'Gesamtbetrag\s+EUR\s+([\d,.]+)', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'TIP Trailer Services Germany GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_euromaster(text):
    """Euromaster - ÑˆÐ¸Ð½Ñ‹ Ð¸ Ð¾Ð±ÑÐ»ÑƒÐ¶Ð¸Ð²Ð°Ð½Ð¸Ðµ"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'Datum\s*:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð°
    truck_match = re.search(r'KFZ-KENNZEICHEN:\s*([A-Z]{2}-[A-Z0-9]+)', text)
    if truck_match:
        data['truck'] = truck_match.group(1)
    else:
        data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ - Ð¿ÐµÑ€Ð²Ð°Ñ Ð¿Ð¾Ð·Ð¸Ñ†Ð¸Ñ
    desc_match = re.search(r'EUROMASTER\s+\d+\s+[^\n]+', text)
    if desc_match:
        data['name'] = desc_match.group(0)[:50]
    else:
        data['name'] = 'Ð¨Ð¸Ð½Ñ‹ Ð¸ Ð¾Ð±ÑÐ»ÑƒÐ¶Ð¸Ð²Ð°Ð½Ð¸Ðµ'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð° (Nettowert Ð¸Ð»Ð¸ Bruttowert)
    total_match = re.search(r'Nettowert\s+Bruttowert\s+([\d,.]+)\s+([\d,.]+)', text)
    if total_match:
        netto_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(netto_str)
    else:
        # ÐÐ»ÑŒÑ‚ÐµÑ€Ð½Ð°Ñ‚Ð¸Ð²Ð½Ñ‹Ð¹ Ð¿Ð¾Ð¸ÑÐº
        total_match2 = re.search(r'([\d,.]+)\s+EUR\s+zum', text)
        if total_match2:
            total_str = total_match2.group(1).replace('.', '').replace(',', '.')
            data['total_price'] = float(total_str)
        else:
            return None
    
    data['seller'] = 'Euromaster GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data

# ===== ÐŸÐ ÐžÐ”ÐžÐ›Ð–Ð•ÐÐ˜Ð• Ð¤Ð£ÐÐšÐ¦Ð˜Ð™ Ð˜Ð—Ð’Ð›Ð•Ð§Ð•ÐÐ˜Ð¯ =====


def extract_man(text):
    """MAN Truck & Bus Deutschland"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnungsnummer:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð° - Ð½Ðµ Ð²ÑÐµÐ³Ð´Ð° ÐµÑÑ‚ÑŒ
    data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    desc_match = re.search(r'Job\s+\d+:\s+([^\n]+)', text)
    if desc_match:
        data['name'] = desc_match.group(1)[:50]
    else:
        data['name'] = 'Ð ÐµÐ¼Ð¾Ð½Ñ‚ MAN'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð° - Ñ€Ð°Ð·Ð»Ð¸Ñ‡Ð½Ñ‹Ðµ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½Ñ‹
    total_match = re.search(r'NETTO\s+([\d,.]+)\s+EUR', text, re.IGNORECASE)
    if not total_match:
        total_match = re.search(r'Nettobetrag:\s+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'MAN Truck & Bus Deutschland GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_schutt(text):
    """W. SchÃ¼tt GmbH"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnung\s*:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð°
    truck_match = re.search(r'Kennzeichen\s*:\s*([A-Z]{2}-[A-Z]{2}\s*\d+)', text)
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
    else:
        data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    data['name'] = 'Ð ÐµÐ¼Ð¾Ð½Ñ‚Ð½Ñ‹Ðµ Ñ€Ð°Ð±Ð¾Ñ‚Ñ‹'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°
    total_match = re.search(r'Netto\s+([\d,.]+)\s*â‚¬', text)
    if not total_match:
        total_match = re.search(r'Gesamt\s+([\d,.]+)\s*â‚¬', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'W. SchÃ¼tt GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_volvo(text):
    """Volvo Group Trucks Service"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnungs-Nr\.\.\s*:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'Auftr\.-Datum\.\.\s*:\s*(\d{2})\.(\d{2})\.(\d{2})', text)
    if date_match:
        year = f"20{date_match.group(3)}"
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{year}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð°
    truck_match = re.search(r'Kennzeichen\.\.\s*:\s*([A-Z]{2}-[A-Z0-9]+)', text)
    if truck_match:
        data['truck'] = truck_match.group(1)
    else:
        data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    desc_match = re.search(r'PFX\s+Ersatzteilnummer\s+[^\n]+', text)
    if desc_match:
        data['name'] = 'Ð—Ð°Ð¿Ñ‡Ð°ÑÑ‚Ð¸ Volvo'
    else:
        data['name'] = 'Ð ÐµÐ¼Ð¾Ð½Ñ‚ Volvo'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°
    total_match = re.search(r'Gesamt\s+EUR\s+([\d,.]+)', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Volvo Group Trucks Service Nord GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_sotecs(text):
    """Sotecs GmbH"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnungsnummer:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    desc_match = re.search(r'Einbau von\s+([^\n]+)', text)
    if desc_match:
        data['name'] = desc_match.group(1)[:50]
    else:
        data['name'] = 'Ð£ÑÑ‚Ð°Ð½Ð¾Ð²ÐºÐ° Ð¾Ð±Ð¾Ñ€ÑƒÐ´Ð¾Ð²Ð°Ð½Ð¸Ñ'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°
    total_match = re.search(r'Total\s+\(netto\)\s+([\d,.]+)\s+EUR', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Sotecs GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_express(text):
    """Express Service"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'Datum\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð°
    truck_match = re.search(r'NGZ\s+(\d+)', text)
    if truck_match:
        data['truck'] = f"NGZ{truck_match.group(1)}"
    else:
        data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    data['name'] = 'Wartung und Filter'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°
    total_match = re.search(r'Gesamtbetrag\s+([\d,.]+)\s*â‚¬', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Express Service GmbH'
    data['buyer'] = 'UAB Groo Transport'
    data['interne_rechnung'] = ''
    
    return data



def extract_kl(text):
    """K&L Kfz Meisterbetrieb"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnungsnummer:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ - Ð¸Ð· Ñ‚Ð°Ð±Ð»Ð¸Ñ†Ñ‹
    desc_match = re.search(r'MEHRFACHKUPPLUNG|GETRIEBE', text)
    if desc_match:
        data['name'] = desc_match.group(0)
    else:
        data['name'] = 'Ð—Ð°Ð¿Ñ‡Ð°ÑÑ‚Ð¸ Ð¸ Ñ€ÐµÐ¼Ð¾Ð½Ñ‚'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°
    total_match = re.search(r'Gesamtsumme netto\s+([\d,.]+)\s*EUR', text)
    if not total_match:
        total_match = re.search(r'Total\s+\(netto\)\s+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'K&L Kfz Meisterbetrieb GmbH'
    data['buyer'] = 'GROO GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_quick(text):
    """Quick Reifendicount"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'Rechnung\s+Nr\.\s*:\s*([A-Z0-9]+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð°
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{2})', text)
    if date_match:
        year = f"20{date_match.group(3)}"
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{year}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð°
    truck_match = re.search(r'GR-(\d+)', text)
    if truck_match:
        data['truck'] = f"GR-{truck_match.group(1)}"
    else:
        data['truck'] = ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    data['name'] = 'Ð¨Ð¸Ð½Ñ‹'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°
    total_match = re.search(r'Gesamtbetrag\s+([\d,.]+)\s+EUR', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Quick Reifendicount'
    data['buyer'] = 'Andreas Groo Handel & Transporte'
    data['interne_rechnung'] = ''
    
    return data



def extract_tankpool24(text):
    """Tankpool24 - Ñ‚Ð¾Ð¿Ð»Ð¸Ð²Ð½Ñ‹Ðµ ÐºÐ°Ñ€Ñ‚Ñ‹"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð°
    invoice_match = re.search(r'(\d{7})', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð° - Ð¾Ð±Ñ‹Ñ‡Ð½Ð¾ Ð² Ð½Ð°Ñ‡Ð°Ð»Ðµ
    date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    data['truck'] = ''
    data['name'] = 'Ð¢Ð¾Ð¿Ð»Ð¸Ð²Ð¾'
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð°
    total_match = re.search(r'([\d,.]+)\s+EUR', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Tankpool24 International GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data


# ===== ÐšÐ›ÐÐ¡Ð¡Ð˜Ð¤Ð˜ÐšÐÐ¢ÐžÐ  =====


def extract_scania(text, filename):
    """Scania - Ð²Ð½ÐµÑˆÐ½Ð¸Ðµ ÑÑ‡ÐµÑ‚Ð° Ð¾Ñ‚ Scania"""
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð° - Ð¸Ñ‰ÐµÐ¼ SCHWL (ÑƒÐ»ÑƒÑ‡ÑˆÐµÐ½Ð½Ñ‹Ðµ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½Ñ‹)
    invoice_match = re.search(r'SCHWL(\d+)', text)
    if not invoice_match:
        # ÐÐ»ÑŒÑ‚ÐµÑ€Ð½Ð°Ñ‚Ð¸Ð²Ð½Ñ‹Ð¹ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½ Ñ Ð¿Ñ€Ð¾Ð±ÐµÐ»Ð°Ð¼Ð¸
        invoice_match = re.search(r'SCHWL\s*(\d+)', text)
    if not invoice_match:
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ Ð¸Ð· splminfo ÑÑ‚Ñ€Ð¾ÐºÐ¸
        invoice_match = re.search(r'SCH_SCHWL(\d+)_', text)
    
    if invoice_match:
        data['invoice'] = f"SCHWL{invoice_match.group(1)}"
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð° - ÑƒÐ»ÑƒÑ‡ÑˆÐµÐ½Ð½Ñ‹Ðµ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½Ñ‹
    # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 1: "RE-DATUM ... 20.10.25"
    date_match = re.search(r'RE-DATUM[^\n]*?(\d{2}\.\d{2}\.\d{2,4})', text)
    if not date_match:
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 2: ÐžÐ±Ñ‹Ñ‡Ð½Ñ‹Ð¹ Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚ Ð´Ð°Ñ‚Ñ‹
        date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{2,4})', text)
    
    if date_match:
        if len(date_match.groups()) == 1:
            # Ð¤Ð¾Ñ€Ð¼Ð°Ñ‚ DD.MM.YY Ð¸Ð»Ð¸ DD.MM.YYYY
            date_str = date_match.group(1)
            parts = date_str.split('.')
            if len(parts[2]) == 2:
                # ÐŸÑ€ÐµÐ¾Ð±Ñ€Ð°Ð·Ð¾Ð²Ð°Ñ‚ÑŒ YY Ð² YYYY
                year = f"20{parts[2]}"
                date_str = f"{parts[0]}.{parts[1]}.{year}"
        else:
            # Ð¤Ð¾Ñ€Ð¼Ð°Ñ‚ Ð¸Ð· Ð³Ñ€ÑƒÐ¿Ð¿
            day = date_match.group(1)
            month = date_match.group(2)
            year = date_match.group(3)
            if len(year) == 2:
                year = f"20{year}"
            date_str = f"{day}.{month}.{year}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð° - ÑƒÐ»ÑƒÑ‡ÑˆÐµÐ½Ð½Ñ‹Ðµ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½Ñ‹
    # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 1: "AMTL.KENNZ: GR-OO 1726"
    truck_match = re.search(r'AMTL\.KENNZ:\s*([A-Z]{2}-[A-Z]{2,4}\s*\d+)', text)
    if not truck_match:
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 2: "Kennzeichen ... GR-OO 1726"
        truck_match = re.search(r'Kennzeichen[:\s]+([A-Z]{2}-[A-Z]{2,4}\s*\d+)', text)
    if not truck_match:
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 3: ÐžÐ±Ñ‰Ð¸Ð¹ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½ Ð½Ð¾Ð¼ÐµÑ€Ð°
        truck_match = re.search(r'([A-Z]{2}\s*[A-Z]{2}\s*\d+)', text)
    
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
        # Ð¤Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€Ð¾Ð²Ð°Ñ‚ÑŒ ÐºÐ°Ðº GR-OO1726
        if len(data['truck']) >= 4 and '-' not in data['truck']:
            data['truck'] = f"{data['truck'][:2]}-{data['truck'][2:]}"
    else:
        data['truck'] = extract_truck_from_filename(filename) or ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    data['name'] = 'Scania service'
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ±Ñ‰Ð°Ñ ÑÑƒÐ¼Ð¼Ð° - ÐŸÐ Ð˜ÐžÐ Ð˜Ð¢Ð•Ð¢ NETTO!
    # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 1: "NETTOBETRAG ... EUR XXX,XX" (Ð’Ð«Ð¡Ð¨Ð˜Ð™ ÐŸÐ Ð˜ÐžÐ Ð˜Ð¢Ð•Ð¢)
    total_match = re.search(r'NETTOBETRAG[^\d]+EUR\s+([\d,.]+)', text)
    if not total_match:
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 2: "LOHN NETTO: XXX,XX"
        total_match = re.search(r'LOHN\s+NETTO:\s*([\d,.]+)', text)
    if not total_match:
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 3: "SUMME TEILE: XXX,XX"
        total_match = re.search(r'SUMME\s+TEILE:\s*([\d,]+)', text)
    if not total_match:
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 4: "SUMME: XX PE XXX,XX"
        total_match = re.search(r'SUMME:\s*\d+\s*PE\s*([\d,]+)', text)
    if not total_match:
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 5: "SUMME EUR XXX,XX" (Ð¼Ð¾Ð¶ÐµÑ‚ Ð±Ñ‹Ñ‚ÑŒ BRUTTO - Ð¾ÑÑ‚Ð¾Ñ€Ð¾Ð¶Ð½Ð¾)
        total_match = re.search(r'SUMME\s+EUR\s+([\d,.]+)', text)
    if not total_match:
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 6: ÐžÐ±Ñ‰Ð¸Ð¹ Gesamt
        total_match = re.search(r'Gesamt[^\d]+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        try:
            data['total_price'] = float(total_str)
        except:
            return None
    else:
        return None
    
    data['seller'] = 'Scania'
    data['buyer'] = 'Auto Compass GmbH'
    
    return data


def extract_universal(text, supplier, filename):
    """
    Ð£Ð½Ð¸Ð²ÐµÑ€ÑÐ°Ð»ÑŒÐ½Ð¾Ðµ Ð¸Ð·Ð²Ð»ÐµÑ‡ÐµÐ½Ð¸Ðµ Ð´Ð»Ñ Ð¿Ð¾ÑÑ‚Ð°Ð²Ñ‰Ð¸ÐºÐ¾Ð² Ð±ÐµÐ· ÑÐ¿ÐµÑ†Ð¸Ð°Ð»ÑŒÐ½Ð¾Ð¹ Ñ„ÑƒÐ½ÐºÑ†Ð¸Ð¸
    """
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡ÐµÑ‚Ð° (ÑƒÐ½Ð¸Ð²ÐµÑ€ÑÐ°Ð»ÑŒÐ½Ð¾Ðµ)
    invoice_patterns = [
        r'Rechnung[s-]?[Nn]r\.?\s*:?\s*(\d+)',
        r'Invoice\s+(?:No\.?|Number)\s*:?\s*(\d+)',
        r'Rg\.?\s*-?\s*Nr\.?\s*:?\s*(\d+)',
        r'Rechnungs-Nr\.:\s*(\d+)',
    ]
    
    for pattern in invoice_patterns:
        match = re.search(pattern, text)
        if match:
            data['invoice'] = match.group(1)
            break
    
    if not data.get('invoice'):
        return None
    
    # Ð”Ð°Ñ‚Ð° (ÑƒÐ½Ð¸Ð²ÐµÑ€ÑÐ°Ð»ÑŒÐ½Ð¾Ðµ)
    date_patterns = [
        r'(\d{2}\.\d{2}\.\d{4})',
        r'(\d{2})-(\d{2})-(\d{4})',
    ]
    
    for pattern in date_patterns:
        match = re.search(pattern, text)
        if match:
            if len(match.groups()) == 1:
                date_str = match.group(1)
            else:
                date_str = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
            
            try:
                date_obj = datetime.strptime(date_str.replace('-', '.'), '%d.%m.%Y')
                data['date'] = date_obj.strftime('%d.%m.%Y')
                data['month'] = date_obj.month
                data['week'] = date_obj.isocalendar()[1]
                break
            except:
                continue
    
    if not data.get('date'):
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð°
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Ð¡ÑƒÐ¼Ð¼Ð° - ÐŸÐ Ð˜ÐžÐ Ð˜Ð¢Ð•Ð¢ NETTO Ð´Ð»Ñ Ð²Ð½ÐµÑˆÐ½Ð¸Ñ… ÑÑ‡ÐµÑ‚Ð¾Ð²!
    amount_patterns = [
        r'Nettobetrag\s*([\d,\.]+)',              # 1. Nettobetrag (Ð’Ð«Ð¡Ð¨Ð˜Ð™)
        r'Netto\s+([\d,\.]+)',                    # 2. Netto
        r'SUMME\s+(?:NETTO|netto)\s*:?\s*([\d,\.]+)',  # 3. SUMME NETTO
        r'Summe\s+netto\s*:?\s*([\d,\.]+)',       # 4. Summe netto
        r'Gesamtbetrag\s+(?:EUR)?\s*([\d,\.]+)',  # 5. Gesamtbetrag (BRUTTO)
        r'Summe\s+brutto\s*:?\s*([\d,\.]+)',      # 6. Summe brutto
        r'Total\s*:?\s*([\d,\.]+)',               # 7. Total
        r'Gesamt\s*:?\s*([\d,\.]+)',              # 8. Gesamt (Ð¼Ð¾Ð¶ÐµÑ‚ Ð±Ñ‹Ñ‚ÑŒ BRUTTO)
    ]
    
    for pattern in amount_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            amount_str = match.group(1).replace(',', '.').replace(' ', '')
            try:
                data['amount'] = 1
                data['price'] = float(amount_str)
                data['total_price'] = float(amount_str)
                break
            except:
                continue
    
    if not data.get('total_price'):
        return None
    
    data['seller'] = supplier
    data['buyer'] = 'Auto Compass GmbH'
    data['name'] = ''
    
    return data


# ===== Ð ÐžÐ£Ð¢Ð•Ð  Ð˜Ð—Ð’Ð›Ð•Ð§Ð•ÐÐ˜Ð¯ =====

def extract_data_by_supplier(text, supplier, filename):
    """
    Роутер: выбрать правильную функцию извлечения
    v7.0: Поддержка 16 поставщиков
    """
    # Специализированные extractors
    extractors = {
        'DEKRA': lambda t: extract_dekra(t, filename),
        'Auto Compass (Internal)': lambda t: extract_autocompass(t, filename),
        'Scania': lambda t: extract_scania(t, filename),
        'Vital Projekt': extract_vital_projekt,
        'Ferronordic': extract_ferronordic,
        'HNS': extract_hns,
        'TIP': extract_tip,
        'Euromaster': extract_euromaster,
        'MAN': extract_man,
        'Schütt': extract_schutt,
        'Volvo': extract_volvo,
        'Sotecs': extract_sotecs,
        'Express': extract_express,
        'K&L': extract_kl,
        'Quick': extract_quick,
        'Tankpool24': extract_tankpool24,
    }
    
    # Проверяем специализированные
    if supplier in extractors:
        result = extractors[supplier](text)
        if result:
            return result
    
    # Fallback: универсальное извлечение
    return extract_universal(text, supplier, filename)


def check_invoice_exists(invoice_number, excel_file):
    """ÐŸÑ€Ð¾Ð²ÐµÑ€Ð¸Ñ‚ÑŒ ÑÑƒÑ‰ÐµÑÑ‚Ð²ÑƒÐµÑ‚ Ð»Ð¸ ÑÑ‡ÐµÑ‚ Ð² Excel"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[8] == str(invoice_number):  # ÐšÐ¾Ð»Ð¾Ð½ÐºÐ° I (Invoice)
                wb.close()
                # ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ Ð´Ð°Ñ‚Ñƒ Ð¾Ñ€Ð¸Ð³Ð¸Ð½Ð°Ð»Ð°
                for r in ws.iter_rows(min_row=2):
                    if str(r[8].value) == str(invoice_number):
                        date = r[3].value  # ÐšÐ¾Ð»Ð¾Ð½ÐºÐ° D (Date)
                        wb.close()
                        return True, date
                wb.close()
                return True, None
        
        wb.close()
        return False, None
    except Exception as e:
        print(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ñ€Ð¾Ð²ÐµÑ€ÐºÐ¸ Ð´ÑƒÐ±Ð»Ð¸ÐºÐ°Ñ‚Ð°: {e}")
        return False, None


def add_to_excel(data, excel_file):
    """Ð”Ð¾Ð±Ð°Ð²Ð¸Ñ‚ÑŒ Ð´Ð°Ð½Ð½Ñ‹Ðµ Ð² Excel"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # ÐÐ°Ð¹Ñ‚Ð¸ Ð¿ÐµÑ€Ð²ÑƒÑŽ Ð¿ÑƒÑÑ‚ÑƒÑŽ ÑÑ‚Ñ€Ð¾ÐºÑƒ
        row_num = 2
        while ws[f'A{row_num}'].value is not None:
            row_num += 1
        
        # Ð—Ð°Ð¿Ð¾Ð»Ð½Ð¸Ñ‚ÑŒ Ð´Ð°Ð½Ð½Ñ‹Ðµ
        ws[f'A{row_num}'] = data.get('month', 0)
        ws[f'B{row_num}'] = data.get('week', 0)
        ws[f'C{row_num}'] = data.get('truck', '')
        ws[f'D{row_num}'] = data.get('date', '')
        ws[f'E{row_num}'] = data.get('name', '')
        ws[f'F{row_num}'] = data.get('amount', 1)
        ws[f'G{row_num}'] = data.get('price', 0)
        ws[f'H{row_num}'] = data.get('total_price', 0)
        ws[f'I{row_num}'] = data.get('invoice', '')
        ws[f'J{row_num}'] = data.get('seller', '')
        ws[f'K{row_num}'] = data.get('buyer', 'Auto Compass GmbH')
        
        wb.save(excel_file)
        wb.close()
        
        return row_num
        
    except Exception as e:
        print(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð´Ð¾Ð±Ð°Ð²Ð»ÐµÐ½Ð¸Ñ Ð² Excel: {e}")
        return None


# ===== Ð¤Ð£ÐÐšÐ¦Ð˜Ð˜ Ð ÐÐ‘ÐžÐ¢Ð« Ð¡ Ð¤ÐÐ™Ð›ÐÐœÐ˜ =====

def rename_and_move_file(source_path, destination_folder, prefix=''):
    """ÐŸÐµÑ€ÐµÐ¸Ð¼ÐµÐ½Ð¾Ð²Ð°Ñ‚ÑŒ Ð¸ Ð¿ÐµÑ€ÐµÐ¼ÐµÑÑ‚Ð¸Ñ‚ÑŒ Ñ„Ð°Ð¹Ð»"""
    try:
        filename = os.path.basename(source_path)
        
        # Ð£Ð±Ñ€Ð°Ñ‚ÑŒ ÑÑ‚Ð°Ñ€Ñ‹Ðµ Ð¿Ñ€ÐµÑ„Ð¸ÐºÑÑ‹
        clean_filename = filename
        for old_prefix in ['checked_', 'manual_', 'error_', 'duplicate_']:
            if clean_filename.startswith(old_prefix):
                clean_filename = clean_filename[len(old_prefix):]
        
        # Ð”Ð¾Ð±Ð°Ð²Ð¸Ñ‚ÑŒ Ð½Ð¾Ð²Ñ‹Ð¹ Ð¿Ñ€ÐµÑ„Ð¸ÐºÑ
        new_filename = f"{prefix}{clean_filename}"
        destination_path = os.path.join(destination_folder, new_filename)
        
        # Ð•ÑÐ»Ð¸ Ñ„Ð°Ð¹Ð» ÑƒÐ¶Ðµ ÑÑƒÑ‰ÐµÑÑ‚Ð²ÑƒÐµÑ‚, Ð´Ð¾Ð±Ð°Ð²Ð¸Ñ‚ÑŒ timestamp
        if os.path.exists(destination_path):
            name, ext = os.path.splitext(new_filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            new_filename = f"{name}_{timestamp}{ext}"
            destination_path = os.path.join(destination_folder, new_filename)
        
        shutil.move(source_path, destination_path)
        return True, new_filename, None
        
    except Exception as e:
        return False, None, str(e)


def log_processing(message):
    """Ð›Ð¾Ð³Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð¸Ðµ Ð² Ñ„Ð°Ð¹Ð»"""
    try:
        os.makedirs(LOG_FOLDER, exist_ok=True)
        log_file = os.path.join(LOG_FOLDER, f"processing_{datetime.now().strftime('%Y%m%d')}.log")
        
        with open(log_file, 'a', encoding='utf-8') as f:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            f.write(f"[{timestamp}] {message}\n")
    except Exception as e:
        print(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð»Ð¾Ð³Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð¸Ñ: {e}")


# ===== Ð“Ð›ÐÐ’ÐÐÐ¯ Ð¤Ð£ÐÐšÐ¦Ð˜Ð¯ =====

def process_all_pdfs():
    """
    ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ð°Ñ‚ÑŒ Ð²ÑÐµ PDF Ñ„Ð°Ð¹Ð»Ñ‹ Ñ Ð´ÐµÑ‚Ð°Ð»ÑŒÐ½Ñ‹Ð¼Ð¸ Telegram ÑƒÐ²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸ÑÐ¼Ð¸
    """
    
    print("\nâ•”â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•—")
    print("â•‘      PDF PROCESSOR v6.0 - Ð£Ð›Ð£Ð§Ð¨Ð•ÐÐÐ«Ð• ÐŸÐÐ¢Ð¢Ð•Ð ÐÐ«               â•‘")
    print("â•‘      Ð˜ÑÐ¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¾: Auto Compass + Scania manual Ñ„Ð°Ð¹Ð»Ñ‹         â•‘")
    print("â•šâ•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•\n")
    
    # Ð¡Ð¾Ð·Ð´Ð°Ñ‚ÑŒ Ð¿Ð°Ð¿ÐºÐ¸ ÐµÑÐ»Ð¸ Ð½Ðµ ÑÑƒÑ‰ÐµÑÑ‚Ð²ÑƒÑŽÑ‚
    os.makedirs(MANUAL_FOLDER, exist_ok=True)
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)
    os.makedirs(LOG_FOLDER, exist_ok=True)
    
    # ÐŸÐ¾Ð»ÑƒÑ‡Ð¸Ñ‚ÑŒ ÑÐ¿Ð¸ÑÐ¾Ðº Ñ„Ð°Ð¹Ð»Ð¾Ð²
    try:
        pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    except Exception as e:
        print(f"âŒ ÐžÑˆÐ¸Ð±ÐºÐ° Ð´Ð¾ÑÑ‚ÑƒÐ¿Ð° Ðº Ð¿Ð°Ð¿ÐºÐµ: {e}")
        return None
    
    if not pdf_files:
        print("ðŸ“­ ÐÐµÑ‚ Ñ„Ð°Ð¹Ð»Ð¾Ð² Ð´Ð»Ñ Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÐ¸")
        return {'total': 0, 'processed': 0, 'duplicates': 0, 'manual': 0}
    
    total_files = len(pdf_files)
    processed_count = 0
    duplicate_count = 0
    error_count = 0
    
    # Ð”Ð»Ñ ÑÐ²Ð¾Ð´ÐºÐ¸
    processed_files = []
    duplicate_files = []
    manual_files = []
    
    supplier_stats = defaultdict(int)
    
    log_processing("="*80)
    log_processing(f"ÐÐÐ§ÐÐ›Ðž ÐžÐ‘Ð ÐÐ‘ÐžÐ¢ÐšÐ˜ v4.0: {total_files} Ñ„Ð°Ð¹Ð»Ð¾Ð²")
    log_processing("="*80)
    
    print(f"\nðŸ“Š ÐÐ°Ð¹Ð´ÐµÐ½Ð¾ Ñ„Ð°Ð¹Ð»Ð¾Ð²: {total_files}")
    print("="*80)
    
    # ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÐ° ÐºÐ°Ð¶Ð´Ð¾Ð³Ð¾ Ñ„Ð°Ð¹Ð»Ð°
    for index, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        
        print(f"\n[{index}/{total_files}] {pdf_file}")
        
        # Ð£Ð²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ðµ Ð¾ Ð½Ð°Ñ‡Ð°Ð»Ðµ Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÐ¸
        if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
            telegram.notify_processing_start(pdf_file)
        
        try:
            # ===== Ð¨ÐÐ“ 1: Ð˜Ð·Ð²Ð»ÐµÑ‡ÑŒ Ñ‚ÐµÐºÑÑ‚ =====
            full_text = ''
            
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + '\n'
            
            if not full_text.strip():
                print(f"  âš  PDF Ð¿ÑƒÑÑ‚Ð¾Ð¹")
                success, new_name, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                
                if success:
                    print(f"  â†’ manual/{new_name}")
                    log_processing(f"MANUAL (Ð¿ÑƒÑÑ‚Ð¾Ð¹): {pdf_file} -> {new_name}")
                    manual_files.append(pdf_file)
                    
                    # Ð£Ð²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ðµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, "PDF Ð¿ÑƒÑÑ‚Ð¾Ð¹")
                
                error_count += 1
                continue
            
            # ===== Ð¨ÐÐ“ 2: ÐšÐ»Ð°ÑÑÐ¸Ñ„Ð¸Ñ†Ð¸Ñ€Ð¾Ð²Ð°Ñ‚ÑŒ =====
            supplier = identify_supplier(full_text)
            print(f"  ÐŸÐ¾ÑÑ‚Ð°Ð²Ñ‰Ð¸Ðº: {supplier}")
            supplier_stats[supplier] += 1
            
            # ===== Ð¨ÐÐ“ 3: Ð˜Ð·Ð²Ð»ÐµÑ‡ÑŒ Ð´Ð°Ð½Ð½Ñ‹Ðµ =====
            data = extract_data_by_supplier(full_text, supplier, pdf_file)
            
            if not data:
                print(f"  âŒ ÐÐµ ÑƒÐ´Ð°Ð»Ð¾ÑÑŒ Ð¸Ð·Ð²Ð»ÐµÑ‡ÑŒ Ð´Ð°Ð½Ð½Ñ‹Ðµ")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'manual_'
                )
                
                if success:
                    print(f"  â†’ manual/{new_name}")
                    log_processing(f"MANUAL ({supplier}): {pdf_file} -> {new_name}")
                    manual_files.append(pdf_file)
                    
                    # Ð£Ð²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ðµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, "ÐÐµ ÑƒÐ´Ð°Ð»Ð¾ÑÑŒ Ð¸Ð·Ð²Ð»ÐµÑ‡ÑŒ Ð´Ð°Ð½Ð½Ñ‹Ðµ", supplier)
                
                error_count += 1
                continue
            
            if data.get('truck'):
                print(f"  ÐœÐ°ÑˆÐ¸Ð½Ð°: {data['truck']}")
            
            print(f"  Ð¡Ñ‡ÐµÑ‚: {data.get('invoice', 'N/A')}, Ð”Ð°Ñ‚Ð°: {data.get('date', 'N/A')}, Ð¡ÑƒÐ¼Ð¼Ð°: {data.get('total_price', 0):.2f} EUR")
            
            # ===== Ð¨ÐÐ“ 4: ÐŸÑ€Ð¾Ð²ÐµÑ€Ð¸Ñ‚ÑŒ Ð´ÑƒÐ±Ð»Ð¸ÐºÐ°Ñ‚ =====
            is_duplicate, original_date = check_invoice_exists(data['invoice'], EXCEL_FILE)
            
            if is_duplicate:
                print(f"  âš  Ð”Ð£Ð‘Ð›Ð˜ÐšÐÐ¢! Ð¡Ñ‡ÐµÑ‚ {data['invoice']} ÑƒÐ¶Ðµ ÑÑƒÑ‰ÐµÑÑ‚Ð²ÑƒÐµÑ‚")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'duplicate_'
                )
                
                if success:
                    print(f"  â†’ processed/{new_name}")
                    log_processing(f"DUPLICATE: {pdf_file} -> {new_name}")
                    duplicate_files.append(pdf_file)
                    
                    # Ð£Ð²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ðµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_duplicate(pdf_file, data['invoice'], original_date)
                
                duplicate_count += 1
                continue
            
            # ===== Ð¨ÐÐ“ 5: Ð”Ð¾Ð±Ð°Ð²Ð¸Ñ‚ÑŒ Ð² Excel =====
            excel_row = add_to_excel(data, EXCEL_FILE)
            
            if excel_row:
                print(f"  âœ“ Ð”Ð¾Ð±Ð°Ð²Ð»ÐµÐ½ Ð² Excel: ÑÑ‚Ñ€Ð¾ÐºÐ° {excel_row}")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'checked_'
                )
                
                if success:
                    print(f"  â†’ processed/{new_name}")
                    log_processing(f"SUCCESS: {pdf_file} -> {new_name}, Excel row {excel_row}")
                    processed_files.append(pdf_file)
                    
                    # Ð£Ð²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ðµ Ð¾Ð± ÑƒÑÐ¿ÐµÑ…Ðµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_success(data, pdf_file, excel_row)
                
                processed_count += 1
            else:
                print(f"  âŒ ÐžÑˆÐ¸Ð±ÐºÐ° Ð´Ð¾Ð±Ð°Ð²Ð»ÐµÐ½Ð¸Ñ Ð² Excel")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'error_'
                )
                
                if success:
                    print(f"  â†’ manual/{new_name}")
                    manual_files.append(pdf_file)
                    
                    # Ð£Ð²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ðµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, "ÐžÑˆÐ¸Ð±ÐºÐ° Ð´Ð¾Ð±Ð°Ð²Ð»ÐµÐ½Ð¸Ñ Ð² Excel", supplier)
                
                error_count += 1
                
        except Exception as e:
            print(f"  âŒ ÐžÑˆÐ¸Ð±ÐºÐ° Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÐ¸: {e}")
            log_processing(f"ERROR: {pdf_file} - {str(e)}")
            
            try:
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'error_'
                )
                if success:
                    print(f"  â†’ manual/{new_name}")
                    manual_files.append(pdf_file)
                    
                    # Ð£Ð²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ðµ Ð¾Ð± Ð¾ÑˆÐ¸Ð±ÐºÐµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_error(pdf_file, str(e))
            except:
                pass
            
            error_count += 1
    
    # ===== Ð˜Ð¢ÐžÐ“Ð˜ =====
    print("\n" + "="*80)
    print("Ð˜Ð¢ÐžÐ“Ð˜ ÐžÐ‘Ð ÐÐ‘ÐžÐ¢ÐšÐ˜")
    print("="*80)
    print(f"Ð’ÑÐµÐ³Ð¾ Ñ„Ð°Ð¹Ð»Ð¾Ð²: {total_files}")
    print(f"âœ“ ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ð°Ð½Ð¾ ÑƒÑÐ¿ÐµÑˆÐ½Ð¾: {processed_count}")
    print(f"âš  Ð”ÑƒÐ±Ð»Ð¸ÐºÐ°Ñ‚Ð¾Ð²: {duplicate_count}")
    print(f"âŒ Ð¢Ñ€ÐµÐ±ÑƒÑŽÑ‚ Ñ€ÑƒÑ‡Ð½Ð¾Ð¹ Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÐ¸: {error_count}")
    
    print("\nÐ¡Ñ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ° Ð¿Ð¾ Ð¿Ð¾ÑÑ‚Ð°Ð²Ñ‰Ð¸ÐºÐ°Ð¼:")
    for supplier, count in sorted(supplier_stats.items(), key=lambda x: x[1], reverse=True):
        print(f"  {supplier}: {count}")
    
    try:
        remaining_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
        remaining_count = len(remaining_files)
        if remaining_count > 0:
            print(f"\nâš  ÐžÑÑ‚Ð°Ð»Ð¸ÑÑŒ Ð½ÐµÐ¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚Ð°Ð½Ð½Ñ‹Ðµ Ñ„Ð°Ð¹Ð»Ñ‹: {remaining_count}")
    except:
        remaining_count = 0
    
    log_processing("="*80)
    log_processing(f"Ð—ÐÐ’Ð•Ð Ð¨Ð•ÐÐž: ÐžÐ±Ñ€Ð°Ð±Ð¾Ñ‚Ð°Ð½Ð¾={processed_count}, Ð”ÑƒÐ±Ð»Ð¸ÐºÐ°Ñ‚Ñ‹={duplicate_count}, Ð ÑƒÑ‡Ð½Ñ‹Ðµ={error_count}")
    log_processing("="*80)
    
    # Ð¤Ð¸Ð½Ð°Ð»ÑŒÐ½Ð°Ñ ÑÐ²Ð¾Ð´ÐºÐ° Ð² Telegram
    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
        summary = {
            'total': total_files,
            'processed': processed_count,
            'duplicates': duplicate_count,
            'manual': error_count,
            'processed_files': processed_files,
            'duplicate_files': duplicate_files,
            'manual_files': manual_files,
        }
        telegram.notify_summary(summary)
        
        # Ð’Ñ‹Ð²ÐµÑÑ‚Ð¸ ÑÑ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÑƒ ÑƒÐ²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ð¹
        stats = telegram.get_stats()
        print(f"\nðŸ“± Ð¡Ñ‚Ð°Ñ‚Ð¸ÑÑ‚Ð¸ÐºÐ° ÑƒÐ²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ð¹:")
        print(f"   ÐžÑ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¾: {stats['sent']}")
        print(f"   ÐžÑˆÐ¸Ð±Ð¾Ðº: {stats['failed']}")
    
    return {
        'total': total_files,
        'processed': processed_count,
        'duplicates': duplicate_count,
        'manual': error_count,
        'remaining': remaining_count,
        'supplier_stats': dict(supplier_stats)
    }


# ===== Ð¢ÐžÐ§ÐšÐ Ð’Ð¥ÐžÐ”Ð =====

if __name__ == "__main__":
    try:
        print("\n")
        print("â•”â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•—")
        print("â•‘         PDF PROCESSOR v6.0 - Ð£Ð›Ð£Ð§Ð¨Ð•ÐÐÐ«Ð• ÐŸÐÐ¢Ð¢Ð•Ð ÐÐ«            â•‘")
        print("â•‘         Ð˜ÑÐ¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¾: Auto Compass + Scania                   â•‘")
        print("â•šâ•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•")
        print("\n")
        
        print("ÐŸÑ€Ð¾Ð²ÐµÑ€ÐºÐ° ÑÐ¸ÑÑ‚ÐµÐ¼Ñ‹...")
        print(f"âœ“ PDF Ð¿Ð°Ð¿ÐºÐ°: {os.path.exists(PDF_FOLDER)}")
        print(f"âœ“ Excel Ñ„Ð°Ð¹Ð»: {os.path.exists(EXCEL_FILE)}")
        print(f"âœ“ Telegram: {TELEGRAM_ENABLED}")
        print(f"âœ“ Ð”ÐµÑ‚Ð°Ð»ÑŒÐ½Ñ‹Ðµ ÑƒÐ²ÐµÐ´Ð¾Ð¼Ð»ÐµÐ½Ð¸Ñ: {DETAILED_NOTIFICATIONS}")
        
        if not os.path.exists(PDF_FOLDER):
            print("\nâŒ ÐŸÐ°Ð¿ÐºÐ° PDF Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½Ð°!")
            exit(1)
        
        if not os.path.exists(EXCEL_FILE):
            print("\nâŒ Excel Ñ„Ð°Ð¹Ð» Ð½Ðµ Ð½Ð°Ð¹Ð´ÐµÐ½!")
            exit(1)
        
        print("\nÐ—Ð°Ð¿ÑƒÑÐº Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÐ¸...")
        results = process_all_pdfs()
        
        print("\nâœ… ÐŸÑ€Ð¾Ð³Ñ€Ð°Ð¼Ð¼Ð° Ð·Ð°Ð²ÐµÑ€ÑˆÐµÐ½Ð° ÑƒÑÐ¿ÐµÑˆÐ½Ð¾!")
        
        if results and results['total'] > 0:
            success_rate = (results['processed'] / results['total']) * 100
            print(f"\nðŸ“Š ÐŸÑ€Ð¾Ñ†ÐµÐ½Ñ‚ Ð°Ð²Ñ‚Ð¾Ð¼Ð°Ñ‚Ð¸Ð·Ð°Ñ†Ð¸Ð¸: {success_rate:.1f}%")
        
    except Exception as e:
        print(f"\nâŒ ÐšÑ€Ð¸Ñ‚Ð¸Ñ‡ÐµÑÐºÐ°Ñ Ð¾ÑˆÐ¸Ð±ÐºÐ°: {e}")
        log_processing(f"CRITICAL ERROR: {str(e)}")
        
        if TELEGRAM_ENABLED:
            try:
                telegram.notify_error("System", str(e))
            except:
                pass
        
        exit(1)
