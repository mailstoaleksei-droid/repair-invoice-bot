"""
PDF PROCESSOR v6.0 - С УЛУЧШЕННЫМИ ПАТТЕРНАМИ
Unified Telegram + улучшенное извлечение для manual файлов
Исправлено: Auto Compass Internal + Scania SCHWL
"""

import os
import shutil
import pdfplumber
import openpyxl
import re
from datetime import datetime
import time
from collections import defaultdict

# ===== НАСТРОЙКИ ПУТЕЙ =====
PDF_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG"
EXCEL_FILE = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG\Repair_2025.xlsx"
PROCESSED_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\RG 2025 Ersatyteile RepRG"
MANUAL_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\EingangsRG\manual"
LOG_FOLDER = r"C:\Users\Aleksei Samosvat\Groo GmbH\Intranet - Groo GmbH - Dokumente\Auto Compass GmbH\AC - Fahrzeuge\Ablage\Eingangs Rechnungen\PDF_Processor\logs"

# ===== НАСТРОЙКИ TELEGRAM =====
TELEGRAM_ENABLED = True
DETAILED_NOTIFICATIONS = True  # Включить детальные уведомления

# Импорт Telegram уведомлений (UNIFIED)
if TELEGRAM_ENABLED:
    try:
        from unified_telegram import create_client
        
        TELEGRAM_BOT_TOKEN = "8127115250:AAHmDuiiRuPSpE6oSwHzmUpSl2-DzVSr3Io"
        TELEGRAM_CHAT_ID = "745125435"
        
        telegram = create_client(TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID, throttle_interval=2.0)
        
    except ImportError as e:
        TELEGRAM_ENABLED = False
        print(f"⚠ Telegram уведомления недоступны: {e}")


# ===== ФУНКЦИИ ИЗВЛЕЧЕНИЯ ИЗ ИМЕНИ ФАЙЛА =====

def extract_truck_from_filename(filename):
    """Извлечь номер машины из имени файла"""
    
    # Паттерн 1: checked_XXX (основной формат)
    patterns = [
        r'checked_(\d+)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            number = match.group(1)
            if len(number) <= 2:
                return f"GR-OO{number.zfill(2)}"
            else:
                return f"GR-OO{number}"
    
    # Паттерн 2: XXXX - Поставщик (новый формат)
    # Примеры: "1726 - AC Intern.pdf", "771 - TIP.pdf"
    simple_number_pattern = r'^(\d{2,4})\s*[-–]\s*'
    match = re.match(simple_number_pattern, filename)
    if match:
        number = match.group(1)
        if len(number) <= 2:
            return f"GR-OO{number.zfill(2)}"
        elif len(number) == 3:
            return f"GR-OO{number}"
        else:  # 4 digits
            return f"GR-OO{number}"
    
    # Паттерн 3: Прямые указания (GR-OO, HH-AG, etc.)
    direct_patterns = [
        (r'GR[- ]?OO(\d+)', 'GR-OO'),
        (r'GR[- ]?(\d+)', 'GR-'),
        (r'WJQY4010', 'WJQY4010'),
        (r'OHAMX771', 'OHAMX771'),
        (r'HH[- ]?AG(\d+)', 'HH-AG'),
        (r'DE[- ]?FN(\d+)', 'DE-FN'),
    ]
    
    for pattern, prefix in direct_patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            if prefix in ['WJQY4010', 'OHAMX771']:
                return prefix
            number = match.group(1) if match.groups() else ''
            return f"{prefix}{number.zfill(3)}"
    
    return None


# ===== ФУНКЦИИ КЛАССИФИКАЦИИ ПОСТАВЩИКОВ =====

def identify_supplier(text):
    """
    Улучшенная классификация поставщиков
    Использует паттерны из process_ultimate.py
    """
    text_start = text[:1500].upper()
    text_upper = text.upper()
    
    # ПРИОРИТЕТ 1: Точные совпадения в начале документа
    
    # Vital Projekt
    if 'VITAL PROJEKT' in text_start:
        return 'Vital Projekt'
    
    # K&L
    elif 'K&L KFZ MEISTERBETRIEB' in text_start or 'K&L-KFZ' in text_start:
        return 'K&L'
    
    # Auto Compass Internal
    elif 'AUTO COMPASS' in text_start and ('KOPIE' in text or 'RANDERSWEIDE' in text):
        return 'Auto Compass'
    
    # Scania External (внешние счета)
    elif '#SPLMINFO' in text_start and 'SCANIA' in text_start:
        return 'Scania'
    
    # Ferronordic
    elif 'FERRONORDIC' in text_start:
        return 'Ferronordic'
    
    # HNS
    elif 'HNS NUTZFAHRZEUGE' in text_start or 'HNS SERVICE' in text_start:
        return 'HNS'
    
    # DEKRA
    elif 'DEKRA AUTOMOBIL' in text_start:
        return 'DEKRA'
    
    # Schütt
    elif 'SCHÜTT GMBH' in text_start or 'W. SCHÜTT' in text_start:
        return 'Schütt'
    
    # Volvo
    elif 'VOLVO GROUP TRUCKS' in text_start:
        return 'Volvo'
    
    # TIP Trailer
    elif 'TIP TRAILER SERVICES' in text_start or 'TRAILER SERVICES GERMANY' in text_start:
        return 'TIP Trailer'
    
    # Euromaster
    elif 'EUROMASTER GMBH' in text_start:
        return 'Euromaster'
    
    # Quick
    elif 'QUICK REIFEN' in text_start or 'REIFENDISCOUNT' in text_start:
        return 'Quick'
    
    # Sotecs
    elif 'SOTECS GMBH' in text_start:
        return 'Sotecs'
    
    # Tankpool24
    elif 'TANKPOOL' in text_start:
        return 'Tankpool24'
    
    # ПРИОРИТЕТ 2: Поиск во всём документе
    
    # MAN - только если действительно от MAN
    elif 'MAN TRUCK & BUS DEUTSCHLAND' in text_upper:
        return 'MAN'
    
    elif 'AUTO COMPASS' in text_start:
        return 'Auto Compass'
    
    elif 'GROO GMBH' in text_start:
        return 'Groo GmbH'
    
    return 'Unknown'


# ===== СПЕЦИАЛИЗИРОВАННЫЕ ФУНКЦИИ ИЗВЛЕЧЕНИЯ =====

def extract_dekra(text, filename):
    """DEKRA Automobil GmbH - Hauptuntersuchung"""
    data = {}
    
    # Номер счёта
    invoice_match = re.search(r'Rechnung\s+Nr\.\s*:?\s*(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Rechnung_Nr\.\s*:?\s*(\d+)', text)
    
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', text)
        if date_match:
            date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
        else:
            return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина
    truck_match = re.search(r'Kennzeichen[^\n]*:\s*([A-Z]{2}\s*[A-Z0-9]+\s*\d+)', text)
    if not truck_match:
        truck_match = re.search(r'KM-Stand\s+([A-Z]{2}\s*[A-Z0-9]+\s*\d+)', text, re.IGNORECASE)
    
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
        # Форматировать как GR-OO1514
        if len(data['truck']) >= 4:
            data['truck'] = f"{data['truck'][:2]}-{data['truck'][2:]}"
    else:
        # Попытка из имени файла
        data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Описание
    data['name'] = 'Hauptuntersuchung'
    data['amount'] = 1
    
    # Цена
    price_match = re.search(r'Nettopreis\s+[\d,]+\s+[\d,]+\s+([\d,]+)', text)
    if price_match:
        price_str = price_match.group(1).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['price'] = 0.0
    
    # Общая сумма - NETTO (без НДС)
    total_match = re.search(r'Nettobetrag\s+([\d,.]+)', text)
    if not total_match:
        # Альтернативный паттерн
        total_match = re.search(r'Gesamt[^\d]+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'DEKRA Automobil GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    
    return data


def extract_autocompass(text, filename):
    """Auto Compass Internal - внутренние счета"""
    data = {}
    
    # Номер счёта - улучшенные паттерны
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Interne\s+Rechnung\s+(\d+)', text)
    if not invoice_match:
        # Альтернативный паттерн для строки "Interne Rechnung 700293 Datum"
        invoice_match = re.search(r'Rechnung\s+(\d+)\s+Datum', text)
    
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Дата - улучшенные паттерны
    date_match = re.search(r'Datum\s+(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        date_match = re.search(r'Leistungsdatum:\s*(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        # Новый паттерн: "700293 Datum 08.10.2025"
        date_match = re.search(r'\d+\s+Datum\s+(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    
    if date_match:
        date_str = date_match.group(1)
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина - улучшенные паттерны
    truck_match = re.search(r'Kennzeichen\s+([A-Z]{2}-[A-Z]{2}\s*\d+)', text)
    if not truck_match:
        # Новый паттерн: "Kennzeichen Fahrgestell-Nr. ... HH-AG 1926"
        truck_match = re.search(r'Kennzeichen[^\n]*\n[^\n]*\n([A-Z]{2}-[A-Z]{2,4}\s*\d+)', text)
    if not truck_match:
        truck_match = re.search(r'([A-Z]{2}-[A-Z]{2,4}\s+\d+)', text)
    
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
    else:
        data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Описание
    data['name'] = 'Ремонтные работы'
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма - улучшенные паттерны
    # Паттерн 1: "Gesamt\n337,50 € 112,35 € 0.00 € 0.00 € 452,10 €"
    gesamt_match = re.search(r'Gesamt\s+(?:[\d,]+\s*€\s+){3,}([\d,]+)\s*€', text)
    if not gesamt_match:
        # Паттерн 2: "Gesamt ... € XXX,XX €" (последняя сумма)
        gesamt_match = re.search(r'Gesamt[^\n]*?([\d,]+)\s*€(?:\s|$)', text)
    if not gesamt_match:
        # Паттерн 3: Простой "Gesamt XXX,XX €"
        gesamt_match = re.search(r'Gesamt\s+([\d,.]+)\s*€', text)
    if not gesamt_match:
        netto_match = re.search(r'Netto\s+([\d,.]+)\s*€', text)
        if netto_match:
            gesamt_match = netto_match
    
    if gesamt_match:
        total_str = gesamt_match.group(1).replace('.', '').replace(',', '.')
        try:
            data['total_price'] = float(total_str)
        except:
            return None
    else:
        return None
    
    data['seller'] = 'Auto Compass GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    
    return data


def extract_scania(text, filename):
    """Scania - внешние счета от Scania"""
    data = {}
    
    # Номер счёта - ищем SCHWL (улучшенные паттерны)
    invoice_match = re.search(r'SCHWL(\d+)', text)
    if not invoice_match:
        # Альтернативный паттерн с пробелами
        invoice_match = re.search(r'SCHWL\s*(\d+)', text)
    if not invoice_match:
        # Паттерн из splminfo строки
        invoice_match = re.search(r'SCH_SCHWL(\d+)_', text)
    
    if invoice_match:
        data['invoice'] = f"SCHWL{invoice_match.group(1)}"
    else:
        return None
    
    # Дата - улучшенные паттерны
    # Паттерн 1: "RE-DATUM ... 20.10.25"
    date_match = re.search(r'RE-DATUM[^\n]*?(\d{2}\.\d{2}\.\d{2,4})', text)
    if not date_match:
        # Паттерн 2: Обычный формат даты
        date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{2,4})', text)
    
    if date_match:
        if len(date_match.groups()) == 1:
            # Формат DD.MM.YY или DD.MM.YYYY
            date_str = date_match.group(1)
            parts = date_str.split('.')
            if len(parts[2]) == 2:
                # Преобразовать YY в YYYY
                year = f"20{parts[2]}"
                date_str = f"{parts[0]}.{parts[1]}.{year}"
        else:
            # Формат из групп
            day = date_match.group(1)
            month = date_match.group(2)
            year = date_match.group(3)
            if len(year) == 2:
                year = f"20{year}"
            date_str = f"{day}.{month}.{year}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # Машина - улучшенные паттерны
    # Паттерн 1: "AMTL.KENNZ: GR-OO 1726"
    truck_match = re.search(r'AMTL\.KENNZ:\s*([A-Z]{2}-[A-Z]{2,4}\s*\d+)', text)
    if not truck_match:
        # Паттерн 2: "Kennzeichen ... GR-OO 1726"
        truck_match = re.search(r'Kennzeichen[:\s]+([A-Z]{2}-[A-Z]{2,4}\s*\d+)', text)
    if not truck_match:
        # Паттерн 3: Общий паттерн номера
        truck_match = re.search(r'([A-Z]{2}\s*[A-Z]{2}\s*\d+)', text)
    
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
        # Форматировать как GR-OO1726
        if len(data['truck']) >= 4 and '-' not in data['truck']:
            data['truck'] = f"{data['truck'][:2]}-{data['truck'][2:]}"
    else:
        data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Описание
    data['name'] = 'Scania service'
    data['amount'] = 1
    data['price'] = 0.0
    
    # Общая сумма - ПРИОРИТЕТ NETTO!
    # Паттерн 1: "NETTOBETRAG ... EUR XXX,XX" (ВЫСШИЙ ПРИОРИТЕТ)
    total_match = re.search(r'NETTOBETRAG[^\d]+EUR\s+([\d,.]+)', text)
    if not total_match:
        # Паттерн 2: "LOHN NETTO: XXX,XX"
        total_match = re.search(r'LOHN\s+NETTO:\s*([\d,.]+)', text)
    if not total_match:
        # Паттерн 3: "SUMME TEILE: XXX,XX"
        total_match = re.search(r'SUMME\s+TEILE:\s*([\d,]+)', text)
    if not total_match:
        # Паттерн 4: "SUMME: XX PE XXX,XX"
        total_match = re.search(r'SUMME:\s*\d+\s*PE\s*([\d,]+)', text)
    if not total_match:
        # Паттерн 5: "SUMME EUR XXX,XX" (может быть BRUTTO - осторожно)
        total_match = re.search(r'SUMME\s+EUR\s+([\d,.]+)', text)
    if not total_match:
        # Паттерн 6: Общий Gesamt
        total_match = re.search(r'Gesamt[^\d]+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        try:
            data['total_price'] = float(total_str)
        except:
            return None
    else:
        return None
    
    data['seller'] = 'Scania'
    data['buyer'] = 'Auto Compass GmbH'
    
    return data


def extract_universal(text, supplier, filename):
    """
    Универсальное извлечение для поставщиков без специальной функции
    """
    data = {}
    
    # Номер счета (универсальное)
    invoice_patterns = [
        r'Rechnung[s-]?[Nn]r\.?\s*:?\s*(\d+)',
        r'Invoice\s+(?:No\.?|Number)\s*:?\s*(\d+)',
        r'Rg\.?\s*-?\s*Nr\.?\s*:?\s*(\d+)',
        r'Rechnungs-Nr\.:\s*(\d+)',
    ]
    
    for pattern in invoice_patterns:
        match = re.search(pattern, text)
        if match:
            data['invoice'] = match.group(1)
            break
    
    if not data.get('invoice'):
        return None
    
    # Дата (универсальное)
    date_patterns = [
        r'(\d{2}\.\d{2}\.\d{4})',
        r'(\d{2})-(\d{2})-(\d{4})',
    ]
    
    for pattern in date_patterns:
        match = re.search(pattern, text)
        if match:
            if len(match.groups()) == 1:
                date_str = match.group(1)
            else:
                date_str = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
            
            try:
                date_obj = datetime.strptime(date_str.replace('-', '.'), '%d.%m.%Y')
                data['date'] = date_obj.strftime('%d.%m.%Y')
                data['month'] = date_obj.month
                data['week'] = date_obj.isocalendar()[1]
                break
            except:
                continue
    
    if not data.get('date'):
        return None
    
    # Машина
    data['truck'] = extract_truck_from_filename(filename) or ''
    
    # Сумма - ПРИОРИТЕТ NETTO для внешних счетов!
    amount_patterns = [
        r'Nettobetrag\s*([\d,\.]+)',              # 1. Nettobetrag (ВЫСШИЙ)
        r'Netto\s+([\d,\.]+)',                    # 2. Netto
        r'SUMME\s+(?:NETTO|netto)\s*:?\s*([\d,\.]+)',  # 3. SUMME NETTO
        r'Summe\s+netto\s*:?\s*([\d,\.]+)',       # 4. Summe netto
        r'Gesamtbetrag\s+(?:EUR)?\s*([\d,\.]+)',  # 5. Gesamtbetrag (BRUTTO)
        r'Summe\s+brutto\s*:?\s*([\d,\.]+)',      # 6. Summe brutto
        r'Total\s*:?\s*([\d,\.]+)',               # 7. Total
        r'Gesamt\s*:?\s*([\d,\.]+)',              # 8. Gesamt (может быть BRUTTO)
    ]
    
    for pattern in amount_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            amount_str = match.group(1).replace(',', '.').replace(' ', '')
            try:
                data['amount'] = 1
                data['price'] = float(amount_str)
                data['total_price'] = float(amount_str)
                break
            except:
                continue
    
    if not data.get('total_price'):
        return None
    
    data['seller'] = supplier
    data['buyer'] = 'Auto Compass GmbH'
    data['name'] = ''
    
    return data


# ===== РОУТЕР ИЗВЛЕЧЕНИЯ =====

def extract_data_by_supplier(text, supplier, filename):
    """
    Роутер: выбрать правильную функцию извлечения
    """
    # Специализированные extractors
    extractors = {
        'DEKRA': extract_dekra,
        'Auto Compass': extract_autocompass,
        'Scania': extract_scania,
    }
    
    extractor = extractors.get(supplier)
    if extractor:
        data = extractor(text, filename)
        if data:
            return data
    
    # Универсальное извлечение для остальных
    return extract_universal(text, supplier, filename)


# ===== ФУНКЦИИ РАБОТЫ С EXCEL =====

def check_invoice_exists(invoice_number, excel_file):
    """Проверить существует ли счет в Excel"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[8] == str(invoice_number):  # Колонка I (Invoice)
                wb.close()
                # Получить дату оригинала
                for r in ws.iter_rows(min_row=2):
                    if str(r[8].value) == str(invoice_number):
                        date = r[3].value  # Колонка D (Date)
                        wb.close()
                        return True, date
                wb.close()
                return True, None
        
        wb.close()
        return False, None
    except Exception as e:
        print(f"Ошибка проверки дубликата: {e}")
        return False, None


def add_to_excel(data, excel_file):
    """Добавить данные в Excel"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Найти первую пустую строку
        row_num = 2
        while ws[f'A{row_num}'].value is not None:
            row_num += 1
        
        # Заполнить данные
        ws[f'A{row_num}'] = data.get('month', 0)
        ws[f'B{row_num}'] = data.get('week', 0)
        ws[f'C{row_num}'] = data.get('truck', '')
        ws[f'D{row_num}'] = data.get('date', '')
        ws[f'E{row_num}'] = data.get('name', '')
        ws[f'F{row_num}'] = data.get('amount', 1)
        ws[f'G{row_num}'] = data.get('price', 0)
        ws[f'H{row_num}'] = data.get('total_price', 0)
        ws[f'I{row_num}'] = data.get('invoice', '')
        ws[f'J{row_num}'] = data.get('seller', '')
        ws[f'K{row_num}'] = data.get('buyer', 'Auto Compass GmbH')
        
        wb.save(excel_file)
        wb.close()
        
        return row_num
        
    except Exception as e:
        print(f"Ошибка добавления в Excel: {e}")
        return None


# ===== ФУНКЦИИ РАБОТЫ С ФАЙЛАМИ =====

def rename_and_move_file(source_path, destination_folder, prefix=''):
    """Переименовать и переместить файл"""
    try:
        filename = os.path.basename(source_path)
        
        # Убрать старые префиксы
        clean_filename = filename
        for old_prefix in ['checked_', 'manual_', 'error_', 'duplicate_']:
            if clean_filename.startswith(old_prefix):
                clean_filename = clean_filename[len(old_prefix):]
        
        # Добавить новый префикс
        new_filename = f"{prefix}{clean_filename}"
        destination_path = os.path.join(destination_folder, new_filename)
        
        # Если файл уже существует, добавить timestamp
        if os.path.exists(destination_path):
            name, ext = os.path.splitext(new_filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            new_filename = f"{name}_{timestamp}{ext}"
            destination_path = os.path.join(destination_folder, new_filename)
        
        shutil.move(source_path, destination_path)
        return True, new_filename, None
        
    except Exception as e:
        return False, None, str(e)


def log_processing(message):
    """Логирование в файл"""
    try:
        os.makedirs(LOG_FOLDER, exist_ok=True)
        log_file = os.path.join(LOG_FOLDER, f"processing_{datetime.now().strftime('%Y%m%d')}.log")
        
        with open(log_file, 'a', encoding='utf-8') as f:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            f.write(f"[{timestamp}] {message}\n")
    except Exception as e:
        print(f"Ошибка логирования: {e}")


# ===== ГЛАВНАЯ ФУНКЦИЯ =====

def process_all_pdfs():
    """
    Обработать все PDF файлы с детальными Telegram уведомлениями
    """
    
    print("\n╔══════════════════════════════════════════════════════════════╗")
    print("║      PDF PROCESSOR v6.0 - УЛУЧШЕННЫЕ ПАТТЕРНЫ               ║")
    print("║      Исправлено: Auto Compass + Scania manual файлы         ║")
    print("╚══════════════════════════════════════════════════════════════╝\n")
    
    # Создать папки если не существуют
    os.makedirs(MANUAL_FOLDER, exist_ok=True)
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)
    os.makedirs(LOG_FOLDER, exist_ok=True)
    
    # Получить список файлов
    try:
        pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    except Exception as e:
        print(f"❌ Ошибка доступа к папке: {e}")
        return None
    
    if not pdf_files:
        print("📭 Нет файлов для обработки")
        return {'total': 0, 'processed': 0, 'duplicates': 0, 'manual': 0}
    
    total_files = len(pdf_files)
    processed_count = 0
    duplicate_count = 0
    error_count = 0
    
    # Для сводки
    processed_files = []
    duplicate_files = []
    manual_files = []
    
    supplier_stats = defaultdict(int)
    
    log_processing("="*80)
    log_processing(f"НАЧАЛО ОБРАБОТКИ v4.0: {total_files} файлов")
    log_processing("="*80)
    
    print(f"\n📊 Найдено файлов: {total_files}")
    print("="*80)
    
    # Обработка каждого файла
    for index, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        
        print(f"\n[{index}/{total_files}] {pdf_file}")
        
        # Уведомление о начале обработки
        if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
            telegram.notify_processing_start(pdf_file)
        
        try:
            # ===== ШАГ 1: Извлечь текст =====
            full_text = ''
            
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + '\n'
            
            if not full_text.strip():
                print(f"  ⚠ PDF пустой")
                success, new_name, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                
                if success:
                    print(f"  → manual/{new_name}")
                    log_processing(f"MANUAL (пустой): {pdf_file} -> {new_name}")
                    manual_files.append(pdf_file)
                    
                    # Уведомление
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, "PDF пустой")
                
                error_count += 1
                continue
            
            # ===== ШАГ 2: Классифицировать =====
            supplier = identify_supplier(full_text)
            print(f"  Поставщик: {supplier}")
            supplier_stats[supplier] += 1
            
            # ===== ШАГ 3: Извлечь данные =====
            data = extract_data_by_supplier(full_text, supplier, pdf_file)
            
            if not data:
                print(f"  ❌ Не удалось извлечь данные")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'manual_'
                )
                
                if success:
                    print(f"  → manual/{new_name}")
                    log_processing(f"MANUAL ({supplier}): {pdf_file} -> {new_name}")
                    manual_files.append(pdf_file)
                    
                    # Уведомление
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, "Не удалось извлечь данные", supplier)
                
                error_count += 1
                continue
            
            if data.get('truck'):
                print(f"  Машина: {data['truck']}")
            
            print(f"  Счет: {data.get('invoice', 'N/A')}, Дата: {data.get('date', 'N/A')}, Сумма: {data.get('total_price', 0):.2f} EUR")
            
            # ===== ШАГ 4: Проверить дубликат =====
            is_duplicate, original_date = check_invoice_exists(data['invoice'], EXCEL_FILE)
            
            if is_duplicate:
                print(f"  ⚠ ДУБЛИКАТ! Счет {data['invoice']} уже существует")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'duplicate_'
                )
                
                if success:
                    print(f"  → processed/{new_name}")
                    log_processing(f"DUPLICATE: {pdf_file} -> {new_name}")
                    duplicate_files.append(pdf_file)
                    
                    # Уведомление
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_duplicate(pdf_file, data['invoice'], original_date)
                
                duplicate_count += 1
                continue
            
            # ===== ШАГ 5: Добавить в Excel =====
            excel_row = add_to_excel(data, EXCEL_FILE)
            
            if excel_row:
                print(f"  ✓ Добавлен в Excel: строка {excel_row}")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'checked_'
                )
                
                if success:
                    print(f"  → processed/{new_name}")
                    log_processing(f"SUCCESS: {pdf_file} -> {new_name}, Excel row {excel_row}")
                    processed_files.append(pdf_file)
                    
                    # Уведомление об успехе
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_success(data, pdf_file, excel_row)
                
                processed_count += 1
            else:
                print(f"  ❌ Ошибка добавления в Excel")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'error_'
                )
                
                if success:
                    print(f"  → manual/{new_name}")
                    manual_files.append(pdf_file)
                    
                    # Уведомление
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, "Ошибка добавления в Excel", supplier)
                
                error_count += 1
                
        except Exception as e:
            print(f"  ❌ Ошибка обработки: {e}")
            log_processing(f"ERROR: {pdf_file} - {str(e)}")
            
            try:
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'error_'
                )
                if success:
                    print(f"  → manual/{new_name}")
                    manual_files.append(pdf_file)
                    
                    # Уведомление об ошибке
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_error(pdf_file, str(e))
            except:
                pass
            
            error_count += 1
    
    # ===== ИТОГИ =====
    print("\n" + "="*80)
    print("ИТОГИ ОБРАБОТКИ")
    print("="*80)
    print(f"Всего файлов: {total_files}")
    print(f"✓ Обработано успешно: {processed_count}")
    print(f"⚠ Дубликатов: {duplicate_count}")
    print(f"❌ Требуют ручной обработки: {error_count}")
    
    print("\nСтатистика по поставщикам:")
    for supplier, count in sorted(supplier_stats.items(), key=lambda x: x[1], reverse=True):
        print(f"  {supplier}: {count}")
    
    try:
        remaining_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
        remaining_count = len(remaining_files)
        if remaining_count > 0:
            print(f"\n⚠ Остались необработанные файлы: {remaining_count}")
    except:
        remaining_count = 0
    
    log_processing("="*80)
    log_processing(f"ЗАВЕРШЕНО: Обработано={processed_count}, Дубликаты={duplicate_count}, Ручные={error_count}")
    log_processing("="*80)
    
    # Финальная сводка в Telegram
    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
        summary = {
            'total': total_files,
            'processed': processed_count,
            'duplicates': duplicate_count,
            'manual': error_count,
            'processed_files': processed_files,
            'duplicate_files': duplicate_files,
            'manual_files': manual_files,
        }
        telegram.notify_summary(summary)
        
        # Вывести статистику уведомлений
        stats = telegram.get_stats()
        print(f"\n📱 Статистика уведомлений:")
        print(f"   Отправлено: {stats['sent']}")
        print(f"   Ошибок: {stats['failed']}")
    
    return {
        'total': total_files,
        'processed': processed_count,
        'duplicates': duplicate_count,
        'manual': error_count,
        'remaining': remaining_count,
        'supplier_stats': dict(supplier_stats)
    }


# ===== ТОЧКА ВХОДА =====

if __name__ == "__main__":
    try:
        print("\n")
        print("╔══════════════════════════════════════════════════════════════╗")
        print("║         PDF PROCESSOR v6.0 - УЛУЧШЕННЫЕ ПАТТЕРНЫ            ║")
        print("║         Исправлено: Auto Compass + Scania                   ║")
        print("╚══════════════════════════════════════════════════════════════╝")
        print("\n")
        
        print("Проверка системы...")
        print(f"✓ PDF папка: {os.path.exists(PDF_FOLDER)}")
        print(f"✓ Excel файл: {os.path.exists(EXCEL_FILE)}")
        print(f"✓ Telegram: {TELEGRAM_ENABLED}")
        print(f"✓ Детальные уведомления: {DETAILED_NOTIFICATIONS}")
        
        if not os.path.exists(PDF_FOLDER):
            print("\n❌ Папка PDF не найдена!")
            exit(1)
        
        if not os.path.exists(EXCEL_FILE):
            print("\n❌ Excel файл не найден!")
            exit(1)
        
        print("\nЗапуск обработки...")
        results = process_all_pdfs()
        
        print("\n✅ Программа завершена успешно!")
        
        if results and results['total'] > 0:
            success_rate = (results['processed'] / results['total']) * 100
            print(f"\n📊 Процент автоматизации: {success_rate:.1f}%")
        
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        log_processing(f"CRITICAL ERROR: {str(e)}")
        
        if TELEGRAM_ENABLED:
            try:
                telegram.notify_error("System", str(e))
            except:
                pass
        
        exit(1)
