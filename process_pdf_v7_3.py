"""
PDF PROCESSOR v7.3 - Ð¤Ð˜ÐÐÐ›Ð¬ÐÐÐ¯ Ð’Ð•Ð Ð¡Ð˜Ð¯
Unified Telegram + 16 Ð¿Ð¾ÑÑ‚Ð°Ð²Ñ‰Ð¸ÐºÐ¾Ð² + Ð¿Ñ€Ð°Ð²Ð¸Ð»ÑŒÐ½Ð°Ñ Ð»Ð¾Ð³Ð¸ÐºÐ° NETTO/BRUTTO
Ð˜ÑÐ¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¾: Auto Compass (Internal BRUTTO, External NETTO)
Ð”Ð¾Ð±Ð°Ð²Ð»ÐµÐ½Ð¾: 13 Ð½Ð¾Ð²Ñ‹Ñ… Ñ„ÑƒÐ½ÐºÑ†Ð¸Ð¹ Ð¸Ð·Ð²Ð»ÐµÑ‡ÐµÐ½Ð¸Ñ + Ð°Ð²Ñ‚Ð¾Ð¾Ð¿Ñ€ÐµÐ´ÐµÐ»ÐµÐ½Ð¸Ðµ Ð³Ð¾Ð´Ð° + Ð·Ð°Ñ‰Ð¸Ñ‚Ð° Ð¾Ñ‚ Ð¾ÑˆÐ¸Ð±Ð¾Ðº ÐºÐ¾Ð´Ð¸Ñ€Ð¾Ð²ÐºÐ¸
"""
import os
import shutil
import pdfplumber
import openpyxl
import re
from datetime import datetime
import time
from collections import defaultdict
from dotenv import load_dotenv

try:
    from truck_reference import (
        extract_reference_trucks,
        normalize_truck_candidate,
        extract_normalized_truck_number,
        strip_truck_number_from_text,
    )
except ImportError:
    def extract_reference_trucks(_text):
        return []

    def normalize_truck_candidate(value):
        return value or ''

    def extract_normalized_truck_number(value):
        return str(value or '').strip()

    def strip_truck_number_from_text(value):
        return str(value or '').strip()

try:
    from supplier_reference import extract_reference_suppliers, extract_company_name_only
except ImportError:
    def extract_reference_suppliers(_text):
        return []

    def extract_company_name_only(value):
        return str(value or '').strip()

try:
    from ai_invoice_extractor import extract_invoice_with_ai, is_ai_available
except ImportError:
    def extract_invoice_with_ai(*_args, **_kwargs):
        return None, "ai_module_unavailable"

    def is_ai_available():
        return False

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

from path_config import DATA_ROOT, EXCEL_FILE, LOG_FOLDER, MANUAL_FOLDER, PDF_FOLDER, PROCESSED_FOLDER, REPORT_FOLDER


def _env_flag(name, default=False):
    value = os.getenv(name)
    if value is None:
        return default
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on'}

# Processing modes:
# - report_only: keep PDF files in place and write a separate per-run report
# - production: write to the master Excel and move files to processed/manual folders
PROCESSING_MODE = os.getenv("PROCESSING_MODE", "report_only").strip().lower() or "report_only"
WRITE_MASTER_EXCEL = PROCESSING_MODE == "production"
MOVE_FILES = PROCESSING_MODE == "production"

# ===== ÃÂÃÂÃÂ¡ÃÂ¢ÃÂ ÃÅ¾Ãâ„¢ÃÅ¡ÃËœ TELEGRAM =====
TELEGRAM_ENABLED = _env_flag("TELEGRAM_ENABLED", True)
DETAILED_NOTIFICATIONS = False  # Per-file Telegram notifications; summary is still sent.

# ÃËœÃÂ¼ÃÂ¿ÃÂ¾Ã‘â‚¬Ã‘â€š Telegram Ã‘Æ’ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸ÃÂ¹ (UNIFIED)
if TELEGRAM_ENABLED:
    try:
        from unified_telegram import create_client

        TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
        TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "").strip()
        if TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID:
            telegram = create_client(TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID, throttle_interval=2.0)
        else:
            TELEGRAM_ENABLED = False

    except Exception as e:
        TELEGRAM_ENABLED = False
        print(f"Telegram notifications unavailable: {e}")


# ===== ÃÂ¤ÃÂ£ÃÂÃÅ¡ÃÂ¦ÃËœÃËœ ÃËœÃâ€”Ãâ€™Ãâ€ºÃâ€¢ÃÂ§Ãâ€¢ÃÂÃËœÃÂ¯ ÃËœÃâ€” ÃËœÃÅ“Ãâ€¢ÃÂÃËœ ÃÂ¤ÃÂÃâ„¢Ãâ€ºÃÂ =====

def extract_truck_from_filename(filename):
    """ÃËœÃÂ·ÃÂ²ÃÂ»ÃÂµÃ‘â€¡Ã‘Å’ ÃÂ½ÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ ÃÂ¼ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½Ã‘â€¹ ÃÂ¸ÃÂ· ÃÂ¸ÃÂ¼ÃÂµÃÂ½ÃÂ¸ Ã‘â€žÃÂ°ÃÂ¹ÃÂ»ÃÂ°"""
    
    # ÃÅ¸ÃÂ°Ã‘â€šÃ‘â€šÃÂµÃ‘â‚¬ÃÂ½ 1: checked_XXX (ÃÂ¾Ã‘ÂÃÂ½ÃÂ¾ÃÂ²ÃÂ½ÃÂ¾ÃÂ¹ Ã‘â€žÃÂ¾Ã‘â‚¬ÃÂ¼ÃÂ°Ã‘â€š)
    patterns = [
        r'checked_(\d+)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            number = match.group(1)
            if len(number) <= 2:
                return normalize_truck_candidate(f"GR-OO{number.zfill(2)}")
            else:
                return normalize_truck_candidate(f"GR-OO{number}")
    
    # ÃÅ¸ÃÂ°Ã‘â€šÃ‘â€šÃÂµÃ‘â‚¬ÃÂ½ 2: XXXX - ÃÅ¸ÃÂ¾Ã‘ÂÃ‘â€šÃÂ°ÃÂ²Ã‘â€°ÃÂ¸ÃÂº (ÃÂ½ÃÂ¾ÃÂ²Ã‘â€¹ÃÂ¹ Ã‘â€žÃÂ¾Ã‘â‚¬ÃÂ¼ÃÂ°Ã‘â€š)
    # ÃÅ¸Ã‘â‚¬ÃÂ¸ÃÂ¼ÃÂµÃ‘â‚¬Ã‘â€¹: "1726 - AC Intern.pdf", "771 - TIP.pdf"
    simple_number_pattern = r'^(\d{2,4})\s*[-Ã¢â‚¬â€œ]\s*'
    match = re.match(simple_number_pattern, filename)
    if match:
        number = match.group(1)
        if len(number) <= 2:
            return normalize_truck_candidate(f"GR-OO{number.zfill(2)}")
        elif len(number) == 3:
            return normalize_truck_candidate(f"GR-OO{number}")
        else:  # 4 digits
            return normalize_truck_candidate(f"GR-OO{number}")
    
    # ÃÅ¸ÃÂ°Ã‘â€šÃ‘â€šÃÂµÃ‘â‚¬ÃÂ½ 3: ÃÅ¸Ã‘â‚¬Ã‘ÂÃÂ¼Ã‘â€¹ÃÂµ Ã‘Æ’ÃÂºÃÂ°ÃÂ·ÃÂ°ÃÂ½ÃÂ¸Ã‘Â (GR-OO, HH-AG, etc.)
    direct_patterns = [
        (r'GR[- ]?OO(\d+)', 'GR-OO'),
        (r'GR[- ]?(\d+)', 'GR-'),
        (r'WJQY4010', 'WJQY4010'),
        (r'OHAMX771', 'OHAMX771'),
        (r'HH[- ]?AG(\d+)', 'HH-AG'),
        (r'DE[- ]?FN(\d+)', 'DE-FN'),
    ]
    
    for pattern, prefix in direct_patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            if prefix in ['WJQY4010', 'OHAMX771']:
                return normalize_truck_candidate(prefix)
            number = match.group(1) if match.groups() else ''
            return normalize_truck_candidate(f"{prefix}{number.zfill(3)}")

    reference_trucks = extract_reference_trucks(filename)
    if reference_trucks:
        return reference_trucks[0]
    
    return None


# ===== ÃÂ¤ÃÂ£ÃÂÃÅ¡ÃÂ¦ÃËœÃËœ ÃÅ¡Ãâ€ºÃÂÃÂ¡ÃÂ¡ÃËœÃÂ¤ÃËœÃÅ¡ÃÂÃÂ¦ÃËœÃËœ ÃÅ¸ÃÅ¾ÃÂ¡ÃÂ¢ÃÂÃâ€™ÃÂ©ÃËœÃÅ¡ÃÅ¾Ãâ€™ =====

def identify_supplier(text):
    """
    ÃÂ£ÃÂ»Ã‘Æ’Ã‘â€¡Ã‘Ë†ÃÂµÃÂ½ÃÂ½Ã‘â€¹ÃÂ¹ ÃÂºÃÂ»ÃÂ°Ã‘ÂÃ‘ÂÃÂ¸Ã‘â€žÃÂ¸ÃÂºÃÂ°Ã‘â€šÃÂ¾Ã‘â‚¬ ÃÂ¿ÃÂ¾Ã‘ÂÃ‘â€šÃÂ°ÃÂ²Ã‘â€°ÃÂ¸ÃÂºÃÂ¾ÃÂ²
    Ãâ€™ÃÂ¾ÃÂ·ÃÂ²Ã‘â‚¬ÃÂ°Ã‘â€°ÃÂ°ÃÂµÃ‘â€š ÃÂ½ÃÂ°ÃÂ·ÃÂ²ÃÂ°ÃÂ½ÃÂ¸ÃÂµ ÃÂ¿ÃÂ¾Ã‘ÂÃ‘â€šÃÂ°ÃÂ²Ã‘â€°ÃÂ¸ÃÂºÃÂ°
    """
    text_start = text[:1500].upper()
    text_upper = text.upper()
    
    # ÃÅ¸ÃÂ ÃËœÃÅ¾ÃÂ ÃËœÃÂ¢Ãâ€¢ÃÂ¢ 1: ÃÂ¢ÃÂ¾Ã‘â€¡ÃÂ½Ã‘â€¹ÃÂµ Ã‘ÂÃÂ¾ÃÂ²ÃÂ¿ÃÂ°ÃÂ´ÃÂµÃÂ½ÃÂ¸Ã‘Â ÃÂ² ÃÂ½ÃÂ°Ã‘â€¡ÃÂ°ÃÂ»ÃÂµ ÃÂ´ÃÂ¾ÃÂºÃ‘Æ’ÃÂ¼ÃÂµÃÂ½Ã‘â€šÃÂ°
    
    # Vital Projekt
    if 'VITAL PROJEKT' in text_start:
        return 'Vital Projekt'
    
    # K&L
    elif 'K&L KFZ MEISTERBETRIEB' in text_start or 'K&L-KFZ' in text_start:
        return 'K&L'
    
    # Scania External (ÃÂ²ÃÂ½ÃÂµÃ‘Ë†ÃÂ½ÃÂ¸ÃÂµ Ã‘ÂÃ‘â€¡ÃÂµÃ‘â€šÃÂ°)
    elif ('#SPLMINFO' in text_start and re.search(r'\bSCH[A-Z]{2}\d{4,}\b', text_upper)) or (
        re.search(r'\bSCH[A-Z]{2}\d{4,}\b', text_upper)
        and any(marker in text_upper for marker in ('RE-NR.', 'AUFTRAGS-NR.', 'AMTL.KENNZ', 'SC/WE EXTERN', 'KENNZEICHEN'))
    ) or (
        'SCANIA' in text_start and any(marker in text_upper for marker in ('RE-NR.', 'AUFTRAGS-NR.', 'AMTL.KENNZ', 'SC/WE EXTERN'))
    ):
        return 'Scania External'

    # Auto Compass Internal
    elif 'AUTO COMPASS' in text_start and ('KOPIE' in text or 'RANDERSWEIDE' in text):
        return 'Auto Compass (Internal)'
    
    # Ferronordic
    elif 'FERRONORDIC' in text_start:
        return 'Ferronordic'
    
    # HNS
    elif 'HNS NUTZFAHRZEUGE' in text_start or 'HNS SERVICE' in text_start:
        return 'HNS'
    
    # DEKRA
    elif 'DEKRA AUTOMOBIL' in text_start:
        return 'DEKRA'
    
    # SchÃƒÂ¼tt
    elif 'SCHÃƒÅ“TT GMBH' in text_start or 'W. SCHÃƒÅ“TT' in text_start:
        return 'SchÃƒÂ¼tt'
    
    # Volvo
    elif 'VOLVO GROUP TRUCKS' in text_start:
        return 'Volvo'
    
    # TIP Trailer
    elif 'TIP TRAILER SERVICES' in text_start or 'TRAILER SERVICES GERMANY' in text_start:
        return 'TIP'
    
    # Euromaster
    elif 'EUROMASTER GMBH' in text_start:
        return 'Euromaster'
    
    # Quick
    elif 'QUICK REIFEN' in text_start or 'REIFENDISCOUNT' in text_start:
        return 'Quick'
    
    # Sotecs
    elif 'SOTECS GMBH' in text_start:
        return 'Sotecs'
    
    # Express
    elif 'EXPRESS' in text_start and 'UAB GROO' in text:
        return 'Express'
    
    # Tankpool24
    elif 'TANKPOOL' in text_start:
        return 'Tankpool24'
    
    # ÃÅ¸ÃÂ ÃËœÃÅ¾ÃÂ ÃËœÃÂ¢Ãâ€¢ÃÂ¢ 2: ÃÅ¸ÃÂ¾ÃÂ¸Ã‘ÂÃÂº ÃÂ²ÃÂ¾ ÃÂ²Ã‘ÂÃ‘â€˜ÃÂ¼ ÃÂ´ÃÂ¾ÃÂºÃ‘Æ’ÃÂ¼ÃÂµÃÂ½Ã‘â€šÃÂµ
    
    # MAN - Ã‘â€šÃÂ¾ÃÂ»Ã‘Å’ÃÂºÃÂ¾ ÃÂµÃ‘ÂÃÂ»ÃÂ¸ ÃÂ´ÃÂµÃÂ¹Ã‘ÂÃ‘â€šÃÂ²ÃÂ¸Ã‘â€šÃÂµÃÂ»Ã‘Å’ÃÂ½ÃÂ¾ ÃÂ¾Ã‘â€š MAN
    elif 'MAN TRUCK & BUS DEUTSCHLAND' in text_upper:
        return 'MAN'

    reference_matches = extract_reference_suppliers(text)
    if reference_matches:
        return reference_matches[0]
    
    return 'Unknown'


# ===== Ãâ€™ÃÂ¡ÃÅ¸ÃÅ¾ÃÅ“ÃÅ¾Ãâ€œÃÂÃÂ¢Ãâ€¢Ãâ€ºÃÂ¬ÃÂÃÂ«Ãâ€¢ ÃÂ¤ÃÂ£ÃÂÃÅ¡ÃÂ¦ÃËœÃËœ =====


def extract_dekra(text, filename):
    """DEKRA Automobil GmbH - Hauptuntersuchung"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnung\s+Nr\.\s*:?\s*(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Rechnung_Nr\.\s*:?\s*(\d+)', text)
    
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', text)
        if date_match:
            date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
        else:
            return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ°
    truck_match = re.search(r'Kennzeichen[^\n]*:\s*([A-Z]{2}\s*[A-Z0-9]+\s*\d+)', text)
    if not truck_match:
        truck_match = re.search(r'KM-Stand\s+([A-Z]{2}\s*[A-Z0-9]+\s*\d+)', text, re.IGNORECASE)
    
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
        # ÃÂ¤ÃÂ¾Ã‘â‚¬ÃÂ¼ÃÂ°Ã‘â€šÃÂ¸Ã‘â‚¬ÃÂ¾ÃÂ²ÃÂ°Ã‘â€šÃ‘Å’ ÃÂºÃÂ°ÃÂº GR-OO1514
        if len(data['truck']) >= 4:
            data['truck'] = f"{data['truck'][:2]}-{data['truck'][2:]}"
    else:
        # ÃÅ¸ÃÂ¾ÃÂ¿Ã‘â€¹Ã‘â€šÃÂºÃÂ° ÃÂ¸ÃÂ· ÃÂ¸ÃÂ¼ÃÂµÃÂ½ÃÂ¸ Ã‘â€žÃÂ°ÃÂ¹ÃÂ»ÃÂ°
        data['truck'] = extract_truck_from_filename(filename) or ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ
    data['name'] = 'Hauptuntersuchung'
    data['amount'] = 1
    
    # ÃÂ¦ÃÂµÃÂ½ÃÂ°
    price_match = re.search(r'Nettopreis\s+[\d,]+\s+[\d,]+\s+([\d,]+)', text)
    if price_match:
        price_str = price_match.group(1).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ° - NETTO (ÃÂ±ÃÂµÃÂ· ÃÂÃâ€ÃÂ¡)
    total_match = re.search(r'Nettobetrag\s+([\d,.]+)', text)
    if not total_match:
        # ÃÂÃÂ»Ã‘Å’Ã‘â€šÃÂµÃ‘â‚¬ÃÂ½ÃÂ°Ã‘â€šÃÂ¸ÃÂ²ÃÂ½Ã‘â€¹ÃÂ¹ ÃÂ¿ÃÂ°Ã‘â€šÃ‘â€šÃÂµÃ‘â‚¬ÃÂ½
        total_match = re.search(r'Gesamt[^\d]+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'DEKRA Automobil GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    
    return data


def extract_autocompass(text, filename):
    """Auto Compass Internal - Ð²Ð½ÑƒÑ‚Ñ€ÐµÐ½Ð½Ð¸Ðµ ÑÑ‡ÐµÑ‚Ð°
    Ð’ÐÐ–ÐÐž: Ð”Ð»Ñ Internal Ð¸ÑÐ¿Ð¾Ð»ÑŒÐ·ÑƒÐµÑ‚ BRUTTO (Gesamt), Ð´Ð»Ñ External - NETTO
    """
    data = {}
    
    # ÐÐ¾Ð¼ÐµÑ€ ÑÑ‡Ñ‘Ñ‚Ð° - ÑƒÐ»ÑƒÑ‡ÑˆÐµÐ½Ð½Ñ‹Ðµ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½Ñ‹
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if not invoice_match:
        invoice_match = re.search(r'Interne\s+Rechnung\s+(\d+)', text)
    if not invoice_match:
        # ÐÐ»ÑŒÑ‚ÐµÑ€Ð½Ð°Ñ‚Ð¸Ð²Ð½Ñ‹Ð¹ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½ Ð´Ð»Ñ ÑÑ‚Ñ€Ð¾ÐºÐ¸ "Interne Rechnung 700293 Datum"
        invoice_match = re.search(r'Rechnung\s+(\d+)\s+Datum', text)
    
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ð”Ð°Ñ‚Ð° - ÑƒÐ»ÑƒÑ‡ÑˆÐµÐ½Ð½Ñ‹Ðµ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½Ñ‹
    date_match = re.search(r'Datum\s+(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        date_match = re.search(r'Leistungsdatum:\s*(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        # ÐÐ¾Ð²Ñ‹Ð¹ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½: "700293 Datum 08.10.2025"
        date_match = re.search(r'\d+\s+Datum\s+(\d{2}\.\d{2}\.\d{4})', text)
    if not date_match:
        date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
    
    if date_match:
        date_str = date_match.group(1)
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÐœÐ°ÑˆÐ¸Ð½Ð° - ÑƒÐ»ÑƒÑ‡ÑˆÐµÐ½Ð½Ñ‹Ðµ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½Ñ‹
    truck_match = re.search(r'Kennzeichen\s+([A-Z]{2}-[A-Z]{2}\s*\d+)', text)
    if not truck_match:
        # ÐÐ¾Ð²Ñ‹Ð¹ Ð¿Ð°Ñ‚Ñ‚ÐµÑ€Ð½: "Kennzeichen Fahrgestell-Nr. ... HH-AG 1926"
        truck_match = re.search(r'Kennzeichen[^\n]*\n[^\n]*\n([A-Z]{2}-[A-Z]{2,4}\s*\d+)', text)
    if not truck_match:
        truck_match = re.search(r'([A-Z]{2}-[A-Z]{2,4}\s+\d+)', text)
    
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
    else:
        data['truck'] = extract_truck_from_filename(filename) or ''
    
    # ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ
    data['name'] = 'Ð ÐµÐ¼Ð¾Ð½Ñ‚Ð½Ñ‹Ðµ Ñ€Ð°Ð±Ð¾Ñ‚Ñ‹'
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÐžÐ¿Ñ€ÐµÐ´ÐµÐ»ÑÐµÐ¼ seller Ð¸ buyer
    data['seller'] = 'Auto Compass GmbH'
    
    # ÐŸÑ‹Ñ‚Ð°ÐµÐ¼ÑÑ Ð½Ð°Ð¹Ñ‚Ð¸ buyer (Firma)
    buyer_match = re.search(r'Firma\s+(.+?)(?:\n|Randersweide)', text)
    if buyer_match:
        buyer = buyer_match.group(1).strip()
        data['buyer'] = buyer
    else:
        data['buyer'] = 'Auto Compass GmbH'  # ÐŸÐ¾ ÑƒÐ¼Ð¾Ð»Ñ‡Ð°Ð½Ð¸ÑŽ

    data['seller'] = normalize_party_name(data.get('seller', ''))
    data['buyer'] = normalize_party_name(data.get('buyer', ''))
    
    # ðŸ”´ ÐšÐ Ð˜Ð¢Ð˜Ð§Ð•Ð¡ÐšÐÐ¯ Ð›ÐžÐ“Ð˜ÐšÐ: Internal (BRUTTO) vs External (NETTO)
    is_internal = (data['seller'] == data['buyer'])
    
    if is_internal:
        # âš ï¸ INTERNAL ÑÑ‡Ñ‘Ñ‚: Ð¸ÑÐ¿Ð¾Ð»ÑŒÐ·ÑƒÐµÐ¼ BRUTTO (Gesamt)
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 1: "Gesamt\n337,50 â‚¬ 112,35 â‚¬ 0.00 â‚¬ 0.00 â‚¬ 452,10 â‚¬"
        gesamt_match = re.search(r'Gesamt\s+(?:[\d,]+\s*â‚¬\s+){3,}([\d,]+)\s*â‚¬', text)
        if not gesamt_match:
            # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 2: "Gesamt ... â‚¬ XXX,XX â‚¬" (Ð¿Ð¾ÑÐ»ÐµÐ´Ð½ÑÑ ÑÑƒÐ¼Ð¼Ð°)
            gesamt_match = re.search(r'Gesamt[^\n]*?([\d,]+)\s*â‚¬(?:\s|$)', text)
        if not gesamt_match:
            # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 3: ÐŸÑ€Ð¾ÑÑ‚Ð¾Ð¹ "Gesamt XXX,XX â‚¬"
            gesamt_match = re.search(r'Gesamt\s+([\d,.]+)\s*â‚¬', text)
        
        if gesamt_match:
            total_str = gesamt_match.group(1).replace('.', '').replace(',', '.')
            try:
                data['total_price'] = float(total_str)
            except:
                return None
        else:
            return None
    else:
        # âœ… EXTERNAL ÑÑ‡Ñ‘Ñ‚: Ð¸ÑÐ¿Ð¾Ð»ÑŒÐ·ÑƒÐµÐ¼ NETTO
        # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 1: Ð¢Ð°Ð±Ð»Ð¸Ñ†Ð° "Lohn Material Fremdleistung Netto"
        netto_pattern = re.search(
            r'Lohn\s+Material\s+Fremdleistung\s+Netto\s+[\d,.]+\s*â‚¬\s+[\d,.]+\s*â‚¬\s+[\d,.]+\s*â‚¬\s+([\d,.]+)\s*â‚¬',
            text
        )
        if not netto_pattern:
            # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 2: 4 Ñ‡Ð¸ÑÐ»Ð° Ð¿ÐµÑ€ÐµÐ´ "Gesamt"
            netto_pattern = re.search(
                r'[\d,.]+\s*â‚¬\s+[\d,.]+\s*â‚¬\s+[\d,.]+\s*â‚¬\s+([\d,.]+)\s*â‚¬\s+Gesamt',
                text
            )
        if not netto_pattern:
            # ÐŸÐ°Ñ‚Ñ‚ÐµÑ€Ð½ 3: ÐŸÑ€Ð¾ÑÑ‚Ð¾Ð¹ "Netto XXX,XX â‚¬"
            netto_pattern = re.search(r'Netto\s+([\d,.]+)\s*â‚¬', text)
        
        if netto_pattern:
            total_str = netto_pattern.group(1).replace('.', '').replace(',', '.')
            try:
                data['total_price'] = float(total_str)
            except:
                return None
        else:
            return None
    
    return data


def extract_vital_projekt(text):
    """Vital Projekt - Ã‘Ë†ÃÂ¸ÃÂ½Ã‘â€¹ ÃÂ¸ Ã‘Æ’Ã‘ÂÃÂ»Ã‘Æ’ÃÂ³ÃÂ¸"""
    data = {}
    
    invoice_match = re.search(r'Rechnungs-Nr\.:\s*(\d+)', text)
    if not invoice_match:
        return None
    data['invoice'] = invoice_match.group(1)
    
    date_match = re.search(r'Hamburg,\s*den\s*(\d{2}\.\d{2}\.\d{4})', text)
    if date_match:
        date_str = date_match.group(1)
    else:
        date_match2 = re.search(r'Leistungsdatum/-Ort:(\d{2})-(\d{2})-(\d{4})', text)
        if date_match2:
            date_str = f"{date_match2.group(1)}.{date_match2.group(2)}.{date_match2.group(3)}"
        else:
            return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    truck_match = re.search(r'Kennzeichen/Fahrer:\s*([A-Z]{2}\s*[A-Z]{2}\s*\d+)', text)
    if not truck_match:
        return None
    
    truck_raw = truck_match.group(1)
    truck_clean = re.sub(r'\s+', '', truck_raw)
    if len(truck_clean) >= 4:
        data['truck'] = f"{truck_clean[:2]}-{truck_clean[2:]}"
    else:
        data['truck'] = truck_clean
    
    table_match = re.search(
        r'^\s*1\s+(\d+)\s+Stk\.\s+[^\s]+\s+(.+?)\s+\d+%\s+([\d,]+)\s*Ã¢â€šÂ¬',
        text, re.MULTILINE
    )
    
    if table_match:
        data['amount'] = int(table_match.group(1))
        description = table_match.group(2).strip()
        data['name'] = description[:50] if len(description) > 50 else description
        price_str = table_match.group(3).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['amount'] = 1
        data['name'] = "ÃÂ¨ÃÂ¸ÃÂ½Ã‘â€¹ ÃÂ¸ Ã‘Æ’Ã‘ÂÃÂ»Ã‘Æ’ÃÂ³ÃÂ¸"
        data['price'] = 0.0
    
    summe_match = re.search(r'Summe\s+([\d,.]+)\s*Ã¢â€šÂ¬', text)
    if summe_match:
        total_str = summe_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Vital Projekt Inh.Vitalij Barth'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_ferronordic(text):
    """Ferronordic - Ã‘ÂÃÂµÃ‘â‚¬ÃÂ²ÃÂ¸Ã‘Â Volvo"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnung:\s*RE(\d+)-(\d+)', text)
    if invoice_match:
        data['invoice'] = f"RE{invoice_match.group(1)}-{invoice_match.group(2)}"
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'Vom:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ°
    truck_match = re.search(r'Kennzeichen:\s*DE\s*FN\s*(\d+)', text)
    if truck_match:
        data['truck'] = f"DE-FN{truck_match.group(1)}"
    else:
        truck_match2 = re.search(r'DE\s*FN\s*(\d+)', text)
        if truck_match2:
            data['truck'] = f"DE-FN{truck_match2.group(1)}"
        else:
            data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ
    desc_match = re.search(r'1\s+Monatstarif[^\n]+', text)
    if desc_match:
        data['name'] = desc_match.group(0)[:50]
    else:
        data['name'] = 'Wartung/Service Volvo'
    
    # ÃÅ¡ÃÂ¾ÃÂ»ÃÂ¸Ã‘â€¡ÃÂµÃ‘ÂÃ‘â€šÃÂ²ÃÂ¾ ÃÂ¸ Ã‘â€ ÃÂµÃÂ½ÃÂ°
    amount_match = re.search(r'1\s+Monatstarif[^\d]+(\d+,\d+)\s*Ã¢â€šÂ¬', text)
    if amount_match:
        data['amount'] = 1
        price_str = amount_match.group(1).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['amount'] = 1
        data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ° (Endsumme Ã‘Â ÃÂÃâ€ÃÂ¡)
    total_match = re.search(r'Endsumme\s+([\d,.]+)\s*Ã¢â€šÂ¬', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Ferronordic Deutschland GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_hns(text):
    """HNS Nutzfahrzeuge Service"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnung\s*:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ°
    truck_match = re.search(r'Fahrzeug:\s*GR-OO\s*(\d+)', text)
    if truck_match:
        data['truck'] = f"GR-OO{truck_match.group(1)}"
    else:
        truck_match2 = re.search(r'GR-OO\s*(\d+)', text)
        if truck_match2:
            data['truck'] = f"GR-OO{truck_match2.group(1)}"
        else:
            data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ - ÃÂ±ÃÂµÃ‘â‚¬Ã‘â€˜ÃÂ¼ ÃÂ¿ÃÂµÃ‘â‚¬ÃÂ²Ã‘Æ’Ã‘Å½ Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃ‘Æ’
    desc_match = re.search(r'HU\s+Begleitung|HU\s+PrÃƒÂ¼fung', text)
    if desc_match:
        data['name'] = desc_match.group(0)
    else:
        data['name'] = 'Werkstattleistungen'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ° (Endbetrag)
    total_match = re.search(r'Endbetrag\s*:\s*([\d,.]+)\s*Ã¢â€šÂ¬', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'HNS Nutzfahrzeuge Service GmbH'
    data['buyer'] = 'Groo GmbH' if 'Groo GmbH' in text else 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_tip(text):
    """TIP Trailer Services - ÃÂ°Ã‘â‚¬ÃÂµÃÂ½ÃÂ´ÃÂ° ÃÂ¿Ã‘â‚¬ÃÂ¸Ã‘â€ ÃÂµÃÂ¿ÃÂ¾ÃÂ²"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnungsnr\.\s*U71/(\d+)', text)
    if invoice_match:
        data['invoice'] = f"U71/{invoice_match.group(1)}"
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})/(\d{2})/(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ° (ÃÂ¿Ã‘â‚¬ÃÂ¸Ã‘â€ ÃÂµÃÂ¿)
    truck_match = re.search(r'Kennzeichen:\s*([A-Z0-9]+)', text)
    if truck_match:
        data['truck'] = truck_match.group(1)
    else:
        # ÃÂÃÂ»Ã‘Å’Ã‘â€šÃÂµÃ‘â‚¬ÃÂ½ÃÂ°Ã‘â€šÃÂ¸ÃÂ²ÃÂ½Ã‘â€¹ÃÂ¹ ÃÂ¿ÃÂ¾ÃÂ¸Ã‘ÂÃÂº
        truck_match2 = re.search(r'Flotten-Nr\.\s*(\d+)', text)
        if truck_match2:
            data['truck'] = truck_match2.group(1)
        else:
            data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ
    data['name'] = 'ÃÂÃ‘â‚¬ÃÂµÃÂ½ÃÂ´ÃÂ° ÃÂ¿Ã‘â‚¬ÃÂ¸Ã‘â€ ÃÂµÃÂ¿ÃÂ°'
    
    data['amount'] = 1
    
    # ÃÂ¦ÃÂµÃÂ½ÃÂ° ÃÂ·ÃÂ° ÃÂ¼ÃÂµÃ‘ÂÃ‘ÂÃ‘â€ 
    price_match = re.search(r'(\d+,\d+)\s+EUR', text)
    if price_match:
        price_str = price_match.group(1).replace(',', '.')
        data['price'] = float(price_str)
    else:
        data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ°
    total_match = re.search(r'Gesamtbetrag\s+EUR\s+([\d,.]+)', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'TIP Trailer Services Germany GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_euromaster(text):
    """Euromaster - Ã‘Ë†ÃÂ¸ÃÂ½Ã‘â€¹ ÃÂ¸ ÃÂ¾ÃÂ±Ã‘ÂÃÂ»Ã‘Æ’ÃÂ¶ÃÂ¸ÃÂ²ÃÂ°ÃÂ½ÃÂ¸ÃÂµ"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'Datum\s*:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ°
    truck_match = re.search(r'KFZ-KENNZEICHEN:\s*([A-Z]{2}-[A-Z0-9]+)', text)
    if truck_match:
        data['truck'] = truck_match.group(1)
    else:
        data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ - ÃÂ¿ÃÂµÃ‘â‚¬ÃÂ²ÃÂ°Ã‘Â ÃÂ¿ÃÂ¾ÃÂ·ÃÂ¸Ã‘â€ ÃÂ¸Ã‘Â
    desc_match = re.search(r'EUROMASTER\s+\d+\s+[^\n]+', text)
    if desc_match:
        data['name'] = desc_match.group(0)[:50]
    else:
        data['name'] = 'ÃÂ¨ÃÂ¸ÃÂ½Ã‘â€¹ ÃÂ¸ ÃÂ¾ÃÂ±Ã‘ÂÃÂ»Ã‘Æ’ÃÂ¶ÃÂ¸ÃÂ²ÃÂ°ÃÂ½ÃÂ¸ÃÂµ'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ° (Nettowert ÃÂ¸ÃÂ»ÃÂ¸ Bruttowert)
    total_match = re.search(r'Nettowert\s+Bruttowert\s+([\d,.]+)\s+([\d,.]+)', text)
    if total_match:
        netto_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(netto_str)
    else:
        # ÃÂÃÂ»Ã‘Å’Ã‘â€šÃÂµÃ‘â‚¬ÃÂ½ÃÂ°Ã‘â€šÃÂ¸ÃÂ²ÃÂ½Ã‘â€¹ÃÂ¹ ÃÂ¿ÃÂ¾ÃÂ¸Ã‘ÂÃÂº
        total_match2 = re.search(r'([\d,.]+)\s+EUR\s+zum', text)
        if total_match2:
            total_str = total_match2.group(1).replace('.', '').replace(',', '.')
            data['total_price'] = float(total_str)
        else:
            return None
    
    data['seller'] = 'Euromaster GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data

# ===== ÃÅ¸ÃÂ ÃÅ¾Ãâ€ÃÅ¾Ãâ€ºÃâ€“Ãâ€¢ÃÂÃËœÃâ€¢ ÃÂ¤ÃÂ£ÃÂÃÅ¡ÃÂ¦ÃËœÃâ„¢ ÃËœÃâ€”Ãâ€™Ãâ€ºÃâ€¢ÃÂ§Ãâ€¢ÃÂÃËœÃÂ¯ =====


def extract_man(text):
    """MAN Truck & Bus Deutschland"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnungsnummer:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ° - ÃÂ½ÃÂµ ÃÂ²Ã‘ÂÃÂµÃÂ³ÃÂ´ÃÂ° ÃÂµÃ‘ÂÃ‘â€šÃ‘Å’
    data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ
    desc_match = re.search(r'Job\s+\d+:\s+([^\n]+)', text)
    if desc_match:
        data['name'] = desc_match.group(1)[:50]
    else:
        data['name'] = 'ÃÂ ÃÂµÃÂ¼ÃÂ¾ÃÂ½Ã‘â€š MAN'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ° - Ã‘â‚¬ÃÂ°ÃÂ·ÃÂ»ÃÂ¸Ã‘â€¡ÃÂ½Ã‘â€¹ÃÂµ ÃÂ¿ÃÂ°Ã‘â€šÃ‘â€šÃÂµÃ‘â‚¬ÃÂ½Ã‘â€¹
    total_match = re.search(r'NETTO\s+([\d,.]+)\s+EUR', text, re.IGNORECASE)
    if not total_match:
        total_match = re.search(r'Nettobetrag:\s+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'MAN Truck & Bus Deutschland GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_schutt(text):
    """W. SchÃƒÂ¼tt GmbH"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnung\s*:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ°
    truck_match = re.search(r'Kennzeichen\s*:\s*([A-Z]{2}-[A-Z]{2}\s*\d+)', text)
    if truck_match:
        truck_raw = truck_match.group(1)
        data['truck'] = re.sub(r'\s+', '', truck_raw)
    else:
        data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ
    data['name'] = 'ÃÂ ÃÂµÃÂ¼ÃÂ¾ÃÂ½Ã‘â€šÃÂ½Ã‘â€¹ÃÂµ Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃ‘â€¹'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ°
    total_match = re.search(r'Netto\s+([\d,.]+)\s*Ã¢â€šÂ¬', text)
    if not total_match:
        total_match = re.search(r'Gesamt\s+([\d,.]+)\s*Ã¢â€šÂ¬', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'W. SchÃƒÂ¼tt GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_volvo(text):
    """Volvo Group Trucks Service"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnungs-Nr\.\.\s*:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'Auftr\.-Datum\.\.\s*:\s*(\d{2})\.(\d{2})\.(\d{2})', text)
    if date_match:
        year = f"20{date_match.group(3)}"
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{year}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ°
    truck_match = re.search(r'Kennzeichen\.\.\s*:\s*([A-Z]{2}-[A-Z0-9]+)', text)
    if truck_match:
        data['truck'] = truck_match.group(1)
    else:
        data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ
    desc_match = re.search(r'PFX\s+Ersatzteilnummer\s+[^\n]+', text)
    if desc_match:
        data['name'] = 'Ãâ€”ÃÂ°ÃÂ¿Ã‘â€¡ÃÂ°Ã‘ÂÃ‘â€šÃÂ¸ Volvo'
    else:
        data['name'] = 'ÃÂ ÃÂµÃÂ¼ÃÂ¾ÃÂ½Ã‘â€š Volvo'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ°
    total_match = re.search(r'Gesamt\s+EUR\s+([\d,.]+)', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Volvo Group Trucks Service Nord GmbH'
    data['buyer'] = 'Auto Compass GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_sotecs(text):
    """Sotecs GmbH"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnungsnummer:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ
    desc_match = re.search(r'Einbau von\s+([^\n]+)', text)
    if desc_match:
        data['name'] = desc_match.group(1)[:50]
    else:
        data['name'] = 'ÃÂ£Ã‘ÂÃ‘â€šÃÂ°ÃÂ½ÃÂ¾ÃÂ²ÃÂºÃÂ° ÃÂ¾ÃÂ±ÃÂ¾Ã‘â‚¬Ã‘Æ’ÃÂ´ÃÂ¾ÃÂ²ÃÂ°ÃÂ½ÃÂ¸Ã‘Â'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ°
    total_match = re.search(r'Total\s+\(netto\)\s+([\d,.]+)\s+EUR', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Sotecs GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_express(text):
    """Express Service"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnung\s+(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'Datum\s+(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ°
    truck_match = re.search(r'NGZ\s+(\d+)', text)
    if truck_match:
        data['truck'] = f"NGZ{truck_match.group(1)}"
    else:
        data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ
    data['name'] = 'Wartung und Filter'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ°
    total_match = re.search(r'Gesamtbetrag\s+([\d,.]+)\s*Ã¢â€šÂ¬', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Express Service GmbH'
    data['buyer'] = 'UAB Groo Transport'
    data['interne_rechnung'] = ''
    
    return data



def extract_kl(text):
    """K&L Kfz Meisterbetrieb"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnungsnummer:\s*(\d+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'Rechnungsdatum:\s*(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ - ÃÂ¸ÃÂ· Ã‘â€šÃÂ°ÃÂ±ÃÂ»ÃÂ¸Ã‘â€ Ã‘â€¹
    desc_match = re.search(r'MEHRFACHKUPPLUNG|GETRIEBE', text)
    if desc_match:
        data['name'] = desc_match.group(0)
    else:
        data['name'] = 'Ãâ€”ÃÂ°ÃÂ¿Ã‘â€¡ÃÂ°Ã‘ÂÃ‘â€šÃÂ¸ ÃÂ¸ Ã‘â‚¬ÃÂµÃÂ¼ÃÂ¾ÃÂ½Ã‘â€š'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ°
    total_match = re.search(r'Gesamtsumme netto\s+([\d,.]+)\s*EUR', text)
    if not total_match:
        total_match = re.search(r'Total\s+\(netto\)\s+([\d,.]+)', text)
    
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'K&L Kfz Meisterbetrieb GmbH'
    data['buyer'] = 'GROO GmbH'
    data['interne_rechnung'] = ''
    
    return data



def extract_quick(text):
    """Quick Reifendicount"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'Rechnung\s+Nr\.\s*:\s*([A-Z0-9]+)', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ°
    date_match = re.search(r'vom\s+(\d{2})\.(\d{2})\.(\d{2})', text)
    if date_match:
        year = f"20{date_match.group(3)}"
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{year}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ°
    truck_match = re.search(r'GR-(\d+)', text)
    if truck_match:
        data['truck'] = f"GR-{truck_match.group(1)}"
    else:
        data['truck'] = ''
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ
    data['name'] = 'ÃÂ¨ÃÂ¸ÃÂ½Ã‘â€¹'
    
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ°
    total_match = re.search(r'Gesamtbetrag\s+([\d,.]+)\s+EUR', text)
    if total_match:
        total_str = total_match.group(1).replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Quick Reifendicount'
    data['buyer'] = 'Andreas Groo Handel & Transporte'
    data['interne_rechnung'] = ''
    
    return data



def extract_tankpool24(text):
    """Tankpool24 - Ã‘â€šÃÂ¾ÃÂ¿ÃÂ»ÃÂ¸ÃÂ²ÃÂ½Ã‘â€¹ÃÂµ ÃÂºÃÂ°Ã‘â‚¬Ã‘â€šÃ‘â€¹"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ°
    invoice_match = re.search(r'(\d{7})', text)
    if invoice_match:
        data['invoice'] = invoice_match.group(1)
    else:
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ° - ÃÂ¾ÃÂ±Ã‘â€¹Ã‘â€¡ÃÂ½ÃÂ¾ ÃÂ² ÃÂ½ÃÂ°Ã‘â€¡ÃÂ°ÃÂ»ÃÂµ
    date_match = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', text)
    if date_match:
        date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    data['truck'] = ''
    data['name'] = 'ÃÂ¢ÃÂ¾ÃÂ¿ÃÂ»ÃÂ¸ÃÂ²ÃÂ¾'
    data['amount'] = 1
    data['price'] = 0.0
    
    # ÃÅ¾ÃÂ±Ã‘â€°ÃÂ°Ã‘Â Ã‘ÂÃ‘Æ’ÃÂ¼ÃÂ¼ÃÂ°
    total_match = re.search(r'([\d,.]+)\s+EUR', text)
    if total_match:
        total_str = total_match.group(1).replace('.', '').replace(',', '.')
        data['total_price'] = float(total_str)
    else:
        return None
    
    data['seller'] = 'Tankpool24 International GmbH'
    data['buyer'] = 'Groo GmbH'
    data['interne_rechnung'] = ''
    
    return data


# ===== ÃÅ¡Ãâ€ºÃÂÃÂ¡ÃÂ¡ÃËœÃÂ¤ÃËœÃÅ¡ÃÂÃÂ¢ÃÅ¾ÃÂ  =====


SCANIA_INVOICE_RE = re.compile(r"\b(SCH[A-Z]{2}\d{4,})\b", re.IGNORECASE)
SCANIA_FINANCE_INVOICE_RE = re.compile(r"\b(SRD\d{4,})\b", re.IGNORECASE)


def extract_invoice_hint_from_text(*values):
    """Extract a likely invoice number for review/reporting from filenames or text."""
    for value in values:
        source = str(value or "")
        if not source:
            continue

        scania_match = SCANIA_INVOICE_RE.search(source)
        if scania_match:
            return scania_match.group(1).upper()

        scania_finance_match = SCANIA_FINANCE_INVOICE_RE.search(source)
        if scania_finance_match:
            return scania_finance_match.group(1).upper()

        generic_patterns = [
            r"\b(RE\d{4,}-\d+)\b",
            r"\b(U\d{2}/\d{5,})\b",
            r"\b([A-Z]{2,5}\d{5,})\b",
        ]
        for pattern in generic_patterns:
            match = re.search(pattern, source, flags=re.IGNORECASE)
            if match:
                return match.group(1).upper()

    return ""


def parse_euro_amount(value):
    """Convert German formatted currency text to float."""
    if value is None:
        return 0.0
    try:
        return float(str(value).replace('.', '').replace(',', '.'))
    except (TypeError, ValueError):
        return 0.0


def extract_scania_invoice_from_re_nr(text):
    """Extract Scania invoice number with strong priority for the value under RE-NR."""
    lines = [line.strip() for line in str(text or "").splitlines()]

    def is_invoice_token(token):
        candidate = str(token or "").strip().upper().strip(".,;:()[]{}")
        if not candidate:
            return False
        if candidate in {"RE-NR", "RE-NR.", "AUFTRAGS-NR", "AUFTRAGS-NR.", "RE-DATUM"}:
            return False
        has_letters = bool(re.search(r"[A-Z]", candidate))
        has_digits = bool(re.search(r"\d", candidate))
        return has_letters and has_digits

    def clean_invoice_token(token):
        return str(token or "").strip().upper().strip(".,;:()[]{}")

    for index, line in enumerate(lines):
        if "RE-NR." not in line.upper():
            continue

        same_line_suffix = re.split(r"RE-NR\.", line, maxsplit=1, flags=re.IGNORECASE)[-1]
        same_line_tokens = re.findall(r"[A-Z0-9-]+", same_line_suffix.upper())
        for token in same_line_tokens:
            if is_invoice_token(token):
                return clean_invoice_token(token)

        for next_index in range(index + 1, min(index + 4, len(lines))):
            candidate_line = lines[next_index]
            if any(marker in candidate_line.upper() for marker in ("AUFTRAGS-NR.", "RE-DATUM", "KD-NR.")):
                if "RE-NR." not in candidate_line.upper():
                    continue

            tokens = re.findall(r"[A-Z0-9-]+", candidate_line.upper())
            for token in tokens:
                if is_invoice_token(token):
                    return clean_invoice_token(token)

    return ""


def extract_scania_total_net(text, line_items=None):
    """Extract Scania net total with priority for the invoice summary table."""
    source = str(text or "")
    patterns = [
        r"NETTOBETRAG[^\n]*\n\s*EUR\s+([\d.]+,\d{2})",
        r"Nettobetrag[^\n]*\n\s*\*?\d*\s*([\d.]+,\d{2})",
        r"Netto\s+gesamt\s*\(EUR\)\s*:?\s*([\d.]+,\d{2})",
        r"Zwischensumme:\s*([\d.]+,\d{2})",
        r"NETTOBETRAG[^\d]+EUR\s+([\d.]+,\d{2})",
        r"Gesamt[^\d]+([\d.]+,\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, source, flags=re.IGNORECASE)
        if match:
            amount = parse_euro_amount(match.group(1))
            if amount > 0:
                return amount

    if line_items:
        total = sum(float(item.get('total_price') or 0) for item in line_items)
        if total > 0:
            return round(total, 2)

    return 0.0


def extract_scania_line_items(text, filename):
    """Extract one Scania maintenance-contract row per Kennzeichen."""
    source = str(text or "")
    lines = source.splitlines()
    items = []
    seen = set()
    is_contract = bool(re.search(r"Wartungsvertrag|Vertragsnr|Laufzeit|Rate\.", source, re.IGNORECASE))

    for index, line in enumerate(lines):
        if "KENNZEICHEN" not in line.upper():
            continue

        truck = extract_normalized_truck_number(line)
        if not truck or truck in seen:
            continue

        price = 0.0
        for lookback in range(max(0, index - 3), index + 1):
            amounts = re.findall(r"\b([\d.]+,\d{2})\b", lines[lookback])
            if amounts:
                price = parse_euro_amount(amounts[-1])
                if price > 0:
                    break

        seen.add(truck)
        items.append({
            'truck': truck,
            'name': 'Wartungsvertrag maintenance contract' if is_contract else 'Scania service',
            'category': 'Service',
            'total_price': price,
        })

    if len(items) <= 1:
        return []

    total = extract_scania_total_net(source)
    missing_prices = [item for item in items if not item.get('total_price')]
    if missing_prices and total > 0:
        per_item = round(total / len(items), 2)
        for item in missing_prices:
            item['total_price'] = per_item

    return items


def extract_scania_invoice_date(text, invoice_value=''):
    """Extract the Scania invoice date, avoiding vehicle dates from the header."""
    source = str(text or "")
    escaped_invoice = re.escape(str(invoice_value or ""))
    patterns = []
    if escaped_invoice:
        patterns.append(rf"Rechnung\s+\d+\s+{escaped_invoice}\s+(\d{{2}}\.\d{{2}}\.\d{{2,4}})")
    patterns.extend([
        r"RE-DATUM[^\n]*\n[^\n]*(\d{2}\.\d{2}\.\d{2,4})",
        r"R\s*E\s*C\s*H\s*N\s*U\s*N\s*G[^\n]*(\d{2}\.\d{2}\.\d{2,4})",
        r"\bRE-DATUM\b[^\d]*(\d{2}\.\d{2}\.\d{2,4})",
    ])

    for pattern in patterns:
        match = re.search(pattern, source, flags=re.IGNORECASE)
        if match:
            return match.group(1)

    spl_match = re.search(r"SCH_[A-Z]+_(\d{8})", source, flags=re.IGNORECASE)
    if spl_match:
        raw_date = spl_match.group(1)
        return f"{raw_date[6:8]}.{raw_date[4:6]}.{raw_date[0:4]}"

    generic_match = re.search(r"(\d{2})\.(\d{2})\.(\d{2,4})", source)
    if generic_match:
        return generic_match.group(0)

    return ""


def extract_scania(text, filename):
    """Scania - ÃÂ²ÃÂ½ÃÂµÃ‘Ë†ÃÂ½ÃÂ¸ÃÂµ Ã‘ÂÃ‘â€¡ÃÂµÃ‘â€šÃÂ° ÃÂ¾Ã‘â€š Scania"""
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡Ã‘â€˜Ã‘â€šÃÂ° - Ð±ÐµÑ€ÐµÐ¼ Ð¸Ð¼ÐµÐ½Ð½Ð¾ RE-NR., Ð° Ð½Ðµ AUFTRAGS-NR.
    invoice_value = extract_scania_invoice_from_re_nr(text)
    if not invoice_value:
        # Fallback: all Scania invoice prefixes, not the AUFTRAGS-NR.
        invoice_match = SCANIA_INVOICE_RE.search(text)
        if not invoice_match:
            invoice_match = re.search(r'SCH_(SCH[A-Z]{2}\d{4,})_', text, re.IGNORECASE)
        if not invoice_match:
            invoice_match = SCANIA_FINANCE_INVOICE_RE.search(text)
        if invoice_match:
            invoice_value = invoice_match.group(1).upper()

    if invoice_value:
        data['invoice'] = invoice_value
    else:
        return None
    
    date_str = extract_scania_invoice_date(text, invoice_value)
    if date_str:
        parts = date_str.split('.')
        if len(parts) == 3 and len(parts[2]) == 2:
            date_str = f"{parts[0]}.{parts[1]}.20{parts[2]}"
    else:
        return None
    
    data['date'] = date_str
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        data['month'] = date_obj.month
        data['week'] = date_obj.isocalendar()[1]
    except:
        return None
    
    line_items = extract_scania_line_items(text, filename)

    labeled_truck_match = re.search(
        r'(?:AMTL\.?\s*KENNZ|Kennzeichen)\s*:?\s*([A-Z]{2}[-\s]?[A-Z]{2,4}\s*\d{2,4})',
        text,
        re.IGNORECASE,
    )
    if labeled_truck_match:
        data['truck'] = normalize_truck_candidate(labeled_truck_match.group(1))
    else:
        reference_trucks = extract_reference_trucks(text)
        data['truck'] = reference_trucks[0] if reference_trucks else (extract_truck_from_filename(filename) or '')
    
    # ÃÅ¾ÃÂ¿ÃÂ¸Ã‘ÂÃÂ°ÃÂ½ÃÂ¸ÃÂµ
    data['name'] = 'Scania service'
    data['amount'] = 1
    data['price'] = 0.0
    
    total_price = extract_scania_total_net(text, line_items)
    if total_price > 0:
        data['total_price'] = total_price
    else:
        return None

    if line_items:
        data['line_items'] = line_items
        if not data.get('truck'):
            data['truck'] = line_items[0].get('truck', '')
    
    if 'SCANIA FINANCE' in str(text or '').upper():
        data['seller'] = 'Scania Finance Deutschland GmbH'
        data['buyer'] = 'Groo GmbH'
    else:
        data['seller'] = 'Scania Vertrieb und Service GmbH'
        data['buyer'] = 'Auto Compass GmbH'
    
    return data


def extract_universal(text, supplier, filename):
    """
    ÃÂ£ÃÂ½ÃÂ¸ÃÂ²ÃÂµÃ‘â‚¬Ã‘ÂÃÂ°ÃÂ»Ã‘Å’ÃÂ½ÃÂ¾ÃÂµ ÃÂ¸ÃÂ·ÃÂ²ÃÂ»ÃÂµÃ‘â€¡ÃÂµÃÂ½ÃÂ¸ÃÂµ ÃÂ´ÃÂ»Ã‘Â ÃÂ¿ÃÂ¾Ã‘ÂÃ‘â€šÃÂ°ÃÂ²Ã‘â€°ÃÂ¸ÃÂºÃÂ¾ÃÂ² ÃÂ±ÃÂµÃÂ· Ã‘ÂÃÂ¿ÃÂµÃ‘â€ ÃÂ¸ÃÂ°ÃÂ»Ã‘Å’ÃÂ½ÃÂ¾ÃÂ¹ Ã‘â€žÃ‘Æ’ÃÂ½ÃÂºÃ‘â€ ÃÂ¸ÃÂ¸
    """
    data = {}
    
    # ÃÂÃÂ¾ÃÂ¼ÃÂµÃ‘â‚¬ Ã‘ÂÃ‘â€¡ÃÂµÃ‘â€šÃÂ° (Ã‘Æ’ÃÂ½ÃÂ¸ÃÂ²ÃÂµÃ‘â‚¬Ã‘ÂÃÂ°ÃÂ»Ã‘Å’ÃÂ½ÃÂ¾ÃÂµ)
    invoice_patterns = [
        r'Rechnung[s-]?[Nn]r\.?\s*:?\s*(\d+)',
        r'Invoice\s+(?:No\.?|Number)\s*:?\s*(\d+)',
        r'Rg\.?\s*-?\s*Nr\.?\s*:?\s*(\d+)',
        r'Rechnungs-Nr\.:\s*(\d+)',
    ]
    
    for pattern in invoice_patterns:
        match = re.search(pattern, text)
        if match:
            data['invoice'] = match.group(1)
            break
    
    if not data.get('invoice'):
        return None
    
    # Ãâ€ÃÂ°Ã‘â€šÃÂ° (Ã‘Æ’ÃÂ½ÃÂ¸ÃÂ²ÃÂµÃ‘â‚¬Ã‘ÂÃÂ°ÃÂ»Ã‘Å’ÃÂ½ÃÂ¾ÃÂµ)
    date_patterns = [
        r'(\d{2}\.\d{2}\.\d{4})',
        r'(\d{2})-(\d{2})-(\d{4})',
    ]
    
    for pattern in date_patterns:
        match = re.search(pattern, text)
        if match:
            if len(match.groups()) == 1:
                date_str = match.group(1)
            else:
                date_str = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
            
            try:
                date_obj = datetime.strptime(date_str.replace('-', '.'), '%d.%m.%Y')
                data['date'] = date_obj.strftime('%d.%m.%Y')
                data['month'] = date_obj.month
                data['week'] = date_obj.isocalendar()[1]
                break
            except:
                continue
    
    if not data.get('date'):
        return None
    
    # ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ°
    truck_candidates = extract_reference_trucks(text)
    data['truck'] = truck_candidates[0] if truck_candidates else (extract_truck_from_filename(filename) or '')
    
    # ÃÂ¡Ã‘Æ’ÃÂ¼ÃÂ¼ÃÂ° - ÃÅ¸ÃÂ ÃËœÃÅ¾ÃÂ ÃËœÃÂ¢Ãâ€¢ÃÂ¢ NETTO ÃÂ´ÃÂ»Ã‘Â ÃÂ²ÃÂ½ÃÂµÃ‘Ë†ÃÂ½ÃÂ¸Ã‘â€¦ Ã‘ÂÃ‘â€¡ÃÂµÃ‘â€šÃÂ¾ÃÂ²!
    amount_patterns = [
        r'Nettobetrag\s*([\d,\.]+)',              # 1. Nettobetrag (Ãâ€™ÃÂ«ÃÂ¡ÃÂ¨ÃËœÃâ„¢)
        r'Netto\s+([\d,\.]+)',                    # 2. Netto
        r'SUMME\s+(?:NETTO|netto)\s*:?\s*([\d,\.]+)',  # 3. SUMME NETTO
        r'Summe\s+netto\s*:?\s*([\d,\.]+)',       # 4. Summe netto
        r'Gesamtbetrag\s+(?:EUR)?\s*([\d,\.]+)',  # 5. Gesamtbetrag (BRUTTO)
        r'Summe\s+brutto\s*:?\s*([\d,\.]+)',      # 6. Summe brutto
        r'Total\s*:?\s*([\d,\.]+)',               # 7. Total
        r'Gesamt\s*:?\s*([\d,\.]+)',              # 8. Gesamt (ÃÂ¼ÃÂ¾ÃÂ¶ÃÂµÃ‘â€š ÃÂ±Ã‘â€¹Ã‘â€šÃ‘Å’ BRUTTO)
    ]
    
    for pattern in amount_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            amount_str = match.group(1).replace(',', '.').replace(' ', '')
            try:
                data['amount'] = 1
                data['price'] = float(amount_str)
                data['total_price'] = float(amount_str)
                break
            except:
                continue
    
    if not data.get('total_price'):
        return None
    
    data['seller'] = supplier
    data['buyer'] = 'Auto Compass GmbH'
    data['name'] = ''
    
    return data


# ===== ÃÂ ÃÅ¾ÃÂ£ÃÂ¢Ãâ€¢ÃÂ  ÃËœÃâ€”Ãâ€™Ãâ€ºÃâ€¢ÃÂ§Ãâ€¢ÃÂÃËœÃÂ¯ =====

def extract_data_by_supplier(text, supplier, filename):
    """
    Ð Ð¾ÑƒÑ‚ÐµÑ€: Ð²Ñ‹Ð±Ñ€Ð°Ñ‚ÑŒ Ð¿Ñ€Ð°Ð²Ð¸Ð»ÑŒÐ½ÑƒÑŽ Ñ„ÑƒÐ½ÐºÑ†Ð¸ÑŽ Ð¸Ð·Ð²Ð»ÐµÑ‡ÐµÐ½Ð¸Ñ
    v7.1: ÐŸÐ¾Ð´Ð´ÐµÑ€Ð¶ÐºÐ° 16 Ð¿Ð¾ÑÑ‚Ð°Ð²Ñ‰Ð¸ÐºÐ¾Ð²
    """
    # Ð¡Ð¿ÐµÑ†Ð¸Ð°Ð»Ð¸Ð·Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð½Ñ‹Ðµ extractors
    extractors = {
        'DEKRA': lambda t: extract_dekra(t, filename),
        'Auto Compass (Internal)': lambda t: extract_autocompass(t, filename),
        'Scania External': lambda t: extract_scania(t, filename),
        'Scania': lambda t: extract_scania(t, filename),
        'Vital Projekt': extract_vital_projekt,
        'Ferronordic': extract_ferronordic,
        'HNS': extract_hns,
        'TIP': extract_tip,
        'Euromaster': extract_euromaster,
        'MAN': extract_man,
        'SchÃ¼tt': extract_schutt,
        'Volvo': extract_volvo,
        'Sotecs': extract_sotecs,
        'Express': extract_express,
        'K&L': extract_kl,
        'Quick': extract_quick,
        'Tankpool24': extract_tankpool24,
    }
    
    # ÐŸÑ€Ð¾Ð²ÐµÑ€ÑÐµÐ¼ ÑÐ¿ÐµÑ†Ð¸Ð°Ð»Ð¸Ð·Ð¸Ñ€Ð¾Ð²Ð°Ð½Ð½Ñ‹Ðµ
    if supplier in extractors:
        result = extractors[supplier](text)
        if result:
            return result
    
    # Fallback: ÑƒÐ½Ð¸Ð²ÐµÑ€ÑÐ°Ð»ÑŒÐ½Ð¾Ðµ Ð¸Ð·Ð²Ð»ÐµÑ‡ÐµÐ½Ð¸Ðµ
    return extract_universal(text, supplier, filename)


def check_invoice_exists(invoice_number, excel_file):
    """Check whether the invoice already exists in the master Excel."""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= 7:
                continue
            if str(row[7]) == str(invoice_number):
                wb.close()
                for r in ws.iter_rows(min_row=2):
                    if len(r) <= 7:
                        continue
                    if str(r[7].value) == str(invoice_number):
                        date = r[3].value if len(r) > 3 else None
                        wb.close()
                        return True, date
                wb.close()
                return True, None
        
        wb.close()
        return False, None
    except Exception as e:
        print(f"Invoice duplicate check error: {e}")
        return False, None


def parse_invoice_date(date_str):
    """Parse invoice date in the formats used by the current extractors."""
    if not date_str:
        return None

    for fmt in ('%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(str(date_str), fmt)
        except (TypeError, ValueError):
            continue

    return None


def format_invoice_date(date_str):
    """Normalize invoice date for output reports."""
    date_obj = parse_invoice_date(date_str)
    if date_obj:
        return date_obj.strftime('%d/%m/%Y')
    return date_str or ''


def normalize_party_name(value):
    """Keep only the company name without address in seller/buyer fields."""
    return extract_company_name_only(value)


def sanitize_business_parties(data):
    """Return a copy of extracted data with cleaned seller/buyer names."""
    data = dict(data or {})
    if 'seller' in data:
        data['seller'] = normalize_party_name(data.get('seller', ''))
    if 'buyer' in data:
        data['buyer'] = normalize_party_name(data.get('buyer', ''))
    return data


def sanitize_truck_value(value):
    """Keep only the first valid truck number from a mixed string."""
    return extract_normalized_truck_number(value)


def sanitize_name_and_truck(truck_value, name_value):
    """Move truck-like values out of Name and into Truck."""
    truck = sanitize_truck_value(truck_value)
    name = str(name_value or '').strip()
    name_truck = extract_normalized_truck_number(name)

    if name_truck and not truck:
        truck = name_truck

    if name_truck:
        cleaned_name = strip_truck_number_from_text(name)
        cleaned_name = re.sub(
            r"(?i)\b(?:kennzeichen|amtl\.?\s*kennz\.?|amtl|truck|fahrzeug|fahrgestell(?:-?nr)?|vin|fz)\b",
            " ",
            cleaned_name,
        )
        cleaned_name = re.sub(r"[()\\[\\]{}:;/,_-]+", " ", cleaned_name)
        cleaned_name = re.sub(r"\s+", " ", cleaned_name).strip()

        compact = re.sub(r"[^A-Z0-9]+", "", cleaned_name.upper())
        for token in ("KENNZEICHEN", "AMTLKENNZ", "AMTL", "TRUCK", "FAHRZEUG", "FAHRGESTELL", "FAHRGESTELLNR", "VIN", "FZ"):
            compact = compact.replace(token, "")

        if not re.search(r"[A-Za-zÄÖÜäöüА-Яа-я]", cleaned_name):
            cleaned_name = ""
        elif compact == "":
            cleaned_name = ""
        elif re.search(r"\d", cleaned_name) and not re.search(r"[a-zäöüа-я]", cleaned_name) and " " not in cleaned_name and len(cleaned_name) >= 6:
            cleaned_name = ""

        name = cleaned_name

    return truck, name


def sanitize_extracted_data(data):
    """Normalize seller/buyer and ensure truck-like text does not remain in Name."""
    data = sanitize_business_parties(data)
    data = dict(data or {})

    data['truck'], data['name'] = sanitize_name_and_truck(
        data.get('truck', ''),
        data.get('name', ''),
    )

    line_items = data.get('line_items')
    if isinstance(line_items, list) and line_items:
        sanitized_items = []
        for item in line_items:
            item_copy = dict(item or {})
            item_copy['truck'], item_copy['name'] = sanitize_name_and_truck(
                item_copy.get('truck', data.get('truck', '')),
                item_copy.get('name', data.get('name', '')),
            )
            sanitized_items.append(item_copy)
        data['line_items'] = sanitized_items

        if not data.get('truck'):
            data['truck'] = sanitized_items[0].get('truck', '')
        if not data.get('name'):
            data['name'] = sanitized_items[0].get('name', '')

    return data


CATEGORY_OPTIONS = [
    'Rent',
    'Repair',
    'Service',
    'Parts',
    'Tyres',
    'Toll',
    'Fuel',
    'TÜV',
    'Wash',
    'Insurance',
    'Tax',
    'Parking',
    'Fees',
    'Accessories',
    'Other',
]


def infer_category(data, supplier=''):
    """Infer a normalized category from a fixed list of allowed business values."""
    data = data or {}
    raw_category = str(data.get('category', '') or '').strip().lower()
    name = str(data.get('name', '') or '').lower()
    seller_name = str(data.get('seller', '') or supplier or '').lower()
    haystack = f"{raw_category} {name} {seller_name}"

    rules = [
        ('Toll', ('maut', 'toll collect', 'toll', 'road toll')),
        ('TÜV', ('tüv', 'tuv', 'hauptuntersuch', 'hu/au', 'hu au', 'abgasuntersuchung', 'dekra', 'inspection', 'prüf', 'pruef')),
        ('Insurance', ('versicherung', 'insurance', 'vhv', 'allianz')),
        ('Tax', ('kfz steuer', 'steuer', 'tax', 'zollamt', 'finanzamt')),
        ('Parking', ('parking', 'parken', 'parkgebühr', 'parkgebuehr')),
        ('Fees', ('gebühr', 'gebuehr', 'gebührenbescheid', 'gebuehrenbescheid', 'bescheid', 'zulassung', 'registration', 'anmeldung', 'abmeldung')),
        ('Wash', ('wash', 'wäsche', 'waesche', 'truck wash', 'waschanlage')),
        ('Fuel', ('diesel', 'fuel', 'kraftstoff', 'tankpool', 'shell', 'tanken', 'adblue', 'ad-blue')),
        ('Rent', ('rent', 'miete', 'miet', 'rental', 'leasing', 'trailer services', 'tip trailer', 'mietfahrzeug')),
        ('Tyres', ('reifen', 'tire', 'tyre', 'euromaster', 'vergölst', 'vergoelst', 'reifendiscount', 'vulkan', 'wheel')),
        ('Service', ('service', 'wartung', 'maintenance', 'inspektion', 'inspektion', 'wartungsvertrag')),
        ('Accessories', ('amazon', 'arbeitsschutz', 'zubehör', 'zubehoer', 'tool', 'werkzeug', 'equipment', 'berner', 'werkstattbedarf')),
        ('Repair', ('repair', 'repar', 'werkstatt', 'instandsetzung', 'bremsen', 'autoglas', 'glass', 'schaden')),
        ('Parts', ('ersatzteil', 'ersatzteile', 'parts', 'part', 'spare', 'europart', 'winkler', 'besko', 'sotecs', 'filter', 'lager', 'kupplung')),
    ]

    for category_name, keywords in rules:
        if any(keyword in haystack for keyword in keywords):
            return category_name

    if raw_category:
        category_map = {
            'rent': 'Rent',
            'repair': 'Repair',
            'service': 'Service',
            'parts': 'Parts',
            'tyres': 'Tyres',
            'tires': 'Tyres',
            'toll': 'Toll',
            'fuel': 'Fuel',
            'tuv': 'TÜV',
            'tüv': 'TÜV',
            'wash': 'Wash',
            'insurance': 'Insurance',
            'tax': 'Tax',
            'parking': 'Parking',
            'fees': 'Fees',
            'accessories': 'Accessories',
            'other': 'Other',
        }
        if raw_category in category_map:
            return category_map[raw_category]

    return 'Other'


def build_business_rows(data, supplier=''):
    """Build output rows in the final A-K format expected by the user."""
    data = sanitize_extracted_data(data)
    date_str = data.get('date', '')
    date_obj = parse_invoice_date(date_str)

    year = data.get('year')
    month = data.get('month')
    week = data.get('week')

    if date_obj:
        year = year or date_obj.year
        month = month or date_obj.month
        week = week or date_obj.isocalendar()[1]

    base_row = {
        'year': year or '',
        'month': month or '',
        'week': week or '',
        'date_invoice': format_invoice_date(date_str),
        'invoice': data.get('invoice', ''),
        'seller': normalize_party_name(data.get('seller', '') or supplier),
        'buyer': normalize_party_name(data.get('buyer', '')),
    }

    line_items = data.get('line_items')
    if isinstance(line_items, list) and line_items:
        business_rows = []
        for item in line_items:
            row = base_row.copy()
            row['truck'] = sanitize_truck_value(item.get('truck', data.get('truck', '')))
            row['name'] = item.get('name', data.get('name', ''))
            row['total_price'] = item.get('total_price', item.get('price', data.get('total_price', 0)))
            row['category'] = infer_category(
                {
                    'category': item.get('category', ''),
                    'name': row['name'],
                    'seller': row['seller'],
                },
                supplier,
            )
            business_rows.append(row)
        return business_rows

    return [{
        **base_row,
        'truck': sanitize_truck_value(data.get('truck', '')),
        'name': data.get('name', ''),
        'total_price': data.get('total_price', 0),
        'category': infer_category(
            {
                'category': data.get('category', ''),
                'name': data.get('name', ''),
                'seller': data.get('seller', '') or supplier,
            },
            supplier,
        ),
    }]


def build_report_row(filename, status, supplier='', data=None, reason='', duplicate=False, original_date=''):
    """Build a single row for the per-run validation report."""
    data = sanitize_extracted_data(data)
    business_rows = build_business_rows(data, supplier)
    first_business_row = business_rows[0] if business_rows else {}
    invoice_value = data.get('invoice', '') or extract_invoice_hint_from_text(filename)
    return {
        'filename': filename,
        'status': status,
        'reason_code': reason,
        'reason': reason,
        'duplicate': 'yes' if duplicate else 'no',
        'original_date': original_date or '',
        'supplier': normalize_party_name(supplier or data.get('seller', '')),
        'invoice': invoice_value,
        'date': first_business_row.get('date_invoice', format_invoice_date(data.get('date', ''))),
        'truck': first_business_row.get('truck', data.get('truck', '')),
        'work_name': first_business_row.get('name', data.get('name', '')),
        'missing_fields': ','.join(get_missing_field_codes(data, supplier)),
        'total_price': first_business_row.get('total_price', data.get('total_price', 0)),
        'seller': normalize_party_name(data.get('seller', '')),
        'buyer': normalize_party_name(data.get('buyer', '')),
        'year': first_business_row.get('year', ''),
        'month': first_business_row.get('month', ''),
        'week': first_business_row.get('week', ''),
        'category': first_business_row.get('category', ''),
        'line_items': data.get('line_items', []),
        'business_rows': business_rows,
        'extraction_source': data.get('extraction_source', 'python' if data else ''),
        'ai_confidence': data.get('ai_confidence', ''),
        'ai_notes': data.get('ai_notes', ''),
    }


def get_missing_field_codes(data, supplier=''):
    """Return a normalized list of missing key fields for extracted data."""
    data = sanitize_extracted_data(data)
    missing = []

    if not str(data.get('invoice', '') or '').strip():
        missing.append('invoice')
    if not str(data.get('date', '') or '').strip():
        missing.append('date')
    if not str(data.get('truck', '') or '').strip():
        missing.append('truck')

    seller = str(data.get('seller', '') or supplier or '').strip()
    if not seller or seller == 'Unknown':
        missing.append('seller')

    if not str(data.get('buyer', '') or '').strip():
        missing.append('buyer')
    if not str(data.get('name', '') or '').strip():
        missing.append('name')

    try:
        total_price = float(data.get('total_price', 0) or 0)
    except (TypeError, ValueError):
        total_price = 0
    if total_price <= 0:
        missing.append('total_price')

    return missing


def should_try_ai_on_partial(data, supplier=''):
    """Return True when Python extraction is incomplete enough to justify AI fallback."""
    critical_fields = {'invoice', 'date', 'truck', 'seller', 'buyer', 'name', 'total_price'}
    missing_fields = set(get_missing_field_codes(data, supplier))
    return bool(missing_fields & critical_fields)


def merge_extracted_data(base_data, ai_data):
    """Merge Python extraction with AI output, preferring the more complete value."""
    if not base_data:
        return ai_data
    if not ai_data:
        return base_data

    merged = dict(base_data)
    for key, value in ai_data.items():
        if key == 'line_items':
            if value:
                merged[key] = value
            continue
        if key == 'total_price':
            try:
                ai_total = float(value or 0)
            except (TypeError, ValueError):
                ai_total = 0
            try:
                current_total = float(merged.get(key, 0) or 0)
            except (TypeError, ValueError):
                current_total = 0
            if ai_total > 0 and (current_total <= 0 or should_try_ai_on_partial(base_data, base_data.get('seller', ''))):
                merged[key] = ai_total
            continue
        if value in (None, '', [], {}):
            continue
        if key not in merged or merged.get(key) in (None, '', [], {}):
            merged[key] = value
            continue
        if key in {'truck', 'name', 'seller', 'buyer', 'invoice', 'date', 'category'} and should_try_ai_on_partial(base_data, base_data.get('seller', '')):
            merged[key] = value

    merged['extraction_source'] = 'python+ai'
    if ai_data.get('ai_confidence') is not None:
        merged['ai_confidence'] = ai_data.get('ai_confidence')
    if ai_data.get('ai_missing_fields') is not None:
        merged['ai_missing_fields'] = ai_data.get('ai_missing_fields')
    if ai_data.get('ai_notes'):
        merged['ai_notes'] = ai_data.get('ai_notes')

    return sanitize_extracted_data(merged)


def run_ai_fallback(pdf_path, pdf_file, supplier='', text_content='', partial_data=None, reason=''):
    """Try AI extraction for unreadable or incomplete invoices."""
    if not is_ai_available():
        return None, "ai_not_configured"

    ai_data, ai_error = extract_invoice_with_ai(
        pdf_path=pdf_path,
        filename=pdf_file,
        supplier_hint=supplier,
        text_content=text_content,
        partial_data=partial_data,
    )

    if ai_data:
        log_processing(f"AI_SUCCESS: {pdf_file} - supplier_hint={supplier} - reason={reason}")
        return ai_data, None

    log_processing(f"AI_FAIL: {pdf_file} - supplier_hint={supplier} - reason={reason} - {ai_error}")
    return None, ai_error


def summarize_report_rows(report_rows):
    """Aggregate manual and partial processing reasons for the Telegram summary."""
    reason_breakdown = defaultdict(int)
    partial_breakdown = defaultdict(int)
    partial_count = 0
    processed_full = 0
    problem_items = []

    for row in report_rows:
        status = row.get('status', '')
        reason_code = row.get('reason_code', '') or row.get('reason', '')
        missing_fields = [
            item.strip() for item in str(row.get('missing_fields', '') or '').split(',') if item.strip()
        ]

        if status == 'extracted':
            if missing_fields:
                partial_count += 1
                for field_code in missing_fields:
                    partial_breakdown[field_code] += 1
                problem_items.append({
                    'filename': row.get('filename', ''),
                    'invoice': row.get('invoice', '') or extract_invoice_hint_from_text(row.get('filename', '')),
                    'status': 'partial_review',
                    'reason': 'missing_fields',
                    'missing_fields': ','.join(missing_fields),
                })
            else:
                processed_full += 1
            continue

        if status == 'duplicate':
            continue

        if reason_code:
            reason_breakdown[reason_code] += 1
        problem_items.append({
            'filename': row.get('filename', ''),
            'invoice': row.get('invoice', '') or extract_invoice_hint_from_text(row.get('filename', '')),
            'status': status,
            'reason': reason_code or status or 'unknown',
            'missing_fields': '',
        })

    return {
        'processed_full': processed_full,
        'partial': partial_count,
        'partial_breakdown': dict(partial_breakdown),
        'reason_breakdown': dict(reason_breakdown),
        'problem_items': problem_items,
    }


def save_run_report(report_rows):
    """Save a standalone report for the current processing run."""
    try:
        os.makedirs(REPORT_FOLDER, exist_ok=True)
        report_path = os.path.join(
            REPORT_FOLDER,
            f"processing_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice Report"

        headers = [
            'Yaer',
            'Month',
            'Week',
            'Date Invoice',
            'Truck',
            'Name',
            'Total Price',
            'Invoice',
            'Seller',
            'Buyer',
            'Kategorie',
        ]
        ws.append(headers)

        for row in report_rows:
            if row.get('status') != 'extracted':
                continue
            business_rows = row.get('business_rows') or []
            if not business_rows:
                business_rows = [{
                    'year': row.get('year', ''),
                    'month': row.get('month', ''),
                    'week': row.get('week', ''),
                    'date_invoice': row.get('date', ''),
                    'truck': row.get('truck', ''),
                    'name': row.get('work_name', ''),
                    'total_price': row.get('total_price', 0),
                    'invoice': row.get('invoice', ''),
                    'seller': row.get('seller', row.get('supplier', '')),
                    'buyer': row.get('buyer', ''),
                    'category': row.get('category', ''),
                }]

            for business_row in business_rows:
                ws.append([
                    business_row.get('year', ''),
                    business_row.get('month', ''),
                    business_row.get('week', ''),
                    business_row.get('date_invoice', ''),
                    business_row.get('truck', ''),
                    business_row.get('name', ''),
                    business_row.get('total_price', 0),
                    business_row.get('invoice', ''),
                    business_row.get('seller', ''),
                    business_row.get('buyer', ''),
                    business_row.get('category', ''),
                ])

        ws_review = wb.create_sheet("Review Queue")
        review_headers = [
            'Source File',
            'Review Status',
            'Current Status',
            'Reason Code',
            'Reason',
            'Missing Fields',
            'Supplier',
            'Invoice',
            'Date',
            'Truck',
            'Work Name',
            'Total Price',
            'Seller',
            'Buyer',
            'Extraction Source',
            'AI Confidence',
            'AI Notes',
        ]
        ws_review.append(review_headers)

        ws_problem = wb.create_sheet("Problem Invoices")
        problem_headers = [
            'Invoice',
            'Source File',
            'Review Status',
            'Current Status',
            'Reason Code',
            'Missing Fields',
            'Supplier',
            'Date',
            'Truck',
            'Total Price',
            'AI Notes',
        ]
        ws_problem.append(problem_headers)

        ws_validation = wb.create_sheet("Validation")
        validation_headers = [
            'Source File',
            'Status',
            'Reason Code',
            'Reason',
            'Missing Fields',
            'Duplicate In Master',
            'Original Date',
            'Supplier',
            'Invoice',
            'Date',
            'Truck',
            'Work Name',
            'Total Price',
            'Seller',
            'Buyer',
            'Extraction Source',
            'AI Confidence',
            'AI Notes',
        ]
        ws_validation.append(validation_headers)

        for row in report_rows:
            missing_fields = str(row.get('missing_fields', '') or '').strip()
            if row.get('status') != 'extracted' or missing_fields:
                review_status = 'not_processed'
                if row.get('status') == 'duplicate':
                    review_status = 'duplicate'
                elif row.get('status') == 'extracted':
                    review_status = 'partial_review'

                ws_review.append([
                    row.get('filename', ''),
                    review_status,
                    row.get('status', ''),
                    row.get('reason_code', ''),
                    row.get('reason', ''),
                    missing_fields,
                    row.get('supplier', ''),
                    row.get('invoice', ''),
                    row.get('date', ''),
                    row.get('truck', ''),
                    row.get('work_name', ''),
                    row.get('total_price', 0),
                    row.get('seller', ''),
                    row.get('buyer', ''),
                    row.get('extraction_source', ''),
                    row.get('ai_confidence', ''),
                    row.get('ai_notes', ''),
                ])
                ws_problem.append([
                    row.get('invoice', '') or extract_invoice_hint_from_text(row.get('filename', '')),
                    row.get('filename', ''),
                    review_status,
                    row.get('status', ''),
                    row.get('reason_code', ''),
                    missing_fields,
                    row.get('supplier', ''),
                    row.get('date', ''),
                    row.get('truck', ''),
                    row.get('total_price', 0),
                    row.get('ai_notes', ''),
                ])

            ws_validation.append([
                row.get('filename', ''),
                row.get('status', ''),
                row.get('reason_code', ''),
                row.get('reason', ''),
                row.get('missing_fields', ''),
                row.get('duplicate', ''),
                row.get('original_date', ''),
                row.get('supplier', ''),
                row.get('invoice', ''),
                row.get('date', ''),
                row.get('truck', ''),
                row.get('work_name', ''),
                row.get('total_price', 0),
                row.get('seller', ''),
                row.get('buyer', ''),
                row.get('extraction_source', ''),
                row.get('ai_confidence', ''),
                row.get('ai_notes', ''),
            ])

        for sheet in (ws, ws_review, ws_problem, ws_validation):
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    value = '' if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(value))
                sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

        wb.save(report_path)
        wb.close()
        return report_path
    except Exception as e:
        print(f"Error saving run report: {e}")
        return None


def add_to_excel(data, excel_file):
    """Ãâ€ÃÂ¾ÃÂ±ÃÂ°ÃÂ²ÃÂ¸Ã‘â€šÃ‘Å’ ÃÂ´ÃÂ°ÃÂ½ÃÂ½Ã‘â€¹ÃÂµ ÃÂ² Excel"""
    try:
        data = sanitize_extracted_data(data)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # ÃÂÃÂ°ÃÂ¹Ã‘â€šÃÂ¸ ÃÂ¿ÃÂµÃ‘â‚¬ÃÂ²Ã‘Æ’Ã‘Å½ ÃÂ¿Ã‘Æ’Ã‘ÂÃ‘â€šÃ‘Æ’Ã‘Å½ Ã‘ÂÃ‘â€šÃ‘â‚¬ÃÂ¾ÃÂºÃ‘Æ’
        row_num = 2
        while ws[f'A{row_num}'].value is not None:
            row_num += 1

        first_row_num = row_num
        for business_row in build_business_rows(data, data.get('seller', '')):
            ws[f'A{row_num}'] = business_row.get('year', '')
            ws[f'B{row_num}'] = business_row.get('month', '')
            ws[f'C{row_num}'] = business_row.get('week', '')
            ws[f'D{row_num}'] = business_row.get('date_invoice', '')
            ws[f'E{row_num}'] = business_row.get('truck', '')
            ws[f'F{row_num}'] = business_row.get('name', '')
            ws[f'G{row_num}'] = business_row.get('total_price', 0)
            ws[f'H{row_num}'] = business_row.get('invoice', '')
            ws[f'I{row_num}'] = business_row.get('seller', '')
            ws[f'J{row_num}'] = business_row.get('buyer', 'Auto Compass GmbH')
            ws[f'K{row_num}'] = business_row.get('category', '')
            row_num += 1
        
        wb.save(excel_file)
        wb.close()
        
        return first_row_num
        
    except Exception as e:
        print(f"ÃÅ¾Ã‘Ë†ÃÂ¸ÃÂ±ÃÂºÃÂ° ÃÂ´ÃÂ¾ÃÂ±ÃÂ°ÃÂ²ÃÂ»ÃÂµÃÂ½ÃÂ¸Ã‘Â ÃÂ² Excel: {e}")
        return None


# ===== ÃÂ¤ÃÂ£ÃÂÃÅ¡ÃÂ¦ÃËœÃËœ ÃÂ ÃÂÃâ€˜ÃÅ¾ÃÂ¢ÃÂ« ÃÂ¡ ÃÂ¤ÃÂÃâ„¢Ãâ€ºÃÂÃÅ“ÃËœ =====

def rename_and_move_file(source_path, destination_folder, prefix=''):
    """ÃÅ¸ÃÂµÃ‘â‚¬ÃÂµÃÂ¸ÃÂ¼ÃÂµÃÂ½ÃÂ¾ÃÂ²ÃÂ°Ã‘â€šÃ‘Å’ ÃÂ¸ ÃÂ¿ÃÂµÃ‘â‚¬ÃÂµÃÂ¼ÃÂµÃ‘ÂÃ‘â€šÃÂ¸Ã‘â€šÃ‘Å’ Ã‘â€žÃÂ°ÃÂ¹ÃÂ»"""
    try:
        filename = os.path.basename(source_path)
        
        # ÃÂ£ÃÂ±Ã‘â‚¬ÃÂ°Ã‘â€šÃ‘Å’ Ã‘ÂÃ‘â€šÃÂ°Ã‘â‚¬Ã‘â€¹ÃÂµ ÃÂ¿Ã‘â‚¬ÃÂµÃ‘â€žÃÂ¸ÃÂºÃ‘ÂÃ‘â€¹
        clean_filename = filename
        for old_prefix in ['checked_', 'manual_', 'error_', 'duplicate_']:
            if clean_filename.startswith(old_prefix):
                clean_filename = clean_filename[len(old_prefix):]
        
        # Ãâ€ÃÂ¾ÃÂ±ÃÂ°ÃÂ²ÃÂ¸Ã‘â€šÃ‘Å’ ÃÂ½ÃÂ¾ÃÂ²Ã‘â€¹ÃÂ¹ ÃÂ¿Ã‘â‚¬ÃÂµÃ‘â€žÃÂ¸ÃÂºÃ‘Â
        new_filename = f"{prefix}{clean_filename}"
        destination_path = os.path.join(destination_folder, new_filename)
        
        # Ãâ€¢Ã‘ÂÃÂ»ÃÂ¸ Ã‘â€žÃÂ°ÃÂ¹ÃÂ» Ã‘Æ’ÃÂ¶ÃÂµ Ã‘ÂÃ‘Æ’Ã‘â€°ÃÂµÃ‘ÂÃ‘â€šÃÂ²Ã‘Æ’ÃÂµÃ‘â€š, ÃÂ´ÃÂ¾ÃÂ±ÃÂ°ÃÂ²ÃÂ¸Ã‘â€šÃ‘Å’ timestamp
        if os.path.exists(destination_path):
            name, ext = os.path.splitext(new_filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            new_filename = f"{name}_{timestamp}{ext}"
            destination_path = os.path.join(destination_folder, new_filename)
        
        shutil.move(source_path, destination_path)
        return True, new_filename, None
        
    except Exception as e:
        return False, None, str(e)




def get_processed_folder_by_year(date_str):
    """
    ÐžÐ¿Ñ€ÐµÐ´ÐµÐ»Ð¸Ñ‚ÑŒ Ð¿Ð°Ð¿ÐºÑƒ Ð´Ð»Ñ Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚Ð°Ð½Ð½Ñ‹Ñ… Ñ„Ð°Ð¹Ð»Ð¾Ð² Ð¿Ð¾ Ð³Ð¾Ð´Ñƒ Ð¸Ð· Ð´Ð°Ñ‚Ñ‹ ÑÑ‡Ñ‘Ñ‚Ð°
    ÐÐ²Ñ‚Ð¾Ð¼Ð°Ñ‚Ð¸Ñ‡ÐµÑÐºÐ¸ ÑÐ¾Ð·Ð´Ð°Ñ‘Ñ‚ Ð¿Ð°Ð¿ÐºÐ¸ RG 2025, RG 2026, RG 2027 Ð¸ Ñ‚.Ð´.
    """
    try:
        # Ð˜Ð·Ð²Ð»ÐµÐºÐ°ÐµÐ¼ Ð³Ð¾Ð´ Ð¸Ð· Ð´Ð°Ñ‚Ñ‹ (Ñ„Ð¾Ñ€Ð¼Ð°Ñ‚: DD.MM.YYYY)
        if '.' in date_str:
            parts = date_str.split('.')
            if len(parts) == 3:
                year = parts[2]
            else:
                year = str(datetime.now().year)
        else:
            year = str(datetime.now().year)
        
        # Ð¤Ð¾Ñ€Ð¼Ð¸Ñ€ÑƒÐµÐ¼ Ð¿ÑƒÑ‚ÑŒ Ðº Ð¿Ð°Ð¿ÐºÐµ Ð³Ð¾Ð´Ð°
        year_folder = os.path.join(str(DATA_ROOT), f"RG {year} Ersatyteile RepRG")
        
        # Ð¡Ð¾Ð·Ð´Ð°Ñ‘Ð¼ Ð¿Ð°Ð¿ÐºÑƒ ÐµÑÐ»Ð¸ Ð½Ðµ ÑÑƒÑ‰ÐµÑÑ‚Ð²ÑƒÐµÑ‚
        os.makedirs(year_folder, exist_ok=True)
        
        return year_folder
        
    except Exception as e:
        print(f"  âš ï¸ ÐžÑˆÐ¸Ð±ÐºÐ° Ð¾Ð¿Ñ€ÐµÐ´ÐµÐ»ÐµÐ½Ð¸Ñ Ð³Ð¾Ð´Ð°: {e}")
        # Ð’Ð¾Ð·Ð²Ñ€Ð°Ñ‰Ð°ÐµÐ¼ Ð¿Ð°Ð¿ÐºÑƒ Ð¿Ð¾ ÑƒÐ¼Ð¾Ð»Ñ‡Ð°Ð½Ð¸ÑŽ (2025)
        return PROCESSED_FOLDER



def log_processing(message):
    """Ãâ€ºÃÂ¾ÃÂ³ÃÂ¸Ã‘â‚¬ÃÂ¾ÃÂ²ÃÂ°ÃÂ½ÃÂ¸ÃÂµ ÃÂ² Ã‘â€žÃÂ°ÃÂ¹ÃÂ»"""
    try:
        os.makedirs(LOG_FOLDER, exist_ok=True)
        log_file = os.path.join(LOG_FOLDER, f"processing_{datetime.now().strftime('%Y%m%d')}.log")
        
        with open(log_file, 'a', encoding='utf-8') as f:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            f.write(f"[{timestamp}] {message}\n")
    except Exception as e:
        print(f"ÃÅ¾Ã‘Ë†ÃÂ¸ÃÂ±ÃÂºÃÂ° ÃÂ»ÃÂ¾ÃÂ³ÃÂ¸Ã‘â‚¬ÃÂ¾ÃÂ²ÃÂ°ÃÂ½ÃÂ¸Ã‘Â: {e}")


# ===== Ãâ€œÃâ€ºÃÂÃâ€™ÃÂÃÂÃÂ¯ ÃÂ¤ÃÂ£ÃÂÃÅ¡ÃÂ¦ÃËœÃÂ¯ =====

def process_all_pdfs():
    """
    ÃÅ¾ÃÂ±Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃÂ°Ã‘â€šÃ‘Å’ ÃÂ²Ã‘ÂÃÂµ PDF Ã‘â€žÃÂ°ÃÂ¹ÃÂ»Ã‘â€¹ Ã‘Â ÃÂ´ÃÂµÃ‘â€šÃÂ°ÃÂ»Ã‘Å’ÃÂ½Ã‘â€¹ÃÂ¼ÃÂ¸ Telegram Ã‘Æ’ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸Ã‘ÂÃÂ¼ÃÂ¸
    """
    
    print("\nÃ¢â€¢â€Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢â€”")
    print("Ã¢â€¢â€˜      PDF PROCESSOR v6.0 - ÃÂ£Ãâ€ºÃÂ£ÃÂ§ÃÂ¨Ãâ€¢ÃÂÃÂÃÂ«Ãâ€¢ ÃÅ¸ÃÂÃÂ¢ÃÂ¢Ãâ€¢ÃÂ ÃÂÃÂ«               Ã¢â€¢â€˜")
    print("Ã¢â€¢â€˜      ÃËœÃ‘ÂÃÂ¿Ã‘â‚¬ÃÂ°ÃÂ²ÃÂ»ÃÂµÃÂ½ÃÂ¾: Auto Compass + Scania manual Ã‘â€žÃÂ°ÃÂ¹ÃÂ»Ã‘â€¹         Ã¢â€¢â€˜")
    print("Ã¢â€¢Å¡Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â\n")
    
    # ÃÂ¡ÃÂ¾ÃÂ·ÃÂ´ÃÂ°Ã‘â€šÃ‘Å’ ÃÂ¿ÃÂ°ÃÂ¿ÃÂºÃÂ¸ ÃÂµÃ‘ÂÃÂ»ÃÂ¸ ÃÂ½ÃÂµ Ã‘ÂÃ‘Æ’Ã‘â€°ÃÂµÃ‘ÂÃ‘â€šÃÂ²Ã‘Æ’Ã‘Å½Ã‘â€š
    os.makedirs(MANUAL_FOLDER, exist_ok=True)
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)
    os.makedirs(LOG_FOLDER, exist_ok=True)
    
    # ÃÅ¸ÃÂ¾ÃÂ»Ã‘Æ’Ã‘â€¡ÃÂ¸Ã‘â€šÃ‘Å’ Ã‘ÂÃÂ¿ÃÂ¸Ã‘ÂÃÂ¾ÃÂº Ã‘â€žÃÂ°ÃÂ¹ÃÂ»ÃÂ¾ÃÂ²
    try:
        pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    except Exception as e:
        print(f"Ã¢ÂÅ’ ÃÅ¾Ã‘Ë†ÃÂ¸ÃÂ±ÃÂºÃÂ° ÃÂ´ÃÂ¾Ã‘ÂÃ‘â€šÃ‘Æ’ÃÂ¿ÃÂ° ÃÂº ÃÂ¿ÃÂ°ÃÂ¿ÃÂºÃÂµ: {e}")
        return None
    
    if not pdf_files:
        print("Ã°Å¸â€œÂ­ ÃÂÃÂµÃ‘â€š Ã‘â€žÃÂ°ÃÂ¹ÃÂ»ÃÂ¾ÃÂ² ÃÂ´ÃÂ»Ã‘Â ÃÂ¾ÃÂ±Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃÂºÃÂ¸")
        return {'total': 0, 'processed': 0, 'duplicates': 0, 'manual': 0}
    
    total_files = len(pdf_files)
    processed_count = 0
    duplicate_count = 0
    error_count = 0
    
    # Ãâ€ÃÂ»Ã‘Â Ã‘ÂÃÂ²ÃÂ¾ÃÂ´ÃÂºÃÂ¸
    processed_files = []
    duplicate_files = []
    manual_files = []
    
    supplier_stats = defaultdict(int)
    
    log_processing("="*80)
    log_processing(f"ÃÂÃÂÃÂ§ÃÂÃâ€ºÃÅ¾ ÃÅ¾Ãâ€˜ÃÂ ÃÂÃâ€˜ÃÅ¾ÃÂ¢ÃÅ¡ÃËœ v4.0: {total_files} Ã‘â€žÃÂ°ÃÂ¹ÃÂ»ÃÂ¾ÃÂ²")
    log_processing("="*80)
    
    print(f"\nÃ°Å¸â€œÅ  ÃÂÃÂ°ÃÂ¹ÃÂ´ÃÂµÃÂ½ÃÂ¾ Ã‘â€žÃÂ°ÃÂ¹ÃÂ»ÃÂ¾ÃÂ²: {total_files}")
    print("="*80)
    
    # ÃÅ¾ÃÂ±Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃÂºÃÂ° ÃÂºÃÂ°ÃÂ¶ÃÂ´ÃÂ¾ÃÂ³ÃÂ¾ Ã‘â€žÃÂ°ÃÂ¹ÃÂ»ÃÂ°
    for index, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        
        print(f"\n[{index}/{total_files}] {pdf_file}")
        
        # ÃÂ£ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸ÃÂµ ÃÂ¾ ÃÂ½ÃÂ°Ã‘â€¡ÃÂ°ÃÂ»ÃÂµ ÃÂ¾ÃÂ±Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃÂºÃÂ¸
        if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
            telegram.notify_processing_start(pdf_file)
        
        try:
            # ===== ÃÂ¨ÃÂÃâ€œ 1: ÃËœÃÂ·ÃÂ²ÃÂ»ÃÂµÃ‘â€¡Ã‘Å’ Ã‘â€šÃÂµÃÂºÃ‘ÂÃ‘â€š =====
            full_text = ''
            
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for page in pdf.pages:
                        try:
                            page_text = page.extract_text()
                        except (UnicodeDecodeError, UnicodeError):
                            # ÐŸÑ€Ð¾Ð±ÑƒÐµÐ¼ Ð±ÐµÐ· layout
                            try:
                                page_text = page.extract_text(layout=False)
                            except:
                                page_text = None
                        if page_text:
                            full_text += page_text + '\n'
            
            except (UnicodeDecodeError, UnicodeError, Exception) as pdf_error:
                print(f"  âŒ ÐžÑˆÐ¸Ð±ÐºÐ° Ñ‡Ñ‚ÐµÐ½Ð¸Ñ PDF: {pdf_error}")
                success, new_name, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                if success:
                    print(f"  â†’ manual/{new_name}")
                    log_processing(f"ERROR (Ñ‡Ñ‚ÐµÐ½Ð¸Ðµ): {pdf_file} -> {new_name} - {pdf_error}")
                    manual_files.append(pdf_file)
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, f"ÐžÑˆÐ¸Ð±ÐºÐ° Ñ‡Ñ‚ÐµÐ½Ð¸Ñ PDF: {str(pdf_error)[:80]}")
                error_count += 1
                continue

            if not full_text.strip():
                print(f"  Ã¢Å¡Â  PDF ÃÂ¿Ã‘Æ’Ã‘ÂÃ‘â€šÃÂ¾ÃÂ¹")
                success, new_name, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                
                if success:
                    print(f"  Ã¢â€ â€™ manual/{new_name}")
                    log_processing(f"MANUAL (ÃÂ¿Ã‘Æ’Ã‘ÂÃ‘â€šÃÂ¾ÃÂ¹): {pdf_file} -> {new_name}")
                    manual_files.append(pdf_file)
                    
                    # ÃÂ£ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸ÃÂµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, "PDF ÃÂ¿Ã‘Æ’Ã‘ÂÃ‘â€šÃÂ¾ÃÂ¹")
                
                error_count += 1
                continue
            
            # ===== ÃÂ¨ÃÂÃâ€œ 2: ÃÅ¡ÃÂ»ÃÂ°Ã‘ÂÃ‘ÂÃÂ¸Ã‘â€žÃÂ¸Ã‘â€ ÃÂ¸Ã‘â‚¬ÃÂ¾ÃÂ²ÃÂ°Ã‘â€šÃ‘Å’ =====
            supplier = identify_supplier(full_text)
            print(f"  ÃÅ¸ÃÂ¾Ã‘ÂÃ‘â€šÃÂ°ÃÂ²Ã‘â€°ÃÂ¸ÃÂº: {supplier}")
            supplier_stats[supplier] += 1
            
            # ===== ÃÂ¨ÃÂÃâ€œ 3: ÃËœÃÂ·ÃÂ²ÃÂ»ÃÂµÃ‘â€¡Ã‘Å’ ÃÂ´ÃÂ°ÃÂ½ÃÂ½Ã‘â€¹ÃÂµ =====
            data = extract_data_by_supplier(full_text, supplier, pdf_file)
            
            if not data:
                print(f"  Ã¢ÂÅ’ ÃÂÃÂµ Ã‘Æ’ÃÂ´ÃÂ°ÃÂ»ÃÂ¾Ã‘ÂÃ‘Å’ ÃÂ¸ÃÂ·ÃÂ²ÃÂ»ÃÂµÃ‘â€¡Ã‘Å’ ÃÂ´ÃÂ°ÃÂ½ÃÂ½Ã‘â€¹ÃÂµ")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'manual_'
                )
                
                if success:
                    print(f"  Ã¢â€ â€™ manual/{new_name}")
                    log_processing(f"MANUAL ({supplier}): {pdf_file} -> {new_name}")
                    manual_files.append(pdf_file)
                    
                    # ÃÂ£ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸ÃÂµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, "ÃÂÃÂµ Ã‘Æ’ÃÂ´ÃÂ°ÃÂ»ÃÂ¾Ã‘ÂÃ‘Å’ ÃÂ¸ÃÂ·ÃÂ²ÃÂ»ÃÂµÃ‘â€¡Ã‘Å’ ÃÂ´ÃÂ°ÃÂ½ÃÂ½Ã‘â€¹ÃÂµ", supplier)
                
                error_count += 1
                continue
            
            if data.get('truck'):
                print(f"  ÃÅ“ÃÂ°Ã‘Ë†ÃÂ¸ÃÂ½ÃÂ°: {data['truck']}")
            
            print(f"  ÃÂ¡Ã‘â€¡ÃÂµÃ‘â€š: {data.get('invoice', 'N/A')}, Ãâ€ÃÂ°Ã‘â€šÃÂ°: {data.get('date', 'N/A')}, ÃÂ¡Ã‘Æ’ÃÂ¼ÃÂ¼ÃÂ°: {data.get('total_price', 0):.2f} EUR")
            
            # ===== ÃÂ¨ÃÂÃâ€œ 4: ÃÅ¸Ã‘â‚¬ÃÂ¾ÃÂ²ÃÂµÃ‘â‚¬ÃÂ¸Ã‘â€šÃ‘Å’ ÃÂ´Ã‘Æ’ÃÂ±ÃÂ»ÃÂ¸ÃÂºÃÂ°Ã‘â€š =====
            is_duplicate, original_date = check_invoice_exists(data['invoice'], EXCEL_FILE)
            
            if is_duplicate:
                print(f"  Ã¢Å¡Â  Ãâ€ÃÂ£Ãâ€˜Ãâ€ºÃËœÃÅ¡ÃÂÃÂ¢! ÃÂ¡Ã‘â€¡ÃÂµÃ‘â€š {data['invoice']} Ã‘Æ’ÃÂ¶ÃÂµ Ã‘ÂÃ‘Æ’Ã‘â€°ÃÂµÃ‘ÂÃ‘â€šÃÂ²Ã‘Æ’ÃÂµÃ‘â€š")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'duplicate_'
                )
                
                if success:
                    print(f"  Ã¢â€ â€™ processed/{new_name}")
                    log_processing(f"DUPLICATE: {pdf_file} -> {new_name}")
                    duplicate_files.append(pdf_file)
                    
                    # ÃÂ£ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸ÃÂµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_duplicate(pdf_file, data['invoice'], original_date)
                
                duplicate_count += 1
                continue
            
            # ===== ÃÂ¨ÃÂÃâ€œ 5: Ãâ€ÃÂ¾ÃÂ±ÃÂ°ÃÂ²ÃÂ¸Ã‘â€šÃ‘Å’ ÃÂ² Excel =====
            excel_row = add_to_excel(data, EXCEL_FILE)
            
            if excel_row:
                print(f"  Ã¢Å“â€œ Ãâ€ÃÂ¾ÃÂ±ÃÂ°ÃÂ²ÃÂ»ÃÂµÃÂ½ ÃÂ² Excel: Ã‘ÂÃ‘â€šÃ‘â‚¬ÃÂ¾ÃÂºÃÂ° {excel_row}")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, PROCESSED_FOLDER, 'checked_'
                )
                
                if success:
                    print(f"  Ã¢â€ â€™ processed/{new_name}")
                    log_processing(f"SUCCESS: {pdf_file} -> {new_name}, Excel row {excel_row}")
                    processed_files.append(pdf_file)
                    
                    # ÃÂ£ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸ÃÂµ ÃÂ¾ÃÂ± Ã‘Æ’Ã‘ÂÃÂ¿ÃÂµÃ‘â€¦ÃÂµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_success(data, pdf_file, excel_row)
                
                processed_count += 1
            else:
                print(f"  Ã¢ÂÅ’ ÃÅ¾Ã‘Ë†ÃÂ¸ÃÂ±ÃÂºÃÂ° ÃÂ´ÃÂ¾ÃÂ±ÃÂ°ÃÂ²ÃÂ»ÃÂµÃÂ½ÃÂ¸Ã‘Â ÃÂ² Excel")
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'error_'
                )
                
                if success:
                    print(f"  Ã¢â€ â€™ manual/{new_name}")
                    manual_files.append(pdf_file)
                    
                    # ÃÂ£ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸ÃÂµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, "ÃÅ¾Ã‘Ë†ÃÂ¸ÃÂ±ÃÂºÃÂ° ÃÂ´ÃÂ¾ÃÂ±ÃÂ°ÃÂ²ÃÂ»ÃÂµÃÂ½ÃÂ¸Ã‘Â ÃÂ² Excel", supplier)
                
                error_count += 1
                
        except Exception as e:
            print(f"  Ã¢ÂÅ’ ÃÅ¾Ã‘Ë†ÃÂ¸ÃÂ±ÃÂºÃÂ° ÃÂ¾ÃÂ±Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃÂºÃÂ¸: {e}")
            log_processing(f"ERROR: {pdf_file} - {str(e)}")
            
            try:
                success, new_name, _ = rename_and_move_file(
                    pdf_path, MANUAL_FOLDER, 'error_'
                )
                if success:
                    print(f"  Ã¢â€ â€™ manual/{new_name}")
                    manual_files.append(pdf_file)
                    
                    # ÃÂ£ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸ÃÂµ ÃÂ¾ÃÂ± ÃÂ¾Ã‘Ë†ÃÂ¸ÃÂ±ÃÂºÃÂµ
                    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_error(pdf_file, str(e))
            except:
                pass
            
            error_count += 1
    
    # ===== ÃËœÃÂ¢ÃÅ¾Ãâ€œÃËœ =====
    print("\n" + "="*80)
    print("ÃËœÃÂ¢ÃÅ¾Ãâ€œÃËœ ÃÅ¾Ãâ€˜ÃÂ ÃÂÃâ€˜ÃÅ¾ÃÂ¢ÃÅ¡ÃËœ")
    print("="*80)
    print(f"Ãâ€™Ã‘ÂÃÂµÃÂ³ÃÂ¾ Ã‘â€žÃÂ°ÃÂ¹ÃÂ»ÃÂ¾ÃÂ²: {total_files}")
    print(f"Ã¢Å“â€œ ÃÅ¾ÃÂ±Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃÂ°ÃÂ½ÃÂ¾ Ã‘Æ’Ã‘ÂÃÂ¿ÃÂµÃ‘Ë†ÃÂ½ÃÂ¾: {processed_count}")
    print(f"Ã¢Å¡Â  Ãâ€Ã‘Æ’ÃÂ±ÃÂ»ÃÂ¸ÃÂºÃÂ°Ã‘â€šÃÂ¾ÃÂ²: {duplicate_count}")
    print(f"Ã¢ÂÅ’ ÃÂ¢Ã‘â‚¬ÃÂµÃÂ±Ã‘Æ’Ã‘Å½Ã‘â€š Ã‘â‚¬Ã‘Æ’Ã‘â€¡ÃÂ½ÃÂ¾ÃÂ¹ ÃÂ¾ÃÂ±Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃÂºÃÂ¸: {error_count}")
    
    print("\nÃÂ¡Ã‘â€šÃÂ°Ã‘â€šÃÂ¸Ã‘ÂÃ‘â€šÃÂ¸ÃÂºÃÂ° ÃÂ¿ÃÂ¾ ÃÂ¿ÃÂ¾Ã‘ÂÃ‘â€šÃÂ°ÃÂ²Ã‘â€°ÃÂ¸ÃÂºÃÂ°ÃÂ¼:")
    for supplier, count in sorted(supplier_stats.items(), key=lambda x: x[1], reverse=True):
        print(f"  {supplier}: {count}")
    
    try:
        remaining_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
        remaining_count = len(remaining_files)
        if remaining_count > 0:
            print(f"\nÃ¢Å¡Â  ÃÅ¾Ã‘ÂÃ‘â€šÃÂ°ÃÂ»ÃÂ¸Ã‘ÂÃ‘Å’ ÃÂ½ÃÂµÃÂ¾ÃÂ±Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃÂ°ÃÂ½ÃÂ½Ã‘â€¹ÃÂµ Ã‘â€žÃÂ°ÃÂ¹ÃÂ»Ã‘â€¹: {remaining_count}")
    except:
        remaining_count = 0
    
    log_processing("="*80)
    log_processing(f"Ãâ€”ÃÂÃâ€™Ãâ€¢ÃÂ ÃÂ¨Ãâ€¢ÃÂÃÅ¾: ÃÅ¾ÃÂ±Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃÂ°ÃÂ½ÃÂ¾={processed_count}, Ãâ€Ã‘Æ’ÃÂ±ÃÂ»ÃÂ¸ÃÂºÃÂ°Ã‘â€šÃ‘â€¹={duplicate_count}, ÃÂ Ã‘Æ’Ã‘â€¡ÃÂ½Ã‘â€¹ÃÂµ={error_count}")
    log_processing("="*80)
    
    # ÃÂ¤ÃÂ¸ÃÂ½ÃÂ°ÃÂ»Ã‘Å’ÃÂ½ÃÂ°Ã‘Â Ã‘ÂÃÂ²ÃÂ¾ÃÂ´ÃÂºÃÂ° ÃÂ² Telegram
    if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
        summary = {
            'total': total_files,
            'processed': processed_count,
            'duplicates': duplicate_count,
            'manual': error_count,
            'processed_files': processed_files,
            'duplicate_files': duplicate_files,
            'manual_files': manual_files,
        }
        telegram.notify_summary(summary)
        
        # Ãâ€™Ã‘â€¹ÃÂ²ÃÂµÃ‘ÂÃ‘â€šÃÂ¸ Ã‘ÂÃ‘â€šÃÂ°Ã‘â€šÃÂ¸Ã‘ÂÃ‘â€šÃÂ¸ÃÂºÃ‘Æ’ Ã‘Æ’ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸ÃÂ¹
        stats = telegram.get_stats()
        print(f"\nÃ°Å¸â€œÂ± ÃÂ¡Ã‘â€šÃÂ°Ã‘â€šÃÂ¸Ã‘ÂÃ‘â€šÃÂ¸ÃÂºÃÂ° Ã‘Æ’ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸ÃÂ¹:")
        print(f"   ÃÅ¾Ã‘â€šÃÂ¿Ã‘â‚¬ÃÂ°ÃÂ²ÃÂ»ÃÂµÃÂ½ÃÂ¾: {stats['sent']}")
        print(f"   ÃÅ¾Ã‘Ë†ÃÂ¸ÃÂ±ÃÂ¾ÃÂº: {stats['failed']}")
    
    return {
        'total': total_files,
        'processed': processed_count,
        'duplicates': duplicate_count,
        'manual': error_count,
        'remaining': remaining_count,
        'supplier_stats': dict(supplier_stats)
    }


def process_all_pdfs():
    """Override: controlled processing with report-only and production modes."""

    print("\n" + "=" * 80)
    print("PDF PROCESSOR")
    print("=" * 80)
    print(f"Mode: {PROCESSING_MODE}")

    if MOVE_FILES:
        os.makedirs(MANUAL_FOLDER, exist_ok=True)
        os.makedirs(PROCESSED_FOLDER, exist_ok=True)
    os.makedirs(LOG_FOLDER, exist_ok=True)
    os.makedirs(REPORT_FOLDER, exist_ok=True)

    try:
        pdf_files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    except Exception as e:
        print(f"Folder access error: {e}")
        return None

    if not pdf_files:
        print("No PDF files found")
        return {
            'total': 0,
            'processed': 0,
            'duplicates': 0,
            'manual': 0,
            'remaining': 0,
            'supplier_stats': {},
            'report_path': None,
            'mode': PROCESSING_MODE,
        }

    total_files = len(pdf_files)
    processed_count = 0
    duplicate_count = 0
    error_count = 0
    ai_assisted_files = set()
    processed_files = []
    duplicate_files = []
    manual_files = []
    report_rows = []
    supplier_stats = defaultdict(int)

    log_processing("=" * 80)
    log_processing(f"START: total_files={total_files}")
    log_processing(f"MODE: {PROCESSING_MODE}")
    log_processing("=" * 80)

    for index, pdf_file in enumerate(pdf_files, 1):
        pdf_path = os.path.join(PDF_FOLDER, pdf_file)
        print(f"\n[{index}/{total_files}] {pdf_file}")

        if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
            telegram.notify_processing_start(pdf_file)

        try:
            full_text = ''
            supplier = 'Unknown'
            data = None

            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for page in pdf.pages:
                        try:
                            page_text = page.extract_text()
                        except (UnicodeDecodeError, UnicodeError):
                            try:
                                page_text = page.extract_text(layout=False)
                            except Exception:
                                page_text = None
                        if page_text:
                            full_text += page_text + '\n'
            except (UnicodeDecodeError, UnicodeError, Exception) as pdf_error:
                reason = str(pdf_error)[:200]
                log_processing(f"READ_ERROR: {pdf_file} - {reason}")
                data, ai_error = run_ai_fallback(
                    pdf_path=pdf_path,
                    pdf_file=pdf_file,
                    supplier=supplier,
                    text_content='',
                    partial_data=None,
                    reason='pdf_read_error',
                )
                if not data:
                    manual_files.append(pdf_file)
                    report_rows.append(build_report_row(pdf_file, 'pdf_read_error', reason='pdf_read_error'))
                    if MOVE_FILES:
                        success, _, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                        if success and TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                            telegram.notify_manual(pdf_file, f"PDF read error: {reason}")
                    error_count += 1
                    continue
                ai_assisted_files.add(pdf_file)
                supplier = data.get('seller', supplier) or supplier

            if data is None and not full_text.strip():
                reason = 'PDF has no extractable text'
                log_processing(f"EMPTY_PDF: {pdf_file}")
                data, ai_error = run_ai_fallback(
                    pdf_path=pdf_path,
                    pdf_file=pdf_file,
                    supplier=supplier,
                    text_content='',
                    partial_data=None,
                    reason='empty_pdf',
                )
                if not data:
                    manual_files.append(pdf_file)
                    report_rows.append(build_report_row(pdf_file, 'empty_pdf', reason='empty_pdf'))
                    if MOVE_FILES:
                        success, _, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                        if success and TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                            telegram.notify_manual(pdf_file, "PDF is empty")
                    error_count += 1
                    continue
                ai_assisted_files.add(pdf_file)
                supplier = data.get('seller', supplier) or supplier

            if data is None:
                supplier = identify_supplier(full_text)
                supplier_stats[supplier] += 1
                data = extract_data_by_supplier(full_text, supplier, pdf_file)

            if not data:
                reason = 'unknown_supplier' if supplier == 'Unknown' else 'extract_failed'
                log_processing(f"NO_DATA: {pdf_file} - supplier={supplier}")
                data, ai_error = run_ai_fallback(
                    pdf_path=pdf_path,
                    pdf_file=pdf_file,
                    supplier=supplier,
                    text_content=full_text,
                    partial_data=None,
                    reason=reason,
                )
                if not data:
                    manual_files.append(pdf_file)
                    report_rows.append(build_report_row(pdf_file, 'no_data', supplier=supplier, reason=reason))
                    if MOVE_FILES:
                        success, _, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'manual_')
                        if success and TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                            telegram.notify_manual(pdf_file, reason, supplier)
                    error_count += 1
                    continue
                ai_assisted_files.add(pdf_file)
                supplier = data.get('seller', supplier) or supplier

            if _env_flag('AI_FALLBACK_ON_PARTIAL', True) and should_try_ai_on_partial(data, supplier):
                ai_data, ai_error = run_ai_fallback(
                    pdf_path=pdf_path,
                    pdf_file=pdf_file,
                    supplier=supplier,
                    text_content=full_text,
                    partial_data=data,
                    reason='partial',
                )
                if ai_data:
                    data = merge_extracted_data(data, ai_data)
                    ai_assisted_files.add(pdf_file)
                    supplier = data.get('seller', supplier) or supplier

            if supplier:
                supplier_stats[supplier] += 0

            invoice_number = data.get('invoice', '')
            is_duplicate = False
            original_date = None
            if invoice_number:
                is_duplicate, original_date = check_invoice_exists(invoice_number, EXCEL_FILE)

            if is_duplicate:
                duplicate_count += 1
                duplicate_files.append(pdf_file)
                log_processing(f"DUPLICATE: {pdf_file} - invoice={invoice_number}")
                report_rows.append(build_report_row(
                    pdf_file,
                    'duplicate',
                    supplier=supplier,
                    data=data,
                    reason='duplicate',
                    duplicate=True,
                    original_date=original_date,
                ))
                if MOVE_FILES:
                    success, _, _ = rename_and_move_file(pdf_path, PROCESSED_FOLDER, 'duplicate_')
                    if success and TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_duplicate(pdf_file, invoice_number, original_date)
                continue

            if WRITE_MASTER_EXCEL:
                excel_row = add_to_excel(data, EXCEL_FILE)
            else:
                excel_row = 1

            if excel_row:
                processed_count += 1
                processed_files.append(pdf_file)
                report_rows.append(build_report_row(
                    pdf_file,
                    'extracted',
                    supplier=supplier,
                    data=data,
                    reason='Duplicate detected in master Excel' if is_duplicate else '',
                    duplicate=is_duplicate,
                    original_date=original_date,
                ))
                if MOVE_FILES:
                    success, new_name, _ = rename_and_move_file(pdf_path, PROCESSED_FOLDER, 'checked_')
                    if success:
                        log_processing(f"SUCCESS: {pdf_file} -> {new_name}, Excel row {excel_row}")
                        if TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                            telegram.notify_success(data, pdf_file, excel_row)
                else:
                    log_processing(f"REPORT_ONLY_SUCCESS: {pdf_file}")
            else:
                reason = 'excel_write_error'
                manual_files.append(pdf_file)
                report_rows.append(build_report_row(
                    pdf_file,
                    'excel_error',
                    supplier=supplier,
                    data=data,
                    reason=reason,
                ))
                if MOVE_FILES:
                    success, _, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'error_')
                    if success and TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_manual(pdf_file, reason, supplier)
                error_count += 1

        except Exception as e:
            reason = str(e)[:200]
            log_processing(f"PROCESSING_ERROR: {pdf_file} - {reason}")
            manual_files.append(pdf_file)
            report_rows.append(build_report_row(pdf_file, 'processing_error', reason='processing_error'))
            if MOVE_FILES:
                try:
                    success, _, _ = rename_and_move_file(pdf_path, MANUAL_FOLDER, 'error_')
                    if success and TELEGRAM_ENABLED and DETAILED_NOTIFICATIONS:
                        telegram.notify_error(pdf_file, reason)
                except Exception:
                    pass
            error_count += 1

    try:
        remaining_count = len([f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')])
    except Exception:
        remaining_count = 0

    report_path = save_run_report(report_rows)
    if report_path:
        print(f"Report saved: {report_path}")
        log_processing(f"REPORT: {report_path}")

    print("\n" + "=" * 80)
    print(f"Total files: {total_files}")
    print(f"Processed: {processed_count}")
    print(f"Duplicates: {duplicate_count}")
    print(f"Manual/Error: {error_count}")
    ai_count = len(ai_assisted_files)
    print(f"AI Assisted: {ai_count}")

    log_processing("=" * 80)
    log_processing(
        f"FINISH: processed={processed_count}, duplicates={duplicate_count}, manual={error_count}, ai={ai_count}, remaining={remaining_count}"
    )
    log_processing("=" * 80)

    if TELEGRAM_ENABLED:
        report_summary = summarize_report_rows(report_rows)
        summary = {
            'total': total_files,
            'processed': processed_count,
            'processed_full': report_summary['processed_full'],
            'partial': report_summary['partial'],
            'duplicates': duplicate_count,
            'manual': error_count,
            'ai': ai_count,
            'reason_breakdown': report_summary['reason_breakdown'],
            'partial_breakdown': report_summary['partial_breakdown'],
            'problem_items': report_summary['problem_items'],
            'report_path': report_path,
            'mode': PROCESSING_MODE,
        }
        telegram.notify_summary(summary)

    return {
        'total': total_files,
        'processed': processed_count,
        'duplicates': duplicate_count,
        'manual': error_count,
        'ai': ai_count,
        'remaining': remaining_count,
        'supplier_stats': dict(supplier_stats),
        'report_path': report_path,
        'mode': PROCESSING_MODE,
    }


# ===== ÃÂ¢ÃÅ¾ÃÂ§ÃÅ¡ÃÂ Ãâ€™ÃÂ¥ÃÅ¾Ãâ€ÃÂ =====

if False and __name__ == "__main__":
    try:
        print("\n")
        print("Ã¢â€¢â€Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢â€”")
        print("Ã¢â€¢â€˜         PDF PROCESSOR v6.0 - ÃÂ£Ãâ€ºÃÂ£ÃÂ§ÃÂ¨Ãâ€¢ÃÂÃÂÃÂ«Ãâ€¢ ÃÅ¸ÃÂÃÂ¢ÃÂ¢Ãâ€¢ÃÂ ÃÂÃÂ«            Ã¢â€¢â€˜")
        print("Ã¢â€¢â€˜         ÃËœÃ‘ÂÃÂ¿Ã‘â‚¬ÃÂ°ÃÂ²ÃÂ»ÃÂµÃÂ½ÃÂ¾: Auto Compass + Scania                   Ã¢â€¢â€˜")
        print("Ã¢â€¢Å¡Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â")
        print("\n")
        
        print("ÃÅ¸Ã‘â‚¬ÃÂ¾ÃÂ²ÃÂµÃ‘â‚¬ÃÂºÃÂ° Ã‘ÂÃÂ¸Ã‘ÂÃ‘â€šÃÂµÃÂ¼Ã‘â€¹...")
        print(f"Ã¢Å“â€œ PDF ÃÂ¿ÃÂ°ÃÂ¿ÃÂºÃÂ°: {os.path.exists(PDF_FOLDER)}")
        print(f"Ã¢Å“â€œ Excel Ã‘â€žÃÂ°ÃÂ¹ÃÂ»: {os.path.exists(EXCEL_FILE)}")
        print(f"Ã¢Å“â€œ Telegram: {TELEGRAM_ENABLED}")
        print(f"Ã¢Å“â€œ Ãâ€ÃÂµÃ‘â€šÃÂ°ÃÂ»Ã‘Å’ÃÂ½Ã‘â€¹ÃÂµ Ã‘Æ’ÃÂ²ÃÂµÃÂ´ÃÂ¾ÃÂ¼ÃÂ»ÃÂµÃÂ½ÃÂ¸Ã‘Â: {DETAILED_NOTIFICATIONS}")
        print(f"Ã¢Å“â€œ Mode: {PROCESSING_MODE}")
        
        if not os.path.exists(PDF_FOLDER):
            print("\nÃ¢ÂÅ’ ÃÅ¸ÃÂ°ÃÂ¿ÃÂºÃÂ° PDF ÃÂ½ÃÂµ ÃÂ½ÃÂ°ÃÂ¹ÃÂ´ÃÂµÃÂ½ÃÂ°!")
            exit(1)
        
        if not os.path.exists(EXCEL_FILE):
            print("\nÃ¢ÂÅ’ Excel Ã‘â€žÃÂ°ÃÂ¹ÃÂ» ÃÂ½ÃÂµ ÃÂ½ÃÂ°ÃÂ¹ÃÂ´ÃÂµÃÂ½!")
            exit(1)
        
        print("\nÃâ€”ÃÂ°ÃÂ¿Ã‘Æ’Ã‘ÂÃÂº ÃÂ¾ÃÂ±Ã‘â‚¬ÃÂ°ÃÂ±ÃÂ¾Ã‘â€šÃÂºÃÂ¸...")
        results = process_all_pdfs()
        
        print("\nÃ¢Å“â€¦ ÃÅ¸Ã‘â‚¬ÃÂ¾ÃÂ³Ã‘â‚¬ÃÂ°ÃÂ¼ÃÂ¼ÃÂ° ÃÂ·ÃÂ°ÃÂ²ÃÂµÃ‘â‚¬Ã‘Ë†ÃÂµÃÂ½ÃÂ° Ã‘Æ’Ã‘ÂÃÂ¿ÃÂµÃ‘Ë†ÃÂ½ÃÂ¾!")
        
        if results and results['total'] > 0:
            success_rate = (results['processed'] / results['total']) * 100
            print(f"\nÃ°Å¸â€œÅ  ÃÅ¸Ã‘â‚¬ÃÂ¾Ã‘â€ ÃÂµÃÂ½Ã‘â€š ÃÂ°ÃÂ²Ã‘â€šÃÂ¾ÃÂ¼ÃÂ°Ã‘â€šÃÂ¸ÃÂ·ÃÂ°Ã‘â€ ÃÂ¸ÃÂ¸: {success_rate:.1f}%")
        
    except Exception as e:
        print(f"\nÃ¢ÂÅ’ ÃÅ¡Ã‘â‚¬ÃÂ¸Ã‘â€šÃÂ¸Ã‘â€¡ÃÂµÃ‘ÂÃÂºÃÂ°Ã‘Â ÃÂ¾Ã‘Ë†ÃÂ¸ÃÂ±ÃÂºÃÂ°: {e}")
        log_processing(f"CRITICAL ERROR: {str(e)}")
        
        if TELEGRAM_ENABLED:
            try:
                telegram.notify_error("System", str(e))
            except:
                pass
        
        exit(1)


if __name__ == "__main__":
    try:
        print()
        print("=" * 80)
        print("PROCESS PDF")
        print("=" * 80)
        print(f"PDF folder exists: {os.path.exists(PDF_FOLDER)}")
        print(f"Excel file exists: {os.path.exists(EXCEL_FILE)}")
        print(f"Telegram enabled: {TELEGRAM_ENABLED}")
        print(f"Detailed notifications: {DETAILED_NOTIFICATIONS}")
        print(f"Mode: {PROCESSING_MODE}")

        if not os.path.exists(PDF_FOLDER):
            print("\nPDF folder not found")
            exit(1)

        if not os.path.exists(EXCEL_FILE):
            print("\nExcel file not found")
            exit(1)

        print("\nStarting processing...")
        results = process_all_pdfs()

        print("\nProcessing finished")
        if results and results.get('total', 0) > 0:
            success_rate = (results['processed'] / results['total']) * 100
            print(f"Automation rate: {success_rate:.1f}%")
            if results.get('report_path'):
                print(f"Report: {results['report_path']}")

    except Exception as e:
        print(f"\nCritical error: {e}")
        log_processing(f"CRITICAL ERROR: {str(e)}")

        if TELEGRAM_ENABLED:
            try:
                telegram.notify_error("System", str(e))
            except Exception:
                pass

        exit(1)
